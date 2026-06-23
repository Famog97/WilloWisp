"""
iscs_reports.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Standalone Reporting Module for the ISCS Closed-Loop Framework.
Generates consolidated HTML Suite Dashboards and structured Excel workbooks.
"""

import os
import json
import re
import datetime
import logging
import traceback
from pathlib import Path

logger = logging.getLogger("AutoClick")

# Try to load optional Excel-related dependencies
try:
    import pandas as pd
    import openpyxl
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


def _size_label(path: Path) -> str:
    """Helper to calculate human-readable file sizes."""
    try:
        b = path.stat().st_size
        if b < 1024:           return f"{b} B"
        if b < 1024 ** 2:      return f"{b/1024:.1f} KB"
        if b < 1024 ** 3:      return f"{b/1024**2:.1f} MB"
        return f"{b/1024**3:.2f} GB"
    except Exception:
        return ""


class ReportManager:
    @staticmethod
    def normalize_results(raw_results):
        """
        Consolidates results across loops, scenarios, and rerun attempts.
        Preserves loop numbering, scenario mappings, and granular steps.
        """
        def determine_sub_status(msg, item_overall):
            if not msg:
                return "SKIP"
            msg_lower = msg.lower()
            if "skip" in msg_lower or "skipped" in msg_lower or "not drawn" in msg_lower or "not configured" in msg_lower:
                return "SKIP"
            if re.search(r'\d{2}:\d{2}:\d{2}', msg):
                return "PASS"
            if any(w in msg_lower for w in ["not found", "not seen", "never seen", "exceeds", "mismatch", "missing", "failed", "not detected", "fail"]):
                return "FAIL"
            if any(w in msg_lower for w in ["found", "seen", "within", "pass", "detected"]):
                return "PASS"
            return "FAIL" if item_overall == "FAIL" else "PASS"

        # Sort raw results by rerun_attempt so later runs naturally overwrite earlier ones in our map
        sorted_raw = sorted(raw_results, key=lambda x: x.get("rerun_attempt", 0))
        by_point = {}

        for item in sorted_raw:
            pid = item.get("point_id", item.get("identifier", ""))
            if not pid:
                continue

            loop_num = item.get("loop_num", 1)
            sc_idx = item.get("scenario_idx", 1)
            sc_name = item.get("scenario_name", "Scenario")
            key = (loop_num, sc_idx, pid)

            overall = item.get("overall", "FAIL")
            screenshot = item.get("screenshot", "")
            attempt = item.get("rerun_attempt", 0)

            # Build structural steps
            steps = []
            sections = {
                "alarm_panel":  {"label": "🚨 Alarm Panel",  "trigger": [], "normalize": [], "overall": "PASS"},
                "alarm_list":   {"label": "📋 Alarm List",   "trigger": [], "normalize": [], "overall": "SKIP"},
                "event_list":   {"label": "📅 Event List",   "trigger": [], "normalize": [], "overall": "SKIP"},
                "equipment":    {"label": "🖥 Equipment Page","trigger": [], "normalize": [], "overall": "SKIP"},
            }

            field_map = {
                "trigger_datetime":    ("alarm_panel", "trigger", "Datetime"),
                "trigger_identifier":  ("alarm_panel", "trigger", "Identifier"),
                "trigger_description": ("alarm_panel", "trigger", "Description"),
                "trigger_value":       ("alarm_panel", "trigger", "Value"),
                "trigger_severity":    ("alarm_panel", "trigger", "Severity"),
                "trigger_color":       ("alarm_panel", "trigger", "Color"),
                "norm_datetime":       ("alarm_panel", "normalize", "Datetime"),
                "norm_identifier":     ("alarm_panel", "normalize", "Identifier"),
                "norm_value":          ("alarm_panel", "normalize", "Value"),
                "norm_severity":       ("alarm_panel", "normalize", "Severity"),
                "norm_color":          ("alarm_panel", "normalize", "Color"),
                "al_trigger_identifier": ("alarm_list", "trigger", "Identifier"),
                "al_trigger_value":      ("alarm_list", "trigger", "Value"),
                "al_trigger_severity":   ("alarm_list", "trigger", "Severity"),
                "al_trigger_color":      ("alarm_list", "trigger", "Color"),
                "al_norm_value":         ("alarm_list", "normalize", "Value"),
                "al_norm_color":         ("alarm_list", "normalize", "Color"),
                "ev_trigger_identifier": ("event_list", "trigger", "Identifier"),
                "ev_trigger_value":      ("event_list", "trigger", "Value"),
                "ev_trigger_severity":   ("event_list", "trigger", "Severity"),
                "ev_trigger_color":      ("event_list", "trigger", "Color"),
                "ev_norm_value":         ("event_list", "normalize", "Value"),
                "ev_norm_color":         ("event_list", "normalize", "Color"),
                "eq_detail":             ("equipment", "trigger", "Detail"),
            }

            # If steps structure exists (Workflow Engine format), parse and map to sections
            if "steps" in item:
                for rs in item["steps"]:
                    step_name = rs.get("step", "").replace("/", " - ").replace("_", " ").title()
                    status = rs.get("status", "FAIL") if rs.get("status") in ["PASS", "FAIL", "SKIP"] else determine_sub_status(rs.get("msg", ""), overall)
                    msg = rs.get("msg", "")
                    
                    step_obj = {"name": step_name, "status": status, "message": msg}
                    steps.append(step_obj)
                    
                    # Distribute steps into sections for UI rendering
                    step_lower = step_name.lower()
                    if "panel" in step_lower or "trigger" in step_lower or "normalize" in step_lower:
                        if "normalize" in step_lower or "reset" in step_lower:
                            sections["alarm_panel"]["normalize"].append(step_obj)
                        else:
                            sections["alarm_panel"]["trigger"].append(step_obj)
                    elif "alarm list" in step_lower:
                        sections["alarm_list"]["trigger"].append(step_obj)
                    elif "event list" in step_lower:
                        sections["event_list"]["trigger"].append(step_obj)
                    elif "equipment" in step_lower or "inspector" in step_lower:
                        sections["equipment"]["trigger"].append(step_obj)

                def _evaluate_section_overall(sec_key):
                    sec = sections[sec_key]
                    all_checks = sec["trigger"] + sec["normalize"]
                    if not all_checks:
                        return "SKIP"
                    if any(c["status"] == "FAIL" for c in all_checks):
                        return "FAIL"
                    return "PASS"

                sections["alarm_panel"]["overall"] = _evaluate_section_overall("alarm_panel")
                sections["alarm_list"]["overall"]  = _evaluate_section_overall("alarm_list")
                sections["event_list"]["overall"]  = _evaluate_section_overall("event_list")
                sections["equipment"]["overall"]   = _evaluate_section_overall("equipment")
            else:
                for k, v in item.items():
                    if k in ("point_id", "overall", "screenshot", "failure_diagnostics", "rerun_attempt") or not v:
                        continue
                    mapping = field_map.get(k)
                    if mapping is None:
                        continue
                    sec, phase, name = mapping
                    status = determine_sub_status(str(v), overall)
                    sections[sec][phase].append({"name": name, "status": status, "message": str(v)})

                def _sec_overall(prefix):
                    val = item.get(f"{prefix}_overall", "")
                    return val if val in ("PASS", "FAIL", "SKIP") else "SKIP"

                sections["alarm_panel"]["overall"] = _sec_overall("trigger") if item.get("trigger_overall") else "PASS"
                sections["alarm_list"]["overall"]  = item.get("al_trigger_overall", "SKIP")
                sections["event_list"]["overall"]  = item.get("ev_trigger_overall", "SKIP")
                sections["equipment"]["overall"]   = item.get("eq_overall", "SKIP")

                for sec_data in sections.values():
                    for step in sec_data["trigger"] + sec_data["normalize"]:
                        steps.append(step)

            # ── Asset-bound custom verify checks (VERIFY_CUSTOM) ───────────────
            custom_checks = item.get("custom_checks", []) or []
            for cc in custom_checks:
                step_obj = {
                    "name":       cc.get("name", "Custom Check"),
                    "status":     cc.get("status", "FAIL"),
                    "message":    cc.get("message", ""),
                    "expected":   cc.get("expected", ""),
                    "actual":     cc.get("actual", ""),
                    "asset_name": cc.get("asset_name", ""),
                    "asset_id":   cc.get("asset_id", ""),
                    "is_custom":  True,
                }
                steps.append(step_obj)

            category = "None"
            fail_reason = ""
            if overall == "FAIL":
                fail_reason = "One or more checks failed"
                for s in steps:
                    if s["status"] == "FAIL":
                        fail_reason = f"{s['name']}: {s['message']}"
                        break
                category = "Other"
                lower_reason = fail_reason.lower()
                if any(s.get("is_custom") and s["status"] == "FAIL" for s in steps):
                    category = "Custom Asset Mismatch"
                elif "color" in lower_reason:         category = "Color Miss"
                elif "timeout" in lower_reason:     category = "Timeout"
                elif "datetime" in lower_reason or "timestamp" in lower_reason: category = "Datetime Out of Bounds"
                elif any(w in lower_reason for w in ["identifier", "description", "value", "severity", "text"]): category = "OCR Mismatch"

            # Per-attempt record — preserves THIS attempt's own data so the
            # report can show the full failure→pass journey for the point.
            attempt_rec = {
                "attempt": attempt,
                "overall": overall,
                "screenshot": screenshot,
                "failure_category": category,
                "failure_reason": fail_reason,
                "steps": steps,
                "failure_diagnostics": item.get("failure_diagnostics", None),
                "trigger_info": item.get("trigger_info", None),
            }

            if key not in by_point:
                by_point[key] = {
                    "id": pid,
                    "overall": overall,
                    "loop_num": loop_num,
                    "scenario_idx": sc_idx,
                    "scenario_name": sc_name,
                    "screenshot": screenshot,
                    "failure_category": category,
                    "failure_reason": fail_reason,
                    "steps": steps,
                    "sections": list(sections.values()) if "sections" in locals() else [],
                    "failure_diagnostics": item.get("failure_diagnostics", None),
                    "trigger_info": item.get("trigger_info", None),
                    "rerun_attempt": attempt,
                    "attempts": [attempt_rec],
                }
            else:
                # Later attempt for the same point. Keep every attempt, and make
                # the top-level fields reflect the LATEST attempt (since raw is
                # sorted ascending by rerun_attempt, this item is newer).
                entry = by_point[key]
                entry["attempts"].append(attempt_rec)
                entry["overall"]             = overall
                entry["screenshot"]          = screenshot
                entry["failure_category"]    = category
                entry["failure_reason"]      = fail_reason
                entry["steps"]               = steps
                entry["sections"]            = list(sections.values()) if "sections" in locals() else []
                entry["failure_diagnostics"] = item.get("failure_diagnostics", None)
                entry["trigger_info"]        = item.get("trigger_info", None)
                entry["rerun_attempt"]       = attempt

        return list(by_point.values())

    @classmethod
    def generate_reports(cls, raw_results, output_dir, start_time, end_time, title="Test Run"):
        """Generates both visual HTML Dashboard and structured Excel sheets."""
        normalized = cls.normalize_results(raw_results)
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)

        # 1. Scan for evidence files using recursive directory traversal
        evidence_list = cls._scan_evidence_files(output_path)

        # 2. Generate HTML Report
        cls._write_html_report(normalized, output_path, start_time, end_time, title, evidence_list)

        # 3. Generate Excel Report (Multi-Sheet Workbook)
        cls._write_excel_report(normalized, output_path, start_time, end_time, title)

        # 4. Persist raw results so any report template can be re-rendered later,
        #    offline, with no re-run (FR-30 / FR-30e). Additive — never blocks the run.
        try:
            (output_path / "suite_results.json").write_text(
                json.dumps(raw_results, indent=2, ensure_ascii=False, default=str),
                encoding="utf-8")
        except Exception as ex:
            logger.warning(f"Could not write suite_results.json: {ex}")

    @classmethod
    def on_suite_completed(cls, event) -> None:
        """EventBus handler (P2.3): generate the consolidated report when a suite
        finishes. Subscribed to `SuiteCompleted` at app startup, so the runner no
        longer calls generate_reports directly. Marks the event handled so the
        runner's safety-net fallback knows the report was produced here.

        Reads everything off the event and reproduces the previous behavior,
        including the success/failure messages to the UI log.
        """
        results    = getattr(event, "results", None) or []
        output_dir = getattr(event, "output_dir", None)
        on_log     = getattr(event, "on_log", None)
        if not results or output_dir is None:
            return
        # Claim handling up-front so a generation failure still suppresses the
        # fallback (the failure is logged here, matching old behavior — we don't
        # want a second attempt).
        event.report_generated = True
        try:
            cls.generate_reports(
                results, output_dir,
                getattr(event, "start_time", None), getattr(event, "end_time", None),
                title=getattr(event, "title", None) or "Test Run",
            )
            if callable(on_log):
                name = getattr(output_dir, "name", str(output_dir))
                on_log(f"✅ Consolidated Suite Report generated successfully inside: {name}")
        except Exception as report_ex:
            if callable(on_log):
                on_log(f"⚠ Failed to generate consolidated suite report: {report_ex}")
            logger.error("Suite report compilation error", exc_info=True)

    @classmethod
    def _scan_evidence_files(cls, output_dir: Path) -> list:
        """Recursively scans the directory using os.walk for robust path resolution."""
        categories = {
            "mp4": "video", "avi": "video",
            "png": "screenshot", "jpg": "screenshot", "jpeg": "screenshot",
            "xlsx": "data", "csv": "data", "json": "data",
            "log": "log", "txt": "log"
        }
        files_list = []
        abs_output_dir = output_dir.resolve()
        if not abs_output_dir.exists():
            return files_list

        try:
            for root, dirs, files in os.walk(abs_output_dir):
                for fname in files:
                    if fname == "Suite_Report.html":
                        continue
                    p = Path(root) / fname
                    ext = p.suffix.lstrip(".").lower()
                    if ext not in categories:
                        continue
                    
                    # Normalize both paths to prevent ValueErrors during cross-platform walk
                    abs_p = p.resolve()
                    rel_path = abs_p.relative_to(abs_output_dir).as_posix()
                    
                    files_list.append(cls._build_file_entry(abs_p, fname, ext, categories[ext], rel_path))
        except Exception as e:
            logger.error(f"ReportManager: Failed to recursively scan directory: {e}")
            logger.error(traceback.format_exc())

        order = ["video", "screenshot", "report", "data", "log"]
        files_list.sort(key=lambda f: (order.index(f["category"]) if f["category"] in order else 99, f["name"]))
        return files_list

    @staticmethod
    def _build_file_entry(path: Path, name: str, ext: str, category: str, rel_path: str) -> dict:
        """Constructs safe absolute and relative path references."""
        size_str = _size_label(path)
        return {
            "name":            name,
            "ext":             ext,
            "category":        category,
            "rel_path":        rel_path,
            "full_path":       str(path.resolve()).replace("\\", "/"),
            "folder_path":     str(path.parent.resolve()).replace("\\", "/"),
            "win_full_path":   str(path.resolve()).replace("/", "\\"),
            "win_folder_path": str(path.parent.resolve()).replace("/", "\\"),
            "size_label":      size_str
        }

    @classmethod
    def _write_html_report(cls, results, output_dir, start_time, end_time, title, evidence_list):
        duration = end_time - start_time
        total = len(results)
        passed = sum(1 for r in results if r["overall"] == "PASS")
        failed = total - passed
        pass_rate = (passed / total * 100) if total > 0 else 0
        
        loops_set = {r["loop_num"] for r in results}
        scenarios_set = {r["scenario_name"] for r in results}
        total_loops = len(loops_set)
        total_scenarios = len(scenarios_set)
        
        rerun_points = sum(1 for r in results if r.get("rerun_attempt", 0) > 0)

        # Categorize failure modes
        categories = {"OCR Mismatch": 0, "Color Miss": 0, "Timeout": 0, "Datetime Out of Bounds": 0, "Other": 0}
        for r in results:
            if r["overall"] == "FAIL" and r["failure_category"] in categories:
                categories[r["failure_category"]] += 1

        # Make paths relative for portability
        for r in results:
            if r["screenshot"] and os.path.exists(r["screenshot"]):
                try:
                    r["relative_screenshot"] = os.path.relpath(r["screenshot"], output_dir).replace('\\', '/')
                except ValueError:
                    r["relative_screenshot"] = ""
            else:
                r["relative_screenshot"] = ""

            # Re-base failure-diagnostics crop images relative to the report's
            # directory (output_dir). They were stored relative to session_dir,
            # which can differ from output_dir, breaking the <img src>. Use the
            # absolute path captured at collection time.
            diag = r.get("failure_diagnostics")
            if diag and isinstance(diag, dict):
                cz = diag.get("cropped_zones")
                if isinstance(cz, dict):
                    for _slot, zdata in cz.items():
                        if not isinstance(zdata, dict):
                            continue
                        abs_p = zdata.get("image_abs")
                        if abs_p and os.path.exists(abs_p):
                            try:
                                zdata["image"] = os.path.relpath(abs_p, output_dir).replace('\\', '/')
                            except ValueError:
                                pass

            # Per-attempt path rebasing — same treatment as the top-level point
            # above, so the EXECUTION HISTORY panel's per-attempt screenshot links
            # and cropped-region thumbnails resolve. Additive: only adds
            # relative_screenshot / rebases each attempt's own crop images.
            for att in r.get("attempts", []):
                if att.get("screenshot") and os.path.exists(att["screenshot"]):
                    try:
                        att["relative_screenshot"] = os.path.relpath(att["screenshot"], output_dir).replace('\\', '/')
                    except ValueError:
                        att["relative_screenshot"] = ""
                else:
                    att["relative_screenshot"] = ""

                adiag = att.get("failure_diagnostics")
                if adiag and isinstance(adiag, dict):
                    acz = adiag.get("cropped_zones")
                    if isinstance(acz, dict):
                        for _slot, zdata in acz.items():
                            if not isinstance(zdata, dict):
                                continue
                            abs_p = zdata.get("image_abs")
                            if abs_p and os.path.exists(abs_p):
                                try:
                                    zdata["image"] = os.path.relpath(abs_p, output_dir).replace('\\', '/')
                                except ValueError:
                                    pass

        # Compute file counters safely
        evidence_counts = {"video": 0, "screenshot_pass": 0, "screenshot_fail": 0, "screenshot_other": 0, "data": 0, "log": 0}
        for ev in evidence_list:
            cat = ev["category"]
            if cat == "screenshot":
                name_upper = ev["name"].upper()
                if "_FAIL" in name_upper or "FAIL_" in name_upper:
                    evidence_counts["screenshot_fail"] += 1
                elif "_PASS" in name_upper or "PASS_" in name_upper:
                    evidence_counts["screenshot_pass"] += 1
                else:
                    evidence_counts["screenshot_other"] += 1
            elif cat in evidence_counts:
                evidence_counts[cat] += 1

        results_json = json.dumps(results)
        evidence_json = json.dumps(evidence_list)

        html_template = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>ISCS Suite Consolidated Report - __TITLE__</title>
    <style>
        body {
            background-color: #0d0e12;
            color: #cdd6f4;
            font-family: 'Consolas', monospace;
            margin: 0;
            padding: 20px;
        }
        .container {
            max-width: 1350px;
            margin: 0 auto;
        }
        h1 {
            color: #ffffff;
            border-bottom: 2px solid #2979ff;
            padding-bottom: 10px;
            margin-bottom: 20px;
        }
        .dashboard-row {
            display: flex;
            gap: 20px;
            margin-bottom: 30px;
            flex-wrap: wrap;
        }
        .card {
            background: #16161a;
            border: 1px solid #222225;
            border-radius: 6px;
            padding: 16px;
            flex: 1;
            min-width: 200px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.3);
        }
        .card-title {
            color: #89b4fa;
            font-size: 0.85em;
            text-transform: uppercase;
            margin-bottom: 8px;
        }
        .card-value {
            font-size: 1.8em;
            font-weight: bold;
            color: #fff;
        }
        .chart-container {
            display: flex;
            align-items: center;
            justify-content: space-around;
            min-width: 300px;
        }
        .search-container {
            background: #16161a;
            padding: 14px;
            border-radius: 6px;
            border: 1px solid #222225;
            margin-bottom: 20px;
            display: flex;
            gap: 12px;
            align-items: center;
        }
        .search-input {
            background: #0f0f11;
            color: #ffffff;
            border: 1px solid #313244;
            padding: 8px 14px;
            border-radius: 4px;
            font-family: inherit;
            flex: 1;
            outline: none;
        }
        .search-input:focus {
            border-color: #2979ff;
        }
        .filters {
            margin-bottom: 15px;
            display: flex;
            gap: 10px;
        }
        .btn {
            background: #222;
            color: #aaa;
            border: 1px solid #333;
            padding: 8px 16px;
            border-radius: 4px;
            cursor: pointer;
            font-family: inherit;
            font-weight: bold;
        }
        .btn:hover {
            background: #2d2d35;
        }
        .btn.active {
            background: #2979ff;
            color: #fff;
            border-color: #2979ff;
        }
        
        /* Trees styling */
        .tree-section {
            background: #16161a;
            border: 1px solid #222225;
            border-radius: 6px;
            margin-bottom: 24px;
            overflow: hidden;
        }
        .tree-header {
            background: #222230;
            padding: 14px 20px;
            font-weight: bold;
            color: #cdd6f4;
            border-bottom: 1px solid #222225;
        }
        .tree-body {
            padding: 16px;
        }
        .collapsible-node {
            margin-bottom: 6px;
            border-radius: 4px;
            overflow: hidden;
        }
        .node-trigger {
            background: #1a1b26;
            padding: 10px 14px;
            cursor: pointer;
            user-select: none;
            display: flex;
            justify-content: space-between;
            align-items: center;
            border-left: 3px solid #313244;
            transition: background 0.12s;
        }
        .node-trigger:hover {
            background: #212230;
        }
        .node-trigger.node-fail {
            border-left-color: #f38ba8;
        }
        .node-trigger.node-pass {
            border-left-color: #a6e3a1;
        }
        .node-title {
            font-weight: bold;
            font-size: 0.95em;
            display: flex;
            align-items: center;
            gap: 10px;
        }
        .node-content {
            background: #11121a;
            padding: 14px;
            border-left: 3px solid #313244;
            display: none;
        }
        .node-content.active {
            display: block;
        }
        
        .point-row {
            background: #16161a;
            border: 1px solid #222225;
            border-radius: 4px;
            padding: 10px 14px;
            margin-bottom: 6px;
            cursor: pointer;
            transition: background 0.12s;
        }
        .point-row:hover {
            background: #1e1e2e;
        }
        .badge {
            padding: 2px 7px;
            border-radius: 3px;
            font-size: 0.75em;
            font-weight: bold;
            text-transform: uppercase;
        }
        .badge-pass {
            background: rgba(166, 227, 161, 0.12);
            color: #a6e3a1;
            border: 1px solid #a6e3a1;
        }
        .badge-fail {
            background: rgba(243, 139, 168, 0.12);
            color: #f38ba8;
            border: 1px solid #f38ba8;
        }
        .caret {
            font-size: 0.75em;
            transition: transform 0.15s;
            display: inline-block;
        }
        .rotated {
            transform: rotate(90deg);
        }
        
        /* Copy action buttons */
        .action-btn {
            background: transparent;
            color: #585b70;
            border: 1px solid #313244;
            padding: 2px 8px;
            border-radius: 4px;
            cursor: pointer;
            font-size: 0.75em;
            font-family: inherit;
            transition: all 0.1s;
        }
        .action-btn:hover {
            color: #cdd6f4;
            border-color: #45475a;
            background: #252538;
        }
        
        .details-panel {
            background: #1e1e2e;
            padding: 15px;
            margin-top: 8px;
            border-left: 3px solid #2979ff;
            border-radius: 4px;
            display: none;
        }
        .nested-step {
            margin: 6px 0;
            padding: 4px 8px;
            display: flex;
            justify-content: space-between;
            border-bottom: 1px dashed #313244;
            font-size: 0.85em;
        }
        .custom-check {
            margin: 8px 0;
            padding: 8px 10px;
            border-radius: 4px;
            font-size: 0.85em;
            background: #181825;
        }
        .custom-check-row {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 4px;
        }
        .custom-check-ea {
            display: grid;
            grid-template-columns: 70px 1fr;
            gap: 4px 8px;
            font-size: 0.92em;
            margin-top: 4px;
        }
        .custom-check-ea span:first-child {
            color: #6c7086;
        }
        .custom-check-ea .ea-expected { color: #a6e3a1; }
        .custom-check-ea .ea-actual-pass { color: #a6e3a1; }
        .custom-check-ea .ea-actual-fail { color: #f38ba8; }
        .asset-pill {
            background: #313244; color: #f9c74f;
            font-size: 0.72em; padding: 1px 7px;
            border-radius: 8px; margin-left: 6px;
        }
        .screenshot-thumb {
            width: 140px;
            height: auto;
            border: 1px solid #313244;
            border-radius: 4px;
            cursor: pointer;
            transition: transform 0.15s;
        }
        .screenshot-thumb:hover {
            transform: scale(1.05);
        }
        .modal {
            display: none;
            position: fixed;
            z-index: 1000;
            left: 0;
            top: 0;
            width: 100%;
            height: 100%;
            background-color: rgba(0,0,0,0.92);
            justify-content: center;
            align-items: center;
        }
        .modal-content {
            max-width: 90%;
            max-height: 90%;
            border: 2px solid #313244;
            border-radius: 4px;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>ISCS Automated Suite Execution Report</h1>
        <p style="color: #6c7086; margin-top: -14px;">Suite Title: <strong>__TITLE__</strong></p>

        <!-- Executive Summary Cards -->
        <div class="dashboard-row">
            <div class="card">
                <div class="card-title">Run Information</div>
                <div style="font-size: 0.9em; line-height: 1.6em; margin-top: 8px;">
                    Started: __START_TIME__<br>
                    Finished: __END_TIME__<br>
                    Duration: __DURATION__<br>
                    Loops Run: __TOTAL_LOOPS__ | Scenarios: __TOTAL_SCENARIOS__
                </div>
            </div>
            
            <div class="card" style="text-align: center;">
                <div class="card-title">Consolidated Pass Rate</div>
                <div class="card-value" style="color: __PASS_RATE_COLOR__; margin-top: 5px;">
                    __PASS_RATE__
                </div>
                <div style="font-size: 0.85em; color: #a6adc8; margin-top: 4px;">
                    __PASSED__ / __TOTAL__ Points Passed
                </div>
                <div style="font-size: 0.8em; color: #f9e2af; margin-top: 2px;">
                    __RERUN_POINTS__ Points Passed with Rerun
                </div>
            </div>

            <div class="card chart-container">
                <svg width="110" height="110" viewBox="0 0 42 42" style="transform: rotate(-90deg); border-radius: 50%;">
                    <circle cx="21" cy="21" r="15.915" fill="transparent" stroke="#f38ba8" stroke-width="6"></circle>
                    <circle cx="21" cy="21" r="15.915" fill="transparent" stroke="#a6e3a1" stroke-width="6" 
                            stroke-dasharray="__PASS_RATE_NUM__ __PASS_RATE_REMAIN__" stroke-dashoffset="0"></circle>
                </svg>
                <div style="text-align: left; font-size: 0.85em;">
                    <div style="color: #a6e3a1;">■ Passed: __PASSED__</div>
                    <div style="color: #f38ba8; margin-top: 4px;">■ Failed: __FAILED__</div>
                </div>
            </div>

            <div class="card">
                <div class="card-title">Failure Cause Analysis</div>
                <div style="font-size: 0.85em; line-height: 1.5em; margin-top: 5px;">
                    OCR Mismatch: __OCR_MISMATCH__<br>
                    Color Miss: __COLOR_MISS__<br>
                    Timeout: __TIMEOUT__<br>
                    Datetime Delta: __DATETIME_OUT__<br>
                    Other Exceptions: __OTHER__<br>
                </div>
            </div>
        </div>

        <!-- Filter & Search Controls -->
        <div class="search-container">
            <span>🔎</span>
            <input type="text" class="search-input" id="suite-search" placeholder="Type Loop, Scenario Name, Point ID, or evidence filename to filter results..." onkeyup="filterBySearch()">
        </div>

        <div class="filters">
            <button class="btn active" id="btn-all" onclick="filterResults('all')">All Points (__TOTAL__)</button>
            <button class="btn" id="btn-pass" onclick="filterResults('pass')">Passed (__PASSED__)</button>
            <button class="btn" id="btn-fail" onclick="filterResults('fail')">Failed (__FAILED__)</button>
        </div>

        <!-- Grouped Results Tree -->
        <div class="tree-section">
            <div class="tree-header">
                <span>📂 HIERARCHICAL RUN RESULTS</span>
                <span style="font-size: 0.8em; color: #a6adc8; float: right;">Collapsible loops & cards. Failed assets expand automatically.</span>
            </div>
            <div class="tree-body" id="suite-tree-body"></div>
        </div>

        <!-- Consolidated Evidence Tree -->
        <div class="tree-section">
            <div class="tree-header">
                <span>📁 CONSOLIDATED RUN EVIDENCE</span>
                <span style="font-size: 0.8em; color: #a6adc8; float: right;">Compact stats summary with physical paths copying</span>
            </div>
            <div class="tree-body">
                <div style="background: #11121a; padding: 12px; border-radius: 6px; border: 1px solid #222330; display: flex; gap: 20px; font-size: 0.85em; margin-bottom: 16px;">
                    <div style="color:#f38ba8;">🔴 Failed Crops: __FAIL_SCREENSHOTS__</div>
                    <div style="color:#a6e3a1;">🟢 Passed Crops: __PASS_SCREENSHOTS__</div>
                    <div style="color:#89b4fa;">🎥 Video Recordings: __RECORDINGS__</div>
                    <div style="color:#f9e2af;">📄 Data Sheets / CSVs: __DATA_FILES__</div>
                    <div style="color:#cdd6f4;">Total Files Index: __TOTAL_FILES__</div>
                </div>
                <div id="evidence-tree-body"></div>
            </div>
        </div>
    </div>

    <!-- Image Modal -->
    <div id="image-modal" class="modal" onclick="closeModal()">
        <img class="modal-content" id="modal-img">
    </div>

    <script>
        const rawResults = __RESULTS_JSON__;
        const rawEvidence = __EVIDENCE_JSON__;

        function copyToClipboard(text) {
            navigator.clipboard.writeText(text).then(() => {
                alert("Copied to clipboard: " + text);
            }).catch(err => {
                const ta = document.createElement("textarea");
                ta.value = text;
                document.body.appendChild(ta);
                ta.select();
                document.execCommand("copy");
                document.body.removeChild(ta);
                alert("Copied to clipboard: " + text);
            });
        }

        function toggleNode(element) {
            const trigger = element;
            const content = trigger.nextElementSibling;
            const caret = trigger.querySelector('.caret');
            
            if (content.classList.contains('active')) {
                content.classList.remove('active');
                if (caret) caret.classList.remove('rotated');
            } else {
                content.classList.add('active');
                if (caret) caret.classList.add('rotated');
            }
        }

        function filterBySearch() {
            const query = document.getElementById("suite-search").value.toLowerCase().trim();
            const nodes = document.querySelectorAll(".collapsible-node");
            
            nodes.forEach(node => {
                const text = node.textContent.toLowerCase();
                if (text.includes(query) || query === "") {
                    node.style.display = "block";
                    if (query !== "") {
                        const content = node.querySelector(".node-content");
                        const caret = node.querySelector(".caret");
                        if (content) content.classList.add("active");
                        if (caret) caret.classList.add("rotated");
                    }
                } else {
                    node.style.display = "none";
                }
            });
        }

        function filterResults(mode) {
            document.querySelectorAll(".filters .btn").forEach(btn => btn.classList.remove("active"));
            document.getElementById(`btn-${mode}`).classList.add("active");
            renderTree(mode);
        }

        function toggleDetails(elementId) {
            const details = document.getElementById(elementId);
            if (details.style.display === "block") {
                details.style.display = "none";
            } else {
                details.style.display = "block";
            }
        }

        function openModal(src, event) {
            event.stopPropagation();
            const modal = document.getElementById("image-modal");
            const modalImg = document.getElementById("modal-img");
            modal.style.display = "flex";
            modalImg.src = src;
        }

        function closeModal() {
            document.getElementById("image-modal").style.display = "none";
        }

        function renderTree(filterMode = "all") {
            const treeBody = document.getElementById("suite-tree-body");
            treeBody.innerHTML = "";

            const loops = {};
            rawResults.forEach(r => {
                if (filterMode === "pass" && r.overall !== "PASS") return;
                if (filterMode === "fail" && r.overall !== "FAIL") return;

                const lNum = r.loop_num;
                if (!loops[lNum]) {
                    loops[lNum] = { hasFailure: false, scenarios: {} };
                }
                if (r.overall === "FAIL") {
                    loops[lNum].hasFailure = true;
                }

                const sName = r.scenario_name;
                if (!loops[lNum].scenarios[sName]) {
                    loops[lNum].scenarios[sName] = { hasFailure: false, points: [] };
                }
                if (r.overall === "FAIL") {
                    loops[lNum].scenarios[sName].hasFailure = true;
                }

                loops[lNum].scenarios[sName].points.push(r);
            });

            for (const [lNum, lData] of Object.entries(loops)) {
                const loopNode = document.createElement("div");
                loopNode.className = "collapsible-node";

                const loopTrigger = document.createElement("div");
                loopTrigger.className = `node-trigger ${lData.hasFailure ? 'node-fail' : 'node-pass'}`;
                loopTrigger.setAttribute("onclick", "toggleNode(this)");
                loopTrigger.innerHTML = `
                    <div class="node-title">
                        <span class="caret ${lData.hasFailure ? 'rotated' : ''}">▶</span>
                        <span>🔄 LOOP ${String(lNum).padStart(4, '0')}</span>
                    </div>
                    <span class="badge ${lData.hasFailure ? 'badge-fail' : 'badge-pass'}">
                        ${lData.hasFailure ? 'Failure Detected' : 'All Clear'}
                    </span>
                `;

                const loopContent = document.createElement("div");
                loopContent.className = `node-content ${lData.hasFailure ? 'active' : ''}`;

                for (const [sName, sData] of Object.entries(lData.scenarios)) {
                    const scNode = document.createElement("div");
                    scNode.className = "collapsible-node";

                    const scTrigger = document.createElement("div");
                    scTrigger.className = `node-trigger ${sData.hasFailure ? 'node-fail' : 'node-pass'}`;
                    scTrigger.setAttribute("onclick", "toggleNode(this)");
                    scTrigger.innerHTML = `
                        <div class="node-title">
                            <span class="caret ${sData.hasFailure ? 'rotated' : ''}">▶</span>
                            <span>📋 Card: ${sName}</span>
                        </div>
                        <span class="badge ${sData.hasFailure ? 'badge-fail' : 'badge-pass'}">
                            ${sData.hasFailure ? 'Fail' : 'Pass'}
                        </span>
                    `;

                    const scContent = document.createElement("div");
                    scContent.className = `node-content ${sData.hasFailure ? 'active' : ''}`;

                    sData.points.forEach((p, idx) => {
                        const ptRow = document.createElement("div");
                        ptRow.className = "point-row";
                        ptRow.setAttribute("onclick", `toggleDetails('details-${lNum}-${sName.replace(/\\s/g,'_')}-${idx}')`);
                        
                        const badgeStr = p.overall === 'PASS' 
                            ? `<span class="badge badge-pass">PASS</span>` 
                            : `<span class="badge badge-fail">FAIL</span>`;
                        
                        const histId = `rerunhist-${lNum}-${sName.replace(/\\s/g,'_')}-${idx}`;
                        const rerunBadge = (p.attempts && p.attempts.length > 1)
                            ? `<span class="badge" onclick="event.stopPropagation(); toggleRerunHistory('${histId}')" title="Click to view all attempts" style="cursor:pointer; background: rgba(255, 111, 0, 0.12); color: #FF6F00; border: 1px solid #FF6F00;">🕒 RERUN ATTEMPT ${p.rerun_attempt} (History Available)</span>`
                            : (p.rerun_attempt > 0 ? `<span class="badge" style="background: rgba(255, 111, 0, 0.1); color: #FF6F00; border: 1px solid #FF6F00;">RERUN ATTEMPT ${p.rerun_attempt}</span>` : ``);

                        ptRow.innerHTML = `
                            <div style="display: flex; justify-content: space-between; align-items: center;">
                                <div><strong>${p.id}</strong> ${rerunBadge}</div>
                                <div>${badgeStr}</div>
                            </div>
                            <div style="font-size:0.8em; color:#a6adc8; margin-top:4px;">
                                ${p.overall === 'FAIL' ? `❌ ${p.failure_reason}` : '🟢 OK - Checks Verified'}
                            </div>

                            <!-- Consolidated Rerun Execution Attempt History -->
                            ${(p.attempts && p.attempts.length > 1) ? `
                            <div class="history-panel" id="${histId}" style="display: none; background: #1a1b26; border: 1px solid #FF6F00; border-radius: 4px; padding: 12px; margin-top: 10px;" onclick="event.stopPropagation()">
                                <div style="font-size: 0.85em; font-weight: bold; color: #FF6F00; border-bottom: 1px solid rgba(255,111,0,0.25); padding-bottom: 4px; margin-bottom: 8px;">🔄 EXECUTION HISTORY (ALL ATTEMPTS)</div>
                                ${p.attempts.map((att) => {
                                    const attBadge = att.overall === 'PASS'
                                        ? `<span class="badge badge-pass">PASS</span>`
                                        : `<span class="badge badge-fail">FAIL</span>`;
                                    return `
                                    <div style="border-bottom: 1px dashed #313244; padding: 6px 0; font-size: 0.85em;">
                                        <div style="display: flex; justify-content: space-between; align-items: center;">
                                            <strong style="color: #f9e2af;">Attempt ${att.attempt}</strong>
                                            ${attBadge}
                                        </div>
                                        <div style="color: #a6adc8; margin-top: 4px; font-size: 0.9em;">
                                            ${att.overall === 'FAIL' ? `❌ ${att.failure_reason}` : '🟢 Checks Verified'}
                                        </div>
                                        <div style="margin-top: 6px; display: flex; gap: 15px; align-items: center; flex-wrap: wrap;">
                                            ${att.relative_screenshot ? `
                                            <div style="font-size: 0.8em;">
                                                <a href="#" style="color: #89b4fa; text-decoration: underline;" onclick="openModal('${att.relative_screenshot}', event)">🖼 View Attempt Screenshot</a>
                                            </div>` : ''}
                                        </div>
                                        ${att.failure_diagnostics ? renderFailureDiagnostics({ failure_diagnostics: att.failure_diagnostics }) : ''}
                                    </div>
                                    `;
                                }).join('')}
                            </div>
                            ` : ''}

                            <div class="details-panel" id="details-${lNum}-${sName.replace(/\\s/g,'_')}-${idx}" onclick="event.stopPropagation()">
                                <div style="display: flex; gap: 20px; align-items: flex-start; text-align: left;">
                                    <div style="flex: 3;">
                                        <h4 style="margin-top:0; color: #89b4fa;">Step-by-Step Exec Trace Log</h4>
                                        ${renderPointDetails(p)}
                                        ${renderCustomChecks(p)}
                                        ${renderFailureDiagnostics(p)}
                                    </div>
                                    <div style="flex: 1; text-align: center;">
                                        <h4 style="margin-top:0; color: #89b4fa;">Captured Evidence</h4>
                                        ${p.relative_screenshot ? `<img class="screenshot-thumb" src="${p.relative_screenshot}" onclick="openModal('${p.relative_screenshot}', event)">` : 'No screenshot'}
                                        ${renderRightColumnExtras(p)}
                                    </div>
                                </div>
                            </div>
                        `;
                        scContent.appendChild(ptRow);
                    });

                    scNode.appendChild(scTrigger);
                    scNode.appendChild(scContent);
                    loopContent.appendChild(scNode);
                }

                loopNode.appendChild(loopTrigger);
                loopNode.appendChild(loopContent);
                treeBody.appendChild(loopNode);
            }
        }

        function renderPointDetails(p) {
            let html = "";
            if (p.sections) {
                p.sections.forEach(sec => {
                    const color = sec.overall === 'PASS' ? '#a6e3a1' : (sec.overall === 'SKIP' ? '#6c7086' : '#f38ba8');
                    html += `<div style="border-left:2px solid ${color}; padding-left:8px; margin-bottom:10px;">
                        <strong>${sec.label}</strong> [${sec.overall}]`;
                    
                    if (sec.trigger) {
                        sec.trigger.forEach(step => {
                            html += `<div class="nested-step"><span>${step.name}</span><span>${step.message}</span></div>`;
                        });
                    }
                    if (sec.normalize) {
                        sec.normalize.forEach(step => {
                            html += `<div class="nested-step"><span>${step.name}</span><span>${step.message}</span></div>`;
                        });
                    }
                    html += `</div>`;
                });
            }
            return html;
        }

        function toggleRerunHistory(histId) {
            const el = document.getElementById(histId);
            if (!el) return;
            el.style.display = (el.style.display === 'block') ? 'none' : 'block';
        }

        function renderRightColumnExtras(p) {
            // Shown in the RIGHT column, below the captured evidence screenshot.
            // This area is ONLY for asset-bound custom checks — it surfaces the
            // trigger timing + the actual OCR text / image score each custom
            // check captured. For default IO tests (no custom checks) this stays
            // empty: the trigger info already lives in Failure Diagnostics on a
            // fail, and isn't meaningful to repeat on a pass.
            if (!p.steps) return "";
            const custom = p.steps.filter(s => s.is_custom);
            if (!custom.length) return "";   // default IO test → nothing here

            let html = "";

            // ── Trigger timing (context for the custom checks below) ─────────
            const ti = p.trigger_info;
            if (ti && (ti.trigger_time && ti.trigger_time !== 'N/A')) {
                html += `
                    <div style="margin-top:12px; background:#11121a; border:1px solid #222330;
                                border-radius:4px; padding:8px; text-align:left;
                                font-size:0.78em; font-family:monospace; color:#a6adc8; line-height:1.6em;">
                        <div style="font-weight:bold; color:#89b4fa; margin-bottom:4px; font-family:Inter,sans-serif;">⚡ Alarm Trigger</div>
                        ${(ti.trigger_value && ti.trigger_value !== 'N/A') ? `🎯 Value: <span style="color:#f9e2af;">${ti.trigger_value}</span><br>` : ``}
                        🚀 Triggered: <span style="color:#ffffff;">${ti.trigger_time}</span>
                    </div>`;
            }

            // ── Custom-check captured results (OCR text / image score) ──────
            html += `<div style="margin-top:12px; text-align:left;">
                <div style="font-weight:bold; color:#f9c74f; font-size:0.82em; margin-bottom:6px;">★ Custom Check Captures</div>`;
            custom.forEach(c => {
                const passed = c.status === 'PASS';
                const col = passed ? '#a6e3a1' : '#f38ba8';
                html += `
                    <div style="background:#11121a; border:1px solid #222330; border-left:3px solid ${col};
                                border-radius:4px; padding:7px; margin-bottom:6px; font-size:0.78em;
                                font-family:monospace; color:#cdd6f4; line-height:1.5em;">
                        <div style="font-family:Inter,sans-serif; color:#cdd6f4; margin-bottom:3px;">
                            <strong>${c.name}</strong> <span style="color:${col};">[${c.status}]</span>
                        </div>
                        <div style="color:#6c7086;">expected</div>
                        <div style="color:#a6e3a1; word-break:break-word;">${c.expected || '\u2014'}</div>
                        <div style="color:#6c7086; margin-top:3px;">actual (captured)</div>
                        <div style="color:${col}; word-break:break-word;">${c.actual || '\u2014'}</div>
                    </div>`;
            });
            html += `</div>`;

            return html;
        }

        function renderCustomChecks(p) {
            // p.steps may contain entries with is_custom === true,
            // produced by VERIFY_CUSTOM asset-binding steps.
            if (!p.steps) return "";
            const custom = p.steps.filter(s => s.is_custom);
            if (!custom.length) return "";

            let html = `<div style="margin-top:10px;">
                <strong style="color:#f9c74f;">\u2605 Custom Asset Verifications</strong>`;

            custom.forEach(c => {
                const passed   = c.status === 'PASS';
                const badge    = passed
                    ? `<span class="badge badge-pass">PASS</span>`
                    : `<span class="badge badge-fail">FAIL</span>`;
                const assetTag = c.asset_name
                    ? `<span class="asset-pill">${c.asset_name}${c.asset_id ? ' ('+c.asset_id+')' : ''}</span>`
                    : '';
                const actualClass = passed ? 'ea-actual-pass' : 'ea-actual-fail';

                html += `<div class="custom-check">
                    <div class="custom-check-row">
                        <span><strong>${c.name}</strong>${assetTag}</span>
                        <span>${badge}</span>
                    </div>
                    ${(c.expected || c.actual) ? `
                    <div class="custom-check-ea">
                        <span>Expected</span><span class="ea-expected">${c.expected || '\u2014'}</span>
                        <span>Actual</span><span class="${actualClass}">${c.actual || '\u2014'}</span>
                    </div>` : ''}
                    ${c.message ? `<div style="color:#6c7086; margin-top:4px; font-size:0.9em;">${c.message}</div>` : ''}
                </div>`;
            });

            html += `</div>`;
            return html;
        }

        function renderFailureDiagnostics(p) {
            if (!p.failure_diagnostics) return "";
            const diag = p.failure_diagnostics;
            
            let zoneCropsHtml = "";
            if (diag.cropped_zones) {
                for (const [zName, zData] of Object.entries(diag.cropped_zones)) {
                    zoneCropsHtml += `
                        <div style="background:#1a1b26; border:1px solid #222330; padding:8px; border-radius:4px; margin-bottom:8px; display:flex; gap:10px; align-items:center;">
                            <div style="flex:1; text-align:center;">
                                <div style="font-size:0.75em; color:#89b4fa; margin-bottom:4px; font-weight:bold;">${zName.replace(/_/g, ' ').toUpperCase()}</div>
                                <img src="${zData.image}" style="max-height:80px; max-width:100%; border:1px solid #313244; border-radius:2px; cursor:pointer;" onclick="openModal('${zData.image}', event)">
                            </div>
                            <div style="flex:2; background:#0d0e12; padding:6px; font-size:0.8em; border-radius:2px; font-family:monospace; max-height:80px; overflow-y:auto; border-left:2px solid #89b4fa; white-space:pre-wrap; text-align:left; color:#cdd6f4;">${zData.text || '(No text parsed)'}</div>
                        </div>
                    `;
                }
            }

            let mbInfoHtml = "";
            if (diag.modbus_info) {
                const mb = diag.modbus_info;
                const alm = mb.alarm_trigger || {};
                const nrm = mb.normalize_trigger || {};

                const almTriggerTime  = alm.trigger_time    || mb.trigger_time    || 'N/A';
                const almScadaClock   = alm.scada_clock_ocr || 'N/A';
                const almVal          = alm.trigger_value    || mb.trigger_value   || 'N/A';
                const nrmTriggerTime  = nrm.trigger_time    || 'N/A';
                const nrmScadaClock   = nrm.scada_clock_ocr || 'N/A';
                const nrmVal          = nrm.trigger_value   || 'N/A';

                mbInfoHtml = `
                    <div style="font-size:0.85em; color:#a6adc8; line-height:1.7em; font-family:monospace; text-align:left;">
                        ⚡ <strong>Modbus Target:</strong> FC=${mb.function_code} | Reg=${mb.register_address} | Bit=${mb.bit_offset} | Unit=${mb.device_address_unit_id}
                        <hr style="border-color:#313244; margin:6px 0;">
                        🎯 <strong>Trigger Val:</strong> <span style="color:#f9e2af;">${almVal}</span><br>
                        🚀 <strong>Trigger Time:</strong> <span style="color:#ffffff;">${almTriggerTime}</span><br>
                        🖥️ <strong>SCADA Clock (OCR):</strong> <span style="color:#a6e3a1;">${almScadaClock}</span>
                        <hr style="border-color:#313244; margin:6px 0;">
                        🎯 <strong>Reset Val:</strong> <span style="color:#f9e2af;">${nrmVal}</span><br>
                        🚀 <strong>Reset Time:</strong> <span style="color:#ffffff;">${nrmTriggerTime}</span><br>
                        🖥️ <strong>SCADA Clock (OCR):</strong> <span style="color:#a6e3a1;">${nrmScadaClock}</span>
                    </div>
                `;
            }

            return `
                <div style="margin-top:15px; border-top:1px solid #313244; padding-top:15px; text-align:left;">
                    <h4 style="margin:0 0 10px 0; color:#f38ba8; font-size: 0.95em;">🔍 Failure Diagnostics</h4>
                    <div style="display:flex; gap:15px; flex-wrap:wrap;">
                        <div style="flex:1; min-width:250px; background:#11121a; padding:10px; border-radius:4px; border:1px solid #222330;">
                            <div style="font-weight:bold; color:#89b4fa; margin-bottom:6px; font-size:0.85em;">Modbus / Protocol Context</div>
                            ${mbInfoHtml}
                        </div>
                        <div style="flex:2; min-width:400px; background:#11121a; padding:10px; border-radius:4px; border:1px solid #222330;">
                            <div style="font-weight:bold; color:#89b4fa; margin-bottom:6px; font-size:0.85em;">Cropped Regions & OCR Results</div>
                            ${zoneCropsHtml || '<div style="color:#585b70; font-size:0.85em;">No region crop diagnostics captured</div>'}
                        </div>
                    </div>
                </div>
            `;
        }

        function renderEvidenceTree() {
            const evBody = document.getElementById("evidence-tree-body");
            evBody.innerHTML = "";

            if (!rawEvidence || rawEvidence.length === 0) {
                evBody.innerHTML = '<p style="color: #585b70; font-size: 0.9em; padding: 10px;">No files indexed. Double-check your path scan logs.</p>';
                return;
            }

            const tree = {};

            rawEvidence.forEach(f => {
                const parts = f.rel_path.split("/");
                const loopPart = parts.find(p => p.startsWith("loop_")) || "Global";
                const loopLabel = loopPart.startsWith("loop_") 
                    ? "LOOP " + loopPart.split("_")[1] 
                    : loopPart.toUpperCase();
                
                const cardPart = parts[1] || "Global Assets";
                
                if (!tree[loopLabel]) tree[loopLabel] = {};
                if (!tree[loopLabel][cardPart]) tree[loopLabel][cardPart] = {};
                
                const cat = f.category;
                if (!tree[loopLabel][cardPart][cat]) tree[loopLabel][cardPart][cat] = [];
                
                tree[loopLabel][cardPart][cat].push(f);
            });

            const categoryMeta = {
                video: { icon: "🎥", label: "Recording" },
                screenshot: { icon: "🖼", label: "Screenshots" },
                data: { icon: "📄", label: "Data Sheets" },
                log: { icon: "📜", label: "Logs" }
            };

            for (const [loop, cards] of Object.entries(tree)) {
                const loopNode = document.createElement("div");
                loopNode.className = "collapsible-node";

                const loopTrigger = document.createElement("div");
                loopTrigger.className = "node-trigger";
                loopTrigger.setAttribute("onclick", "toggleNode(this)");
                loopTrigger.innerHTML = `
                    <div class="node-title">
                        <span class="caret">▶</span>
                        <span>📂 ${loop}</span>
                    </div>
                `;

                const loopContent = document.createElement("div");
                loopContent.className = "node-content";

                for (const [card, categories] of Object.entries(cards)) {
                    const cardNode = document.createElement("div");
                    cardNode.className = "collapsible-node";
                    cardNode.style.marginLeft = "10px";

                    const cardTrigger = document.createElement("div");
                    cardTrigger.className = "node-trigger";
                    cardTrigger.setAttribute("onclick", "toggleNode(this)");
                    cardTrigger.innerHTML = `
                        <div class="node-title">
                            <span class="caret">▶</span>
                            <span>📋 Card: ${card}</span>
                        </div>
                    `;

                    const cardContent = document.createElement("div");
                    cardContent.className = "node-content";

                    for (const [catType, files] of Object.entries(categories)) {
                        const meta = categoryMeta[catType] || { icon: "📁", label: catType.toUpperCase() };
                        const catNode = document.createElement("div");
                        catNode.className = "collapsible-node";
                        catNode.style.marginLeft = "15px";

                        const catTrigger = document.createElement("div");
                        catTrigger.className = "node-trigger";
                        catTrigger.setAttribute("onclick", "toggleNode(this)");
                        catTrigger.innerHTML = `
                            <div class="node-title">
                                <span class="caret">▶</span>
                                <span>${meta.icon} ${meta.label}</span>
                            </div>
                            <span class="badge" style="background:#1e1e2e; color:#89b4fa;">${files.length}</span>
                        `;

                        const catContent = document.createElement("div");
                        catContent.className = "node-content";

                        if (catType === "screenshot") {
                            const subGroups = { FAIL: [], PASS: [], OTHER: [] };
                            files.forEach(f => {
                                const nameUpper = f.name.toUpperCase();
                                if (nameUpper.includes("_FAIL") || nameUpper.includes("FAIL_")) {
                                    subGroups.FAIL.push(f);
                                } else if (nameUpper.includes("_PASS") || nameUpper.includes("PASS_")) {
                                    subGroups.PASS.push(f);
                                } else {
                                    subGroups.OTHER.push(f);
                                }
                            });

                            for (const [groupName, groupFiles] of Object.entries(subGroups)) {
                                if (groupFiles.length === 0) continue;
                                
                                const groupNode = document.createElement("div");
                                groupNode.className = "collapsible-node";
                                groupNode.style.marginLeft = "15px";

                                const isFail = groupName === "FAIL";
                                const isPass = groupName === "PASS";
                                const groupColor = isFail ? "#f38ba8" : (isPass ? "#a6e3a1" : "#a6adc8");

                                const groupTrigger = document.createElement("div");
                                groupTrigger.className = "node-trigger";
                                groupTrigger.setAttribute("onclick", "toggleNode(this)");
                                groupTrigger.innerHTML = `
                                    <div class="node-title">
                                        <span class="caret ${isFail ? 'rotated' : ''}">▶</span>
                                        <span style="color: ${groupColor}; font-weight: bold;">${groupName}</span>
                                    </div>
                                    <span class="badge" style="background: ${isFail ? 'rgba(243, 139, 168, 0.12)' : 'rgba(166, 227, 161, 0.12)'}; color: ${groupColor}; border: 1px solid ${groupColor};">${groupFiles.length}</span>
                                `;

                                const groupContent = document.createElement("div");
                                groupContent.className = `node-content ${isFail ? 'active' : ''}`;

                                groupFiles.forEach(f => {
                                    groupContent.appendChild(createFileRowElement(f, isFail));
                                });

                                groupNode.appendChild(groupTrigger);
                                groupNode.appendChild(groupContent);
                                catContent.appendChild(groupNode);
                            }
                        } else {
                            files.forEach(f => {
                                catContent.appendChild(createFileRowElement(f, false));
                            });
                        }

                        catNode.appendChild(catTrigger);
                        catNode.appendChild(catContent);
                        cardContent.appendChild(catNode);
                    }

                    cardNode.appendChild(cardTrigger);
                    cardNode.appendChild(cardContent);
                    loopContent.appendChild(cardNode);
                }

                loopNode.appendChild(loopTrigger);
                loopNode.appendChild(loopContent);
                evBody.appendChild(loopNode);
            }
        }

        function createFileRowElement(f, isFail) {
            const fileRow = document.createElement("div");
            fileRow.style.display = "flex";
            fileRow.style.justify = "space-between";
            fileRow.style.alignItems = "center";
            fileRow.style.padding = "6px 12px";
            fileRow.style.borderBottom = "1px solid #1e1e2e";
            fileRow.style.fontSize = "0.85em";
            fileRow.style.marginLeft = "20px";

            const textStyle = isFail ? "color: #f38ba8; font-weight: bold;" : "color: #89b4fa;";
            const escapedFullPath = f.win_full_path.replace(/\\\\/g, '\\\\\\\\').replace(/'/g, "\\'");
            const escapedFolderPath = f.win_folder_path.replace(/\\\\/g, '\\\\\\\\').replace(/'/g, "\\'");

            fileRow.innerHTML = `
                <div>
                    <span>📄 </span>
                    <a href="${f.rel_path}" target="_blank" style="${textStyle} text-decoration:none;">${f.name}</a>
                    <span style="color:#585b70; margin-left:10px;">(${f.size_label})</span>
                </div>
                <div style="display:flex; gap:8px;">
                    <button class="action-btn" onclick="copyToClipboard('${escapedFullPath}')">📋 Copy Path</button>
                    <button class="action-btn" onclick="copyToClipboard('${escapedFolderPath}')">📁 Copy Folder Path</button>
                </div>
            `;
            return fileRow;
        }

        renderTree();
        renderEvidenceTree();
    </script>
</body>
</html>"""

        report_content = (html_template
            .replace("__TITLE__", str(title))
            .replace("__START_TIME__", start_time.strftime('%Y-%m-%d %H:%M:%S'))
            .replace("__END_TIME__", end_time.strftime('%Y-%m-%d %H:%M:%S'))
            .replace("__DURATION__", str(duration).split('.')[0])
            .replace("__TOTAL_LOOPS__", str(total_loops))
            .replace("__TOTAL_SCENARIOS__", str(total_scenarios))
            .replace("__PASS_RATE__", f"{pass_rate:.1f}%")
            .replace("__PASS_RATE_NUM__", str(pass_rate))
            .replace("__PASS_RATE_REMAIN__", str(100 - pass_rate))
            .replace("__PASS_RATE_COLOR__", "#a6e3a1" if pass_rate == 100 else "#f9e2af")
            .replace("__PASSED__", str(passed))
            .replace("__FAILED__", str(failed))
            .replace("__TOTAL__", str(total))
            .replace("__RERUN_POINTS__", str(rerun_points))
            .replace("__OCR_MISMATCH__", str(categories["OCR Mismatch"]))
            .replace("__COLOR_MISS__", str(categories["Color Miss"]))
            .replace("__TIMEOUT__", str(categories["Timeout"]))
            .replace("__DATETIME_OUT__", str(categories["Datetime Out of Bounds"]))
            .replace("__OTHER__", str(categories["Other"]))
            .replace("__FAIL_SCREENSHOTS__", str(evidence_counts["screenshot_fail"]))
            .replace("__PASS_SCREENSHOTS__", str(evidence_counts["screenshot_pass"]))
            .replace("__RECORDINGS__", str(evidence_counts["video"]))
            .replace("__DATA_FILES__", str(evidence_counts["data"]))
            .replace("__TOTAL_FILES__", str(len(evidence_list)))
            .replace("__RESULTS_JSON__", results_json)
            .replace("__EVIDENCE_JSON__", evidence_json)
        )

        try:
            with open(output_dir / "Suite_Report.html", "w", encoding="utf-8") as f:
                f.write(report_content)
            logger.info("ReportManager: HTML Dashboard generated successfully.")
        except Exception as e:
            logger.error(f"ReportManager: HTML generation failed: {e}")

    @classmethod
    def _write_excel_report(cls, results, output_dir, start_time, end_time, title):
        if not PANDAS_AVAILABLE:
            logger.warning("ReportManager: Pandas/Openpyxl not installed. Skipping Excel generation.")
            return
        
        duration = end_time - start_time
        total = len(results)
        passed = sum(1 for r in results if r["overall"] == "PASS")
        failed = total - passed
        pass_rate = (passed / total) if total > 0 else 0

        categories = {"OCR Mismatch": 0, "Color Miss": 0, "Timeout": 0, "Datetime Out of Bounds": 0, "Other": 0}
        for r in results:
            if r["overall"] == "FAIL" and r["failure_category"] in categories:
                categories[r["failure_category"]] += 1

        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            excel_path = output_dir / "Test_Run_Workbook.xlsx"
            writer = pd.ExcelWriter(excel_path, engine='openpyxl')
            
            dash_df = pd.DataFrame([
                ["Metric", "Value"],
                ["Scenario Name", title],
                ["Start Time", start_time.strftime('%Y-%m-%d %H:%M:%S')],
                ["End Time", end_time.strftime('%Y-%m-%d %H:%M:%S')],
                ["Duration", str(duration).split('.')[0]],
                ["Total Points", total],
                ["Passed Points", passed],
                ["Failed Points", failed],
                ["Pass Rate (%)", f"{pass_rate*100:.1f}%"]
            ])
            dash_df.to_excel(writer, sheet_name="Dashboard", index=False, header=False)
            
            detailed_data = []
            for r in results:
                detailed_data.append({
                    "Point ID": r["id"],
                    "Status": r["overall"],
                    "Failure Category": r["failure_category"] if r["overall"] == "FAIL" else "—",
                    "Primary Failure Reason": r["failure_reason"] if r["overall"] == "FAIL" else "None",
                    "Screenshot File": Path(r["screenshot"]).name if r["screenshot"] else "None"
                })
            
            det_df = pd.DataFrame(detailed_data)
            det_df.to_excel(writer, sheet_name="Detailed Results", index=False)
            
            workbook = writer.book
            
            # Format Dashboard Tab
            ws_dash = workbook["Dashboard"]
            ws_dash.views.sheetView[0].showGridLines = True
            
            title_font = Font(name="Consolas", size=14, bold=True, color="FFFFFF")
            header_font = Font(name="Consolas", size=11, bold=True, color="FFFFFF")
            bold_font = Font(name="Consolas", size=11, bold=True)
            norm_font = Font(name="Consolas", size=11)
            
            banner_fill = PatternFill(start_color="1F1F2E", end_color="1F1F2E", fill_type="solid")
            accent_fill = PatternFill(start_color="2979FF", end_color="2979FF", fill_type="solid")
            gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            
            thin_border = Border(
                left=Side(style='thin', color='DDDDDD'),
                right=Side(style='thin', color='DDDDDD'),
                top=Side(style='thin', color='DDDDDD'),
                bottom=Side(style='thin', color='DDDDDD')
            )
            
            ws_dash.insert_rows(1, 2)
            ws_dash.merge_cells("A1:B2")
            ws_dash["A1"] = f"ISCS AUTO RUN SUMMARY EXECUTIVE DASHBOARD"
            ws_dash["A1"].font = title_font
            ws_dash["A1"].fill = banner_fill
            ws_dash["A1"].alignment = Alignment(horizontal="center", vertical="center")
            
            for row in range(3, 12):
                ws_dash.cell(row=row, column=1).font = bold_font
                ws_dash.cell(row=row, column=2).font = norm_font
                ws_dash.cell(row=row, column=1).border = thin_border
                ws_dash.cell(row=row, column=2).border = thin_border
                ws_dash.cell(row=row, column=1).fill = gray_fill
                
            ws_dash.cell(row=13, column=1, value="Failure Profiling Categories").font = bold_font
            ws_dash.cell(row=13, column=1).fill = gray_fill
            ws_dash.cell(row=13, column=1).border = thin_border
            ws_dash.cell(row=13, column=2, value="Count").font = bold_font
            ws_dash.cell(row=13, column=2).fill = gray_fill
            ws_dash.cell(row=13, column=2).border = thin_border
            
            curr_row = 14
            for cat, val in categories.items():
                ws_dash.cell(row=curr_row, column=1, value=cat).font = norm_font
                ws_dash.cell(row=curr_row, column=1).border = thin_border
                ws_dash.cell(row=curr_row, column=2, value=val).font = norm_font
                ws_dash.cell(row=curr_row, column=2).border = thin_border
                curr_row += 1
                
            ws_dash.column_dimensions['A'].width = 32
            ws_dash.column_dimensions['B'].width = 25

            # Format Detailed Tab
            ws_det = workbook["Detailed Results"]
            ws_det.views.sheetView[0].showGridLines = True
            
            for col in range(1, 6):
                cell = ws_det.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = accent_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            
            for row in range(2, len(results) + 2):
                status_cell = ws_det.cell(row=row, column=2)
                row_fill = green_fill if status_cell.value == "PASS" else red_fill
                
                for col in range(1, 6):
                    c = ws_det.cell(row=row, column=col)
                    c.font = norm_font
                    c.fill = row_fill
                    c.border = thin_border
            
            for col in ws_det.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = col[0].column_letter
                ws_det.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
            writer.close()
            logger.info("ReportManager: Excel Multi-Sheet report compiled successfully.")
        except Exception as e:
            logger.error(f"ReportManager: Excel generation failed: {e}")

    @classmethod
    def _inject_evidence_manifest(cls, output_dir: Path):
        """Deprecated compatibility stub. Standard scanning is handled natively prior to HTML construction."""
        pass