#!/usr/bin/env python3
"""
WilloWisp_v1_iscs.py 
UI Testing, Closed-Loop Test Automation Framework
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Modes:
  1. Targeted Sequence (RPA)
  2. Grid Scan (Fuzzer)
  3. Suite Runner (Modbus + OCR Closed-Loop + Expanded Zones)
"""

# ── DPI awareness — MUST be first ───────────
import ctypes
import ctypes.wintypes
try:
    ctypes.windll.user32.SetProcessDpiAwarenessContext(-4)  # Per-Monitor v2
except Exception:
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(2)      # Per-Monitor v1
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()       # System-aware
        except Exception:
            pass
# ─────────────────────────────────────────────

import os
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import json
import time
import threading
import datetime
import logging
import traceback
import sys
import math
import re
import pathlib
import copy
from pathlib import Path

# ── Upgrade modules (Feature 1: Visual Anchoring, Feature 2: Frame Sampler) ──
try:
    from iscs_Sampler_Anchor import VisualAnchor, AnchorManager, FrameSampler
    UPGRADES_AVAILABLE = True
except ImportError:
    UPGRADES_AVAILABLE = False

# ── OCR Subsystem ─────────────────────────────────────────────────────────────
try:
    import iscs_OCR
except ImportError:
    iscs_OCR = None

# ── Dynamic Report Manager Module ───────────
try:
    from iscs_reports import ReportManager
except Exception as e:
    ReportManager = None
    print(f"CRITICAL: Failed to import ReportManager from iscs_reports.py. Error: {e}")

# ── Workflow / Procedure Engine ──────────────
try:
    from iscs_workflow import (
        ProcedureFlow, ProcedureRunner,
        auto_register_procedures, build_runner_from_scenario,
        open_procedure_flow_dialog,
    )
    WORKFLOW_AVAILABLE = True
except ImportError as e:
    WORKFLOW_AVAILABLE = False
    print(f"WARNING: iscs_workflow.py not found — procedure engine disabled. ({e})")

# ── Lifecycle event bus (optional; additive — see ARCHITECTURE_DESIGN.md) ──────
try:
    from iscs_core import (
        bus as CORE_BUS,
        SuiteStarted, SuiteCompleted, CardStarted, CardCompleted,
        discover_directory,
    )
    _CORE_EVENTS_OK = True
except Exception as _ce:
    CORE_BUS = None
    SuiteStarted = SuiteCompleted = CardStarted = CardCompleted = None
    discover_directory = None
    _CORE_EVENTS_OK = False
    print(f"INFO: iscs_core events unavailable — lifecycle events disabled. ({_ce})")

# ── Capability load manifest (optional; FR-18/FR-19 — diagnostics only) ────────
try:
    from iscs_core import (
        LoadManifest, evaluate_requirements,
        registry as CORE_REGISTRY,
    )
    _CORE_MANIFEST_OK = True
except Exception:
    LoadManifest = evaluate_requirements = CORE_REGISTRY = None
    _CORE_MANIFEST_OK = False

# Populated by _load_plugins(): the single startup snapshot of what loaded /
# what's unavailable (unmet requirements) / what failed to import (FR-19).
PLUGIN_MANIFEST = None


# Plugin categories discovered at startup (extend as capabilities are ported out
# of the engine). Each discovered file self-registers, overriding its legacy
# adapter by key — see plugins/README.md.
_PLUGIN_CATEGORIES = ("utilities", "verifications", "actions")


def _load_plugins():
    """Discover ported capability plugins at app startup. Best-effort and isolated:
    a broken plugin is logged + skipped, never blocking launch.

    Also builds the capability load manifest (FR-18/FR-19): what loaded, what has
    unmet requirements, and what failed to import — printed once for diagnostics.
    Capabilities with unmet requirements are reported, NOT disabled (the live
    engine keeps its legacy fallback), so this is purely additive."""
    if discover_directory is None:
        return
    global PLUGIN_MANIFEST
    manifest = LoadManifest() if _CORE_MANIFEST_OK else None
    base = Path(__file__).parent / "plugins"
    for category in _PLUGIN_CATEGORIES:
        try:
            loaded = discover_directory(base / category, manifest=manifest)
            if loaded:
                print(f"INFO: loaded plugin(s) from plugins/{category}: {loaded}")
        except Exception as _pe:
            print(f"WARNING: plugin discovery failed for plugins/{category}: {_pe}")

    if manifest is not None and CORE_REGISTRY is not None:
        manifest.record_registry(CORE_REGISTRY)          # include legacy adapters
        unavailable = evaluate_requirements(CORE_REGISTRY, manifest, disable=False)
        PLUGIN_MANIFEST = manifest
        print("INFO: " + manifest.summary().replace("\n", "\n      "))
        if unavailable:
            print(f"INFO: capabilities with unmet requirements (left enabled): {unavailable}")

    # Coverage check (P6.3): confirm every step type resolves in the registry, so
    # the legacy _exec_* fallback is provably vestigial. Guarded — never blocks launch.
    try:
        from iscs_workflow import registry_step_coverage
        covered, missing = registry_step_coverage()
        if missing:
            print(f"WARNING: {len(missing)} step type(s) have no registered capability "
                  f"(legacy fallback would be used): {missing}")
        else:
            print(f"INFO: registry covers all {len(covered)} step types "
                  f"(legacy fallback inactive).")
    except Exception:
        pass


def _wire_subscribers():
    """Subscribe event-driven subsystems to the shared bus at startup (P2.3).
    The report subsystem generates the consolidated report on SuiteCompleted, so
    SuiteRunner no longer calls ReportManager directly (it keeps a safety-net
    fallback if no subscriber handled the event)."""
    if not _CORE_EVENTS_OK or CORE_BUS is None:
        return
    if ReportManager is not None:
        CORE_BUS.subscribe(SuiteCompleted, ReportManager.on_suite_completed)
        print("INFO: report subsystem subscribed to SuiteCompleted.")

# ── Asset repository ──────────────────────────────────────────────────────────
try:
    from iscs_assets import set_app_dir as _set_asset_app_dir, AssetManager
    _set_asset_app_dir(Path(__file__).parent)
    _ASSETS_AVAILABLE = True
except Exception as _ae:
    _ASSETS_AVAILABLE = False
    print(f"INFO: iscs_assets not available — asset binding disabled. ({_ae})")

# ── Screen Recorder ───────────────────────────────────────────────────────────
try:
    from iscs_recorder import (
        Recorder, RecorderSettings, pre_flight_check,
        FPS_OPTIONS, DEFAULT_FPS, RECORDER_AVAILABLE,
    )
    _RECORDER_MODULE_OK = True
except ImportError as e:
    _RECORDER_MODULE_OK  = False
    RECORDER_AVAILABLE   = False
    RecorderSettings     = None
    Recorder             = None
    pre_flight_check     = None
    FPS_OPTIONS          = [1, 5, 10, 15, 24, 30, 60]
    DEFAULT_FPS          = 5
    print(f"WARNING: iscs_recorder.py not found — recording disabled. ({e})")

# ── Optional deps ─────────────────────────────────────────────────────────────
try:
    import pyautogui
    pyautogui.FAILSAFE = False
    pyautogui.PAUSE = 0.05
    PYAUTOGUI_AVAILABLE = True
except ImportError:
    PYAUTOGUI_AVAILABLE = False

try:
    from PIL import ImageTk, ImageGrab, ImageDraw, Image
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

try:
    import keyboard
    KEYBOARD_AVAILABLE = True
except ImportError:
    KEYBOARD_AVAILABLE = False

try:
    from screeninfo import get_monitors
    SCREENINFO_AVAILABLE = True
except ImportError:
    SCREENINFO_AVAILABLE = False

# ── ISCS / Falani Deps ────────────────────────────────────────────────────────

try:
    import pandas as pd
    import openpyxl
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

# OCR and Tesseract availability will be configured dynamically via Settings
TESSERACT_AVAILABLE = False

try:
    import asyncio
    from pymodbus.server import ModbusTcpServer
    from pymodbus.datastore import ModbusSequentialDataBlock, ModbusDeviceContext, ModbusServerContext
    PYMODBUS_AVAILABLE = True
except Exception as e:
    print(f"PyModbus Import Error: {e}")
    PYMODBUS_AVAILABLE = False

# ── Path & Config Management ──────────────────────────────────────────────────
if getattr(sys, 'frozen', False):
    BASE_DIR = Path(sys.executable).parent
else:
    BASE_DIR = Path(__file__).resolve().parent

LOG_DIR = BASE_DIR / "test_logs"
CONFIG_PATH   = BASE_DIR / "config.json"
TEMPLATE_PATH = BASE_DIR / "iscs_template.json"
LOG_DIR.mkdir(parents=True, exist_ok=True)

# M3.4: tell the core where the app base dir is, so relocated core services
# (the run coordinator) resolve the same test-logs output directory headlessly.
try:
    from core.services.config import set_base_dir as _set_core_base_dir
    _set_core_base_dir(BASE_DIR)
except Exception:
    pass

def _load_template() -> dict:
    """Load iscs_template.json. Returns empty structure if not found."""
    try:
        if TEMPLATE_PATH.exists():
            with open(TEMPLATE_PATH, "r") as f:
                return json.load(f)
    except Exception as e:
        logger.warning(f"Failed to load iscs_template.json: {e}")
    return {"zones": {}, "navigation": {}}

def _save_template(data: dict):
    """Merge data into iscs_template.json and write it."""
    try:
        existing = _load_template()
        if "zones" in data:
            existing.setdefault("zones", {}).update(data["zones"])
        if "navigation" in data:
            existing.setdefault("navigation", {}).update(data["navigation"])
        with open(TEMPLATE_PATH, "w") as f:
            json.dump(existing, f, indent=2)
    except Exception as e:
        logger.warning(f"Failed to save iscs_template.json: {e}")

# Default ISCS Configuration Matrix (Now customizable in JSON)
# Added "name" attributes to make color lookups dynamic and maintainable
# M2.2: severity matrix + config defaults + load/save relocated to
# core/services/config.py. The provider holds the live config; APP_CONFIG is the
# same dict instance, so all existing mutations/reads are unchanged.
from core.services.config import (
    SEVERITY_MATRIX, DEFAULT_CONFIG, ConfigProvider, SeverityColorClassifier,
)
_config_provider = ConfigProvider(CONFIG_PATH)
APP_CONFIG = _config_provider.config

def initialize_tesseract():
    global TESSERACT_AVAILABLE
    if iscs_OCR:
        TESSERACT_AVAILABLE = iscs_OCR.initialize(APP_CONFIG.get("tesseract_cmd", ""))
    else:
        TESSERACT_AVAILABLE = False

initialize_tesseract()

def save_config():
    # M2.2: delegates to the relocated ConfigProvider (same live config dict).
    _config_provider.save()

# ── Constants ─────────────────────────────────────────────────────────────────
GRID_SPACING     = APP_CONFIG["grid_spacing"]
CLICK_DELAY      = APP_CONFIG["click_delay"]
MOUSE_DRIFT_PX   = APP_CONFIG["mouse_drift_px"]
SCREENSHOT_DELAY = 0.25
WIDE_CROP_PAD    = 200   
HEARTBEAT_SEC    = 900   

# Falani IO List Mapping Definitions (Expanded for all IO list columns)
ISCS_ALIASES = {
    "point_id":        ["point_id", "point id", "tag", "tag id", "eqpt_identifier", "eqpt identifier",
                        "identifier", "point_name", "id", "point"],
    "equipment_desc":  ["eqpt_description", "eqpt description", "equipment_description",
                        "equipment description", "equip_desc", "equipment", "eq_desc"],
    "location":        ["equipment_location", "equipment location", "location", "room",
                        "specific_door_location", "specific door location", "area"],
    "attribute_desc":  ["attribute_description", "attribute description", "point_description",
                        "point description", "attr_desc", "alarm_description", "alarm description"],
    "station_code":    ["station_code", "station code", "station", "site"],
    "data_type":       ["dc_data_type", "dc data type", "data_type", "data type", "io_type"],
    "severity":        ["severity", "alarm level", "priority", "level", "alarm_level"],
    "protocol":        ["protocol", "comm", "interface"],
    # Modbus Specifics
    "device_address":  ["device_address", "device address", "unit_id", "unit id", "slave_id", "slave id"],
    "fc":              ["function_code", "function code", "fc", "modbus_fc"],
    "reg":             ["register_address", "register address", "register_addressfile",
                        "register address/file", "reg_addr", "address", "register", "reg"],
    "bit":             ["bit_offset", "bit offset", "bit", "offset"],
    "addr_size":       ["dc_addr_size", "dc addr size", "addr_size", "size"],
    "dc_io_type":      ["dc_io_type", "dc io type"],
    # SNMP (Future)
    "oid":             ["oid", "snmp_oid"],
    "value":           ["value", "trigger_value"]
}
ISCS_REQUIRED = ["point_id"]
IO_SHEET_KEYWORDS = ["io", "iolist", "io list", "ams", "fas", "bas", "tms", "ecs", "iscs",
                     "alarm", "point", "register", "modbus", "snmp", "iscs."]

# ── Metadata Store path ───────────────────────────────────────────────────────
METADATA_DB_PATH = BASE_DIR / "iscs_metadata.db"

# ── Logging Setup ─────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s [%(levelname)s] %(threadName)s: %(message)s',
    handlers=[
        logging.FileHandler(LOG_DIR / "app_debug.log", encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger("WilloWisp")

# ── Modbus traffic log — inside test_logs/modbus_logs/, rotating 10 x 100MB ──
MODBUS_LOG_DIR = LOG_DIR / "modbus_logs"
MODBUS_LOG_DIR.mkdir(parents=True, exist_ok=True)

from logging.handlers import RotatingFileHandler as _RotatingFileHandler
mb_traffic_handler = _RotatingFileHandler(
    MODBUS_LOG_DIR / "modbus_traffic.log",
    maxBytes    = 100 * 1024 * 1024,   # 100 MB per file
    backupCount = 9,                    # 9 backups + 1 active = 10 files max
    encoding    = "utf-8"
)
mb_traffic_handler.setFormatter(logging.Formatter('%(asctime)s [%(levelname)s] %(threadName)s: %(message)s'))

for name in ["pymodbus", "modbus_traffic"]:
    l = logging.getLogger(name)
    l.propagate = False
    l.addHandler(mb_traffic_handler)
    l.setLevel(logging.DEBUG)

# M3.4: per-suite run logger relocated to core/services/run_coordinator.py; shim.
from core.services.run_coordinator import test_run_logger, init_test_run_log

# UI COLORS (Updated with v4 ISCS Zones)
INCLUDE_COLOR     = "#00C853"
EXCLUDE_COLOR     = "#FF1744"
TARGET_COLOR      = "#AA00FF"
ALARM_PANEL_COLOR = "#FF00FF"
POINT_COLOR       = "#2979FF"
FILTERED_COLOR    = "#FF6F00"
PAUSE_COLOR       = "#FFD600"
WARN_COLOR        = "#FF6F00"
EQUIP_ZONE_COLOR  = "#00BCD4"
ALARM_LIST_COLOR  = "#FF9800"
EVENT_LIST_COLOR  = "#8BC34A"
ANCHOR_COLOR      = "#FFD600"   # gold — visual anchor zones
SCREEN_COLORS     = ["#2979FF", "#FF6F00", "#AA00FF", "#00BCD4", "#FF4081"]

HUD_W, HUD_H, HUD_MARGIN, HUD_ALPHA = 310, 150, 18, 0.62
HANDLE_SIZE = 8
MIN_ZONE_PX = 20
CROSSHAIR_R, CROSSHAIR_ARM = 10, 7

# ── Excel Parsing Engine (Falani Logic) ───────────────────────────────────────

def _normalize(text):
    if not text: return ""
    return " ".join(str(text).lower().split())


# ── OCR-tolerant matching ─────────────────────────────────────────────────────
# SCADA panel OCR is noisy: dropped hyphens (OCC-0008 -> OCC0008), stray pipes,
# doubled spaces, O/0 and l/1/I confusion. Exact substring matching causes
# false fails even when the text is clearly correct on screen. These helpers
# match the way a human reading the panel would.

# M2.4: OCR text-match helpers relocated to core/services/text_match.py;
# re-exported here as shims (ISCSVerifier + tests use them unchanged).
from core.services.text_match import (
    _ocr_canon, _ocr_contains, _ocr_fuzzy_contains, TextMatcher,
)


# ── State Table Parser ────────────────────────────────────────────────────────
def _find_state_table_cols(norm_headers):
    result = []
    for vi in range(8):
        label_col = sev_col = state_col = None
        for ci, h in enumerate(norm_headers):
            h = h.strip()
            if re.search(rf'\bv{vi}[_\s]label\b|\blabel[_\s]*\({vi}\)|\blabel[_\s]{vi}\b', h):
                label_col = ci
            elif re.search(rf'\bv{vi}[_\s]severity\b|\bseverity[_\s]*\({vi}\)|\bseverity[_\s]{vi}\b', h):
                sev_col = ci
            elif re.search(rf'\bv{vi}[_\s]state\b|\bstate[_\s]*\({vi}\)|\bstate[_\s]{vi}\b', h):
                state_col = ci
        if label_col is not None or sev_col is not None:
            result.append({
                "value_index": vi,
                "label_col":    label_col,
                "severity_col": sev_col,
                "state_col":    state_col,
            })
    return result

def _extract_states(row, state_cols):
    states = {}
    for sc in state_cols:
        vi = sc["value_index"]
        label    = ""
        severity = 0
        state    = "N"
        if sc["label_col"] is not None and sc["label_col"] < len(row) and row[sc["label_col"]] is not None:
            label = str(row[sc["label_col"]]).strip()
        if sc["severity_col"] is not None and sc["severity_col"] < len(row) and row[sc["severity_col"]] is not None:
            try: severity = int(row[sc["severity_col"]])
            except: severity = 0
        if sc["state_col"] is not None and sc["state_col"] < len(row) and row[sc["state_col"]] is not None:
            state = str(row[sc["state_col"]]).strip().upper()
        if label: 
            states[vi] = {"label": label, "severity": severity, "state": state}
    return states

# M3.4: expected-state helpers relocated to core/services/expected_state.py;
# re-exported here as shims (baru call sites unchanged).
from core.services.expected_state import (
    _get_state_indices, _get_expected_for_value, build_expected,
)


# ── Metadata Store ────────────────────────────────────────────────────────────
import sqlite3 as _sqlite3
import hashlib as _hashlib

from contextlib import contextmanager

@contextmanager
def db_session():
    """Thread-safe connection context manager that guarantees cleanup."""
    conn = _sqlite3.connect(str(METADATA_DB_PATH), check_same_thread=False)
    conn.row_factory = _sqlite3.Row
    try:
        yield conn
        conn.commit()
    except Exception as e:
        conn.rollback()
        logger.error(f"Database transaction failed: {e}")
        raise e
    finally:
        conn.close()

def _metadata_get_db():
    # Keep for schema initialization and compatibility
    conn = _sqlite3.connect(str(METADATA_DB_PATH), check_same_thread=False)
    conn.row_factory = _sqlite3.Row
    cur = conn.cursor()
    cur.executescript("""
        CREATE TABLE IF NOT EXISTS profiles (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            name            TEXT NOT NULL,
            source_file     TEXT NOT NULL,
            sheet_name      TEXT NOT NULL,
            file_hash       TEXT NOT NULL,
            imported_at     TEXT NOT NULL,
            point_count     INTEGER NOT NULL DEFAULT 0,
            column_map_json TEXT NOT NULL DEFAULT '{}',
            UNIQUE(file_hash, sheet_name)
        );
        CREATE TABLE IF NOT EXISTS io_points (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            profile_id      INTEGER NOT NULL REFERENCES profiles(id) ON DELETE CASCADE,
            point_id        TEXT NOT NULL,
            equipment_desc  TEXT,
            location        TEXT,
            attribute_desc  TEXT,
            station_code    TEXT,
            data_type       TEXT,
            alarm_list_desc TEXT,
            payload_json    TEXT NOT NULL DEFAULT '{}',
            states_json     TEXT NOT NULL DEFAULT '{}'
        );
        CREATE INDEX IF NOT EXISTS idx_io_points_profile ON io_points(profile_id);
        CREATE INDEX IF NOT EXISTS idx_io_points_pid     ON io_points(point_id);
    """)
    conn.commit()
    _migrate_columns(conn, "profiles", {"column_map_json": "TEXT NOT NULL DEFAULT '{}'"})
    _migrate_columns(conn, "io_points", {
        "point_id":        "TEXT NOT NULL DEFAULT ''",
        "equipment_desc":  "TEXT",
        "location":        "TEXT",
        "attribute_desc":  "TEXT",
        "station_code":    "TEXT",
        "data_type":       "TEXT",
        "alarm_list_desc": "TEXT",
        "payload_json":    "TEXT NOT NULL DEFAULT '{}'",
        "states_json":     "TEXT NOT NULL DEFAULT '{}'",
    })
    return conn


def _migrate_columns(conn, table, expected_cols):
    """Add any missing columns to `table`. Never removes or renames existing ones."""
    try:
        cur = conn.execute(f"PRAGMA table_info({table})")
        existing = {row["name"] for row in cur.fetchall()}
        for col, typedef in expected_cols.items():
            if col not in existing:
                conn.execute(f"ALTER TABLE {table} ADD COLUMN {col} {typedef}")
                logger.info(f"DB migration: added column '{col}' to '{table}'")
        conn.commit()
    except Exception as ex:
        logger.warning(f"DB migration failed for '{table}': {ex}")

def _metadata_file_hash(path):
    h = _hashlib.md5()
    try:
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(8192), b""):
                h.update(chunk)
    except Exception:
        pass
    return h.hexdigest()

def _metadata_save_profile(file_path, sheet_name, column_map, points):
    """Save IO list profile to DB. Returns (True, profile_name) on success, (False, error_str) on failure."""
    try:
        fhash = _metadata_file_hash(file_path)
        name = f"{pathlib.Path(file_path).stem} — {sheet_name}"
        now  = datetime.datetime.now().isoformat(timespec="seconds")
        col_json  = json.dumps({k: int(v) for k, v in column_map.items() if isinstance(v, int)})

        conn = _metadata_get_db()
        cur  = conn.cursor()
        cur.execute("""
            INSERT INTO profiles (name, source_file, sheet_name, file_hash, imported_at, point_count, column_map_json)
            VALUES (?,?,?,?,?,?,?)
            ON CONFLICT(file_hash, sheet_name) DO UPDATE SET
                name=excluded.name, imported_at=excluded.imported_at,
                point_count=excluded.point_count, column_map_json=excluded.column_map_json
        """, (name, str(file_path), sheet_name, fhash, now, len(points), col_json))
        conn.commit()

        cur.execute("SELECT id FROM profiles WHERE file_hash=? AND sheet_name=?", (fhash, sheet_name))
        row = cur.fetchone()
        if not row:
            conn.close()
            return False, "Profile row not found after insert"
        profile_id = row["id"]

        cur.execute("DELETE FROM io_points WHERE profile_id=?", (profile_id,))
        for pt in points:
            cur.execute("""
                INSERT INTO io_points
                    (profile_id, point_id, equipment_desc, location, attribute_desc,
                     station_code, data_type, alarm_list_desc, payload_json, states_json)
                VALUES (?,?,?,?,?,?,?,?,?,?)
            """, (
                profile_id, pt.get("point_id", ""), pt.get("equipment_desc", ""),
                pt.get("location", ""), pt.get("attribute_desc", ""), pt.get("station_code", ""),
                pt.get("data_type", ""), pt.get("alarm_list_desc", ""),
                json.dumps(pt.get("payload", {})), json.dumps({str(k): v for k, v in pt.get("states", {}).items()}),
            ))
        conn.commit()
        conn.close()
        logger.info(f"Metadata: saved profile '{name}' ({len(points)} points).")
        return True, name
    except Exception as ex:
        logger.warning(f"Metadata save failed: {ex}")
        return False, str(ex)

def _metadata_list_profiles():
    try:
        conn = _metadata_get_db()
        cur  = conn.cursor()
        cur.execute("SELECT id, name, sheet_name, imported_at, point_count, source_file FROM profiles ORDER BY imported_at DESC")
        rows = [dict(r) for r in cur.fetchall()]
        conn.close()
        return rows
    except Exception:
        return []

def _metadata_load_profile(profile_id):
    try:
        conn = _metadata_get_db()
        cur  = conn.cursor()
        cur.execute("SELECT * FROM profiles WHERE id=?", (profile_id,))
        prof = dict(cur.fetchone())
        cur.execute("SELECT * FROM io_points WHERE profile_id=?", (profile_id,))
        points = []
        for r in cur.fetchall():
            pt = dict(r)
            try: pt["payload"] = json.loads(pt.pop("payload_json", "{}"))
            except: pt["payload"] = {}
            try:
                raw_states = json.loads(pt.pop("states_json", "{}"))
                pt["states"] = {int(k): v for k, v in raw_states.items()}
            except: pt["states"] = {}
            # Restore top-level fields that the engine expects but may only be in payload
            if "protocol" not in pt:
                pt["protocol"] = pt["payload"].get("protocol", "MODBUS")
            if "severity" not in pt:
                pt["severity"] = pt["payload"].get("severity", 0)
            points.append(pt)
        conn.close()
        return prof, points
    except Exception as ex:
        logger.warning(f"Metadata load failed: {ex}")
        return None, []

def _metadata_delete_profile(profile_id):
    try:
        with db_session() as conn:
            conn.execute("DELETE FROM profiles WHERE id=?", (profile_id,))
        return True
    except Exception as e:
        logger.error(f"Failed to delete profile: {e}")
        return False

def detect_header_row(ws, max_scan=10):
    best_row_idx, best_headers, best_score = 1, [], 0
    all_aliases = [a for aliases in ISCS_ALIASES.values() for a in aliases]
    
    # Also score vN column patterns as valid header indicators
    import re as _re
    vn_pattern = _re.compile(r'v\d[_\s](?:label|severity|state)|(?:label|severity|state)[_\s]*\(\d\)')
    
    for row_idx, row in enumerate(ws.iter_rows(max_row=max_scan, values_only=True), 1):
        if not any(row): continue
        normalized = [_normalize(c) for c in row if c]
        score = 0
        for cell_text in normalized:
            for alias in all_aliases:
                if alias in cell_text or cell_text in alias:
                    score += 1
                    break
            # Also score vN columns so row 2 wins over row 1
            if vn_pattern.search(cell_text):
                score += 2
        if score > best_score:
            best_score, best_row_idx, best_headers = score, row_idx, list(row)
    return best_row_idx, best_headers

def auto_map_columns(headers):
    mapping, used_indices = {}, set()
    normalized = [_normalize(h) for h in headers]
    for field, aliases in ISCS_ALIASES.items():
        best_idx, best_score = None, 0
        for idx, norm_h in enumerate(normalized):
            if idx in used_indices or not norm_h: continue
            for alias in aliases:
                if alias == norm_h: score = 100
                elif norm_h.startswith(alias) or alias in norm_h: score = 50 + len(alias)
                else: score = 0
                if score > best_score:
                    best_score, best_idx = score, idx
        if best_idx is not None and best_score > 0:
            mapping[field] = best_idx
            used_indices.add(best_idx)
    return mapping

class SheetSelectorDialog(tk.Toplevel):
    def __init__(self, master, sheets):
        super().__init__(master)
        self.title("Select IO List Sheet")
        self.configure(bg="#0f0f0f")
        self.geometry("400x350")
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()
        self.result = None

        tk.Label(self, text="Select the sheet containing the IO list:", bg="#0f0f0f", fg="#cdd6f4", font=("Consolas", 11, "bold")).pack(pady=(15,5))
        suggested = [s for s in sheets if any(k in s.lower() for k in IO_SHEET_KEYWORDS)]
        if suggested: tk.Label(self, text=f"Suggested: {', '.join(suggested)}", bg="#0f0f0f", fg="#00C853", font=("Consolas", 9)).pack(pady=(0,10))

        self.listbox = tk.Listbox(self, bg="#1a1a1a", fg="#fff", font=("Consolas", 11), selectbackground="#2979FF", height=10)
        self.listbox.pack(fill="both", expand=True, padx=20)
        
        for i, sheet in enumerate(sheets):
            self.listbox.insert("end", sheet)
            if sheet in suggested: self.listbox.itemconfig(i, {'fg': '#00C853'})

        if suggested:
            idx = sheets.index(suggested[0])
            self.listbox.selection_set(idx)
            self.listbox.see(idx)
        else: self.listbox.selection_set(0)

        btn_f = tk.Frame(self, bg="#0f0f0f")
        btn_f.pack(pady=15)
        tk.Button(btn_f, text="Load Sheet", bg="#2979FF", fg="#fff", font=("Consolas", 10, "bold"), relief="flat", padx=15, command=self.accept).pack(side="left", padx=10)
        tk.Button(btn_f, text="Cancel", bg="#222", fg="#aaa", font=("Consolas", 10), relief="flat", padx=15, command=self.destroy).pack(side="left", padx=10)
        self.listbox.bind("<Double-Button-1>", lambda e: self.accept())

    def accept(self):
        sel = self.listbox.curselection()
        if sel:
            self.result = self.listbox.get(sel[0])
            self.destroy()

class ColumnMapperDialog(tk.Toplevel):
    def __init__(self, master, headers, auto_mapping):
        super().__init__(master)
        self.title("Verify Column Mapping")
        self.configure(bg="#0f0f0f")
        self.geometry("600x550")
        self.transient(master)
        self.grab_set()
        
        self.headers = [str(h) if h else "(empty)" for h in headers]
        self.auto_mapping = auto_mapping
        self.result = None
        self.combos = {}

        tk.Label(self, text="Verify that columns are correctly mapped.", bg="#0f0f0f", fg="#fff", font=("Consolas", 12, "bold")).pack(pady=(15,2))
        tk.Label(self, text="Green = auto-detected. Red = required but not found.", bg="#0f0f0f", fg="#888", font=("Consolas", 9)).pack(pady=(0,15))

        grid_canvas = tk.Canvas(self, bg="#0f0f0f", highlightthickness=0)
        sb = tk.Scrollbar(self, orient="vertical", command=grid_canvas.yview)
        grid_f = tk.Frame(grid_canvas, bg="#0f0f0f")
        
        grid_canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        grid_canvas.pack(fill="both", expand=True, padx=20)
        grid_canvas.create_window((0,0), window=grid_f, anchor="nw")
        grid_f.bind("<Configure>", lambda e: grid_canvas.configure(scrollregion=grid_canvas.bbox("all")))

        tk.Label(grid_f, text="Field", bg="#0f0f0f", fg="#89b4fa", font=("Consolas", 10, "bold")).grid(row=0, column=0, sticky="w", pady=5)
        tk.Label(grid_f, text="Maps to column", bg="#0f0f0f", fg="#89b4fa", font=("Consolas", 10, "bold")).grid(row=0, column=1, sticky="w", pady=5)
        tk.Label(grid_f, text="Status", bg="#0f0f0f", fg="#89b4fa", font=("Consolas", 10, "bold")).grid(row=0, column=2, sticky="w", pady=5)

        options = ["(not mapped)"] + self.headers

        for row_idx, field in enumerate(ISCS_ALIASES.keys(), 1):
            required = field in ISCS_REQUIRED
            flbl = field.replace("_", " ").title() + (" *" if required else "")
            tk.Label(grid_f, text=flbl, bg="#0f0f0f", fg="#f9e2af" if required else "#cdd6f4", font=("Consolas", 10)).grid(row=row_idx, column=0, sticky="w", pady=5)

            cb = ttk.Combobox(grid_f, values=options, state="readonly", width=35)
            self.combos[field] = cb
            cb.grid(row=row_idx, column=1, padx=10, pady=5)

            mapped_idx = self.auto_mapping.get(field)
            if mapped_idx is not None and mapped_idx < len(self.headers):
                cb.set(self.headers[mapped_idx])
                tk.Label(grid_f, text="✓ detected", bg="#0f0f0f", fg="#a6e3a1", font=("Consolas", 9)).grid(row=row_idx, column=2, sticky="w")
            else:
                cb.set(options[0])
                if required: tk.Label(grid_f, text="⚠ required", bg="#0f0f0f", fg="#f38ba8", font=("Consolas", 9)).grid(row=row_idx, column=2, sticky="w")
                else: tk.Label(grid_f, text="optional", bg="#0f0f0f", fg="#6c7086", font=("Consolas", 9)).grid(row=row_idx, column=2, sticky="w")

        btn_f = tk.Frame(self, bg="#0f0f0f")
        btn_f.pack(pady=20)
        tk.Button(btn_f, text="Confirm & Load", bg="#2979FF", fg="#fff", font=("Consolas", 10, "bold"), relief="flat", padx=15, command=self.accept).pack(side="left", padx=10)
        tk.Button(btn_f, text="Cancel", bg="#222", fg="#aaa", font=("Consolas", 10), relief="flat", padx=15, command=self.destroy).pack(side="left", padx=10)

    def accept(self):
        mapping = {}
        for field, cb in self.combos.items():
            val = cb.get()
            if val != "(not mapped)": mapping[field] = self.headers.index(val)
        
        for req in ISCS_REQUIRED:
            if req not in mapping:
                messagebox.showerror("Error", f"Required field '{req}' is not mapped!", parent=self)
                return
                
        self.result = mapping
        self.destroy()

# ── Protocol Plugin System (Falani Architecture) ──────────────────────────────
# M1.4: BaseProtocol promoted to the Hexagonal port (core/ports/protocol.py).
# Imported here as a shim so existing `ModbusProtocol(BaseProtocol)` is unchanged.
from core.ports.protocol import ProtocolPort, BaseProtocol

# M3.4: protocol layer relocated to adapters/driven/protocol/; re-exported as shims.
from adapters.driven.protocol.modbus import ModbusProtocol
from adapters.driven.protocol.manager import ProtocolManager

# ── Verify Result ─────────────────────────────────────────────────────────────
# M2.1: relocated to core/domain/results.py; re-exported here as a shim.
from core.domain.results import VerifyResult

# ── Report Generation Manager decoupled──────────────────────────────────────────────────
# ── Shared OCR helpers (module-level so ISCSVerifier and OcrMonitorPanel both use them) ──

def ocr_analyze_image(img, region=None):
    return iscs_OCR.analyze_image(img, region) if iscs_OCR else {}


def ocr_preprocess(img, config=None):
    return iscs_OCR.preprocess(img) if iscs_OCR else img


# In baru.py — Replace ocr_run with this updated design
def ocr_run(img, lang="eng", single_line=False, layout="tabular"):
    return iscs_OCR.run(img, lang, single_line, layout) if iscs_OCR else ""

# M2.4: ISCSVerifier relocated to core/services/verifier.py; re-exported as a shim.
from core.services.verifier import ISCSVerifier


# ── Monitor & Zone Models ─────────────────────────────────────────────────────
# M2.1: Monitor + Zone relocated to core/domain/; re-exported here as shims.
from core.domain.scenario import Monitor

def detect_monitors():
    monitors = []
    if SCREENINFO_AVAILABLE:
        try:
            raw_monitors = sorted(get_monitors(), key=lambda m: (m.x, m.y))
            for i, m in enumerate(raw_monitors): monitors.append(Monitor(i, m.x, m.y, m.width, m.height, m.name))
            return monitors
        except Exception: pass
    root = tk.Tk()
    root.withdraw()
    w, h = root.winfo_screenwidth(), root.winfo_screenheight()
    root.destroy()
    return [Monitor(0, 0, 0, w, h, "Primary")]

def get_physical_monitor_rects() -> dict:
    rects = []
    try:
        MonitorEnumProc = ctypes.WINFUNCTYPE(ctypes.c_int, ctypes.c_ulong, ctypes.c_ulong, ctypes.POINTER(ctypes.wintypes.RECT), ctypes.c_double)
        def _cb(hmon, hdc, lprect, lparam):
            r = lprect.contents
            rects.append((r.left, r.top, r.right - r.left, r.top + (r.bottom - r.top)))
            return 1
        ctypes.windll.user32.EnumDisplayMonitors(None, None, MonitorEnumProc(_cb), 0)
    except Exception: pass
    return rects

def match_physical_rect(monitor: "Monitor", phys_rects: list):
    if not phys_rects: return None
    return min(phys_rects, key=lambda r: abs(r[0] - monitor.x) + abs(r[1] - monitor.y))

from core.domain.zone import Zone

# M3.4: generate_points relocated to core/services/run_coordinator.py; shim.
from core.services.run_coordinator import generate_points

def zone_has_points(zone: Zone, monitor: Monitor, spacing: int):
    for y in range(monitor.y + spacing, monitor.y + monitor.height, spacing):
        for x in range(monitor.x + spacing, monitor.x + monitor.width, spacing):
            if zone.contains(x, y): return True
    return False

UNDO_LIMIT = 15  

# M2.1: Scenario relocated to core/domain/scenario.py; re-exported here as a shim.
from core.domain.scenario import Scenario


# ── Execution Engines ─────────────────────────────────────────────────────────
# M2.6: FailureEvidenceCollector relocated to core/services/evidence_collector.py; shim.
from core.services.evidence_collector import FailureEvidenceCollector

# M3.4: SuiteRunner relocated to core/services/run_coordinator.py; re-exported as a shim.
from core.services.run_coordinator import SuiteRunner

# M2.1: SuiteCard relocated to core/domain/scenario.py; re-exported here as a shim.
from core.domain.scenario import SuiteCard


class ISCS_Engine(threading.Thread):
    def __init__(self, card_or_points, zones_or_protocols, protocols_or_config,
                 config_or_log_dir=None, log_dir_or_run_id=None,
                 run_id_or_progress=None, on_progress_or_paused=None,
                 on_paused_or_done=None, on_done_or_log=None, on_log_or_suite=None,
                 suite_info=None, zones_per_page=None):
        """
        Supports two call signatures:

        Suite path (SuiteRunner / _add_current):
            ISCS_Engine(card, points, protocols, config, log_dir, run_id,
                        on_progress, on_paused, on_done, on_log, suite_info=…)

        Direct path (_run_test):
            ISCS_Engine(points, zones, protocols, config, log_dir,
                        on_progress, on_paused, on_done, on_log, zones_per_page=…)
        """
        super().__init__(name="ISCSEngineThread", daemon=True)

        if isinstance(card_or_points, SuiteCard):
            # ── Suite path ────────────────────────────────────────────────────
            self.card       = card_or_points
            self.points     = zones_or_protocols          # list of IO points
            self.protocols  = protocols_or_config         # ProtocolManager
            self.config     = config_or_log_dir           # dict
            self.log_dir    = log_dir_or_run_id           # Path
            self.run_id     = run_id_or_progress or 0     # int
            self.on_progress = on_progress_or_paused
            self.on_paused   = on_paused_or_done
            self.on_done     = on_done_or_log
            self.on_log      = on_log_or_suite
            self.suite_info  = suite_info
        else:
            # ── Direct path (_run_test passes: points, zones, protocols, config,
            #    log_dir, on_progress, on_paused, on_done, on_log) ─────────────
            points    = card_or_points            # list of IO points
            zones     = zones_or_protocols        # list of Zone
            protocols = protocols_or_config       # ProtocolManager
            config    = config_or_log_dir         # dict
            log_dir   = log_dir_or_run_id         # Path
            # remaining positional args shift left by one (no run_id in direct path)
            on_progress = run_id_or_progress
            on_paused   = on_progress_or_paused
            on_done     = on_paused_or_done
            on_log      = on_done_or_log

            self.card = SuiteCard.from_direct(
                name="Direct Run",
                zones=zones,
                protocol="MODBUS",
                zones_per_page=zones_per_page or {},
            )
            self.points     = points
            self.protocols  = protocols
            self.config     = config
            self.log_dir    = log_dir
            self.run_id     = 0
            self.on_progress = on_progress
            self.on_paused   = on_paused
            self.on_done     = on_done
            self.on_log      = on_log
            self.suite_info  = None

        self._stop_event = threading.Event()
        self._pause_event = threading.Event(); self._pause_event.set()
        self.results, self.current_idx, self._pause_reason = [], 0, ""
        self.active_samplers = []  # Added sampler tracking

    def stop(self):  
        self._stop_event.set()
        self._pause_event.set()
        # Abort all running samplers instantly
        for s in list(self.active_samplers):
            try: s.stop()
            except: pass
    def pause(self, r="manual"): self._pause_reason=r; self._pause_event.clear()
    def resume(self): self._pause_reason=""; self._pause_event.set()
    @property
    def is_paused(self): return not self._pause_event.is_set()

    def _sleep(self, seconds: float, granularity: float = 0.1):
        """Interruptible sleep — returns early if stop is requested."""
        deadline = time.monotonic() + seconds
        while time.monotonic() < deadline:
            if self._stop_event.is_set():
                return
            time.sleep(min(granularity, deadline - time.monotonic()))

    def _check_pause(self, idx, total):
        """Helper to block execution and alert UI if paused (manual or drift)."""
        if not self._pause_event.is_set():
            self.on_paused(idx, total, self._pause_reason)
            self._pause_event.wait(timeout=0.2)

    def _nav_click(self, x: int, y: int, label: str, idx: int, total: int):
        if not PYAUTOGUI_AVAILABLE or (x == 0 and y == 0): return
        try: 
            pyautogui.click(x, y)
            time.sleep(0.06)
            
            # Mouse Drift Safety Check
            ax, ay = pyautogui.position()
            drift_px = self.config.get("mouse_drift_px", 15)
            if abs(ax - x) + abs(ay - y) > drift_px:
                self.pause("mouse moved")
                self._check_pause(idx, total)
                
            time.sleep(self.config.get("nav_wait_sec", 1.5))
            self.on_log(f"  Nav: clicked {label}")
        except Exception as e:
            self.on_log(f"  Nav Error: {e}")

    def _build_verifier_zones(self, page_name: str = None) -> dict:
        """
        Build zones dict for ISCSVerifier.
        If page_name given and zones_per_page has that page, use per-page zones.
        Falls back to zones_per_page["Global"], then flat self.card.zones.
        """
        zone_types = ["alarm_panel", "equipment_page", "alarm_list", "event_list"]
        zpp = self.card.zones_per_page

        # Try page-specific zones first
        if page_name and page_name in zpp:
            page_zones = zpp[page_name]
            # Fill any missing zone types from Global
            global_zones = zpp.get("Global", {})
            return {key: page_zones.get(key) or global_zones.get(key) or
                    next((z for z in self.card.zones if z.zone_type == key), None)
                    for key in zone_types}

        # Try Global zones
        if "Global" in zpp:
            global_zones = zpp["Global"]
            return {key: global_zones.get(key) or
                    next((z for z in self.card.zones if z.zone_type == key), None)
                    for key in zone_types}

        # Fallback: flat zones list (old behaviour, backward compat)
        return {key: next((z for z in self.card.zones if z.zone_type == key), None)
                for key in zone_types}

    def _take_screenshot(self, prefix: str, zone=None, full_screen: bool = False) -> str:
        """
        Save a screenshot and return the file path string.
        Naming convention mirrors ClickEngine: {idx:04d}_{prefix}.png
        If zone is given, crops to that zone; otherwise uses full active monitor.
        """
        if not PIL_AVAILABLE:
            return ""
        try:
            import __main__
            mon = getattr(__main__.app, 'active_mon', None)
            if full_screen or zone is None:
                if mon:
                    bbox = (mon.x, mon.y, mon.x + mon.width, mon.y + mon.height)
                else:
                    bbox = None
            else:
                bbox = (zone.x1, zone.y1, zone.x2, zone.y2)

            img = ImageGrab.grab(bbox=bbox, all_screens=True) if bbox else ImageGrab.grab(all_screens=True)
            fname = f"{self.current_idx:04d}_{prefix}.png"
            path  = self.session_dir / fname
            img.save(str(path))
            return str(path)
        except Exception as e:
            logger.warning(f"ISCS_Engine._take_screenshot failed ({prefix}): {e}")
            return ""

    def run(self):
        try:
            self.log_dir.mkdir(parents=True, exist_ok=True)
            sid = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            self.session_dir = self.log_dir / f"iscs_{self.card.name.replace(' ','_')}_{sid}"
            self.session_dir.mkdir(parents=True, exist_ok=True)

            # Feature 1: load AnchorManager for this session
            _anchor_mgr = None
            if UPGRADES_AVAILABLE:
                _anchor_mgr = AnchorManager(self.session_dir)
                _anchor_mgr.load()

            verifier = ISCSVerifier(self._build_verifier_zones(), self.config, anchor_mgr=_anchor_mgr, stop_event=self._stop_event)
            total = len(self.points)
            current_page_name = None  # track which page we're on
            # Track execution start
            start_time = datetime.datetime.now()
            
            import __main__
            app_ref = __main__.app
            
            def init_hud():
                app_ref.hud = HudOverlay(app_ref, total, app_ref.active_mon, name=self.card.name)
                if self.suite_info:
                    app_ref.hud.set_suite_info(
                        self.suite_info["card_idx"], self.suite_info["total_cards"],
                        self.suite_info["pass_num"], self.suite_info["loop_count"], self.card.name
                    )
            app_ref.after(0, init_hud)

            # Heartbeat screenshot at start — same as ClickEngine
            self._take_screenshot("heartbeat_START", full_screen=True)
            last_heartbeat = time.time()

            for i in range(self.current_idx, total):
                if self._stop_event.is_set(): break
                self._check_pause(i, total)

                self.current_idx = i
                pt = self.points[i]
                identifier = pt.get('point_id', pt.get('identifier', f'pt_{i}'))
                self.on_log(f"[{i+1}/{total}] Testing: {identifier}")
                point_results = []
                point_pass = True

                # Derive trigger/reset value indices and build expected states early
                trigger_idx, reset_idx = _get_state_indices(pt)
                expected_alarm = build_expected(pt, trigger_idx)
                expected_norm  = build_expected(pt, reset_idx)

                self.on_progress(i+1, total, identifier, "…")

                # Resolve bbox once before triggering
                alarm_zone = verifier.alarm_zone
                resolved_bbox = (alarm_zone.x1, alarm_zone.y1, alarm_zone.x2, alarm_zone.y2) if alarm_zone else None
                if alarm_zone and verifier.anchor_mgr:
                    resolved = verifier.anchor_mgr.resolve("alarm_panel")
                    if resolved:
                        resolved_bbox = resolved

                try:
                    handler = self.protocols.get_protocol(self.card.protocol)

                    # Start Sampler
                    sampler = None
                    trigger_ns = None
                    if UPGRADES_AVAILABLE and resolved_bbox:
                        _dur_sec = float(self.config.get("detection_duration_sec", 8.0))
                        _int_ms  = int(self.config.get("sampler_interval_ms", 100))
                        sampler  = FrameSampler(resolved_bbox, duration_sec=_dur_sec, interval_ms=_int_ms)
                        self.active_samplers.append(sampler)  # Track sampler
                        sampler.start()

                    handler.trigger_alarm(pt)
                    trigger_time = datetime.datetime.now()
                    trigger_ns   = time.time_ns()

                    # ── Check pause right after trigger ───────────────────────
                    self._check_pause(i, total)

                    # No premature `.join()` here. Let it run concurrently!

                    # ── Verify trigger state ──────────────────────────────────
                    panel_res = verifier.verify_alarm_panel(expected_alarm, self.session_dir,
                                                            point_idx=i, trigger_time=trigger_time, 
                                                            file_suffix="alarm_panel_trigger",
                                                            sampler=sampler, trigger_ns=trigger_ns)
                    point_results.extend(panel_res)
                    if any(r.status == "FAIL" for r in panel_res): point_pass = False
                    
                    # LOG TRIGGER PANEL RESULTS IMMEDIATELY
                    self.on_log("    ┌─ TRIGGER RESULTS ─────────────────")
                    for r in panel_res:
                        icon  = "✓" if r.status == "PASS" else ("–" if r.status == "SKIP" else "✗")
                        field = r.step.split("/")[-1].upper().ljust(12)
                        self.on_log(f"    │ [{icon}] {field} {r.msg}")
                    self.on_log("    └" + "─" * 40)
                    if sampler and sampler in self.active_samplers:
                        self.active_samplers.remove(sampler)

                except Exception as ex:
                    point_pass = False
                    self.on_log(f"Error testing {identifier}: {ex}")
                    logger.error(f"ISCS_Engine point error: {ex}\n{traceback.format_exc()}")

                    try:
                        # ── Right-click → Equipment Page (optional) ──────────────
                        rc_x = self.card.rightclick_row1_x
                        rc_y = self.card.rightclick_row1_y
                        pg_x = self.card.rightclick_page_btn_x
                        pg_y = self.card.rightclick_page_btn_y
                        eq_zone = verifier.zones.get("equipment_page")
    
                        if rc_x != 0 and rc_y != 0 and pg_x != 0 and pg_y != 0:
                            self.on_log("  [Nav] Right-clicking alarm row 1 → navigating to equipment page...")
                            if PYAUTOGUI_AVAILABLE:
                                click_delay = self.config.get("click_delay", 1.5)
                                pyautogui.rightClick(rc_x, rc_y)
                                self._sleep(click_delay)
                                pyautogui.click(pg_x, pg_y)
                                time.sleep(self.config.get("nav_wait_sec", 1.0))
    
                            if eq_zone is not None:
                                img, abs_pos, conf = verifier.find_equipment_on_page(identifier)
                                if abs_pos:
                                    if PYAUTOGUI_AVAILABLE:
                                        cx, cy = abs_pos[0], abs_pos[1]
                                        pyautogui.click(cx, cy)
                                        time.sleep(self.config.get("inspector_wait_sec", 1.0))
                                    insp_results = verifier.verify_inspector(expected_alarm, eq_zone, self.session_dir, point_idx=i)
                                    point_results.extend(insp_results)
                                    if any(r.status == "FAIL" for r in insp_results):
                                        point_pass = False
                                else:
                                    self.on_log(f"  [✗] Equipment '{identifier}' not found on equipment page.")
                                    point_results.append(VerifyResult("equipment_page", "FAIL", f"Equipment '{identifier}' not found on page."))
                                    point_pass = False
                            else:
                                self.on_log("  [Skip] No equipment_page zone drawn. Skipping equipment verify.")
                                point_results.append(VerifyResult("equipment_page", "SKIP", "No equipment_page zone drawn."))
    
                            # Navigate home before alarm list
                            if self.card.home_x != 0 and self.card.home_y != 0:
                                self._nav_click(self.card.home_x, self.card.home_y, "home", i, total)
                        else:
                            point_results.append(VerifyResult("equipment_page", "SKIP", "Right-click coords not configured."))
    
                        # ── Alarm list verification (optional) ───────────────────
                        if self.card.alarm_list_x != 0 and self.card.alarm_list_y != 0:
                            self.on_log("  [Nav] Navigating to Alarm List...")
                            self._nav_click(self.card.alarm_list_x, self.card.alarm_list_y, "alarm list", i, total)
                            list_zone = verifier.zones.get("alarm_list")
                            if list_zone is not None:
                                _al_bbox = (list_zone.x1, list_zone.y1, list_zone.x2, list_zone.y2)
                                _al_sampler = FrameSampler(_al_bbox,
                                    duration_sec=float(self.config.get("sampler_duration_sec", 2.0)),
                                    interval_ms=int(self.config.get("sampler_interval_ms", 100))
                                ) if UPGRADES_AVAILABLE else None
                                if _al_sampler: _al_sampler.start()
                                if _al_sampler: _al_sampler.join(timeout=float(self.config.get("sampler_duration_sec", 2.0)) + 0.5)
                                _al_ns = time.time_ns()
                                list_results = verifier.verify_list("alarm_list", expected_alarm, list_zone, self.session_dir, point_idx=i, sampler=_al_sampler, trigger_ns=_al_ns)
                                point_results.extend(list_results)
                                if any(r.status == "FAIL" for r in list_results):
                                    point_pass = False
                            else:
                                self.on_log("  [Skip] 'alarm_list' zone not drawn. Skipping Alarm List verification.")
                                point_results.append(VerifyResult("alarm_list", "SKIP", "No alarm_list zone drawn."))

                        # ── Event list verification (optional) ───────────────────
                        if self.card.event_list_x != 0 and self.card.event_list_y != 0:
                            self.on_log("  [Nav] Navigating to Event List...")
                            self._nav_click(self.card.event_list_x, self.card.event_list_y, "event list", i, total)
                            event_zone = verifier.zones.get("event_list")
                            if event_zone is not None:
                                _ev_bbox = (event_zone.x1, event_zone.y1, event_zone.x2, event_zone.y2)
                                _ev_sampler = FrameSampler(_ev_bbox,
                                    duration_sec=float(self.config.get("sampler_duration_sec", 2.0)),
                                    interval_ms=int(self.config.get("sampler_interval_ms", 100))
                                ) if UPGRADES_AVAILABLE else None
                                if _ev_sampler: _ev_sampler.start()
                                if _ev_sampler: _ev_sampler.join(timeout=float(self.config.get("sampler_duration_sec", 2.0)) + 0.5)
                                _ev_ns = time.time_ns()
                                event_results = verifier.verify_list("event_list", expected_alarm, event_zone, self.session_dir, point_idx=i, sampler=_ev_sampler, trigger_ns=_ev_ns)
                                point_results.extend(event_results)
                                if any(r.status == "FAIL" for r in event_results):
                                    point_pass = False
                            else:
                                self.on_log("  [Skip] 'event_list' zone not drawn. Skipping Event List verification.")
                                point_results.append(VerifyResult("event_list", "SKIP", "No event_list zone drawn."))
    
                        if self.card.home_x != 0 and self.card.home_y != 0:
                            self._nav_click(self.card.home_x, self.card.home_y, "home", i, total)
    
                        if any(r.status == "FAIL" for r in point_results): point_pass = False

                    except Exception as ex:
                        point_pass = False
                        self.on_log(f"Error testing {identifier}: {ex}")
                        logger.error(f"ISCS_Engine point error: {ex}\n{traceback.format_exc()}")

                                # ── Reset (normalize) — always run ────────────────────────────
                reset_ok = False
                norm_sampler = None
                reset_ns = None

                if UPGRADES_AVAILABLE and resolved_bbox:
                    _dur_sec = float(self.config.get("sampler_duration_sec", 2.0))
                    _int_ms  = int(self.config.get("sampler_interval_ms", 100))
                    norm_sampler = FrameSampler(resolved_bbox, duration_sec=_dur_sec, interval_ms=_int_ms)
                    self.active_samplers.append(norm_sampler)
                    norm_sampler.start()

                try:
                    handler.reset_alarm(pt)
                    reset_ns = time.time_ns()
                    reset_ok = True
                except Exception:
                    pass

                if reset_ok and not self._stop_event.is_set():
                    # Wait for normalize sampler to finish recording
                    if norm_sampler:
                        if self._stop_event.is_set():
                            norm_sampler.stop()
                        norm_sampler.join(timeout=float(self.config.get("sampler_duration_sec", 2.0)) + 0.5)
                        if norm_sampler in self.active_samplers:
                            self.active_samplers.remove(norm_sampler)

                    # ── Verify reset state ────────────────────────
                    norm_res = verifier.verify_alarm_panel(
                        expected_norm, self.session_dir, point_idx=i, trigger_time=None, 
                        file_suffix="alarm_panel_normalize", sampler=norm_sampler, trigger_ns=reset_ns
                    )
                    for r in norm_res:
                        r.step = r.step.replace("alarm_panel/", "normalize/")
                    point_results.extend(norm_res)
                    if any(r.status == "FAIL" for r in norm_res):
                        point_pass = False
                        
                    # LOG NORMALIZE RESULTS IMMEDIATELY
                    self.on_log("    ┌─ NORMALIZE RESULTS ───────────────")
                    for r in norm_res:
                        icon  = "✓" if r.status == "PASS" else ("–" if r.status == "SKIP" else "✗")
                        field = r.step.split("/")[-1].upper().ljust(12)
                        self.on_log(f"    │ [{icon}] {field} {r.msg}")
                    self.on_log("    └" + "─" * 40)

                # Periodic heartbeat screenshot — same as ClickEngine
                if time.time() - last_heartbeat > HEARTBEAT_SEC:
                    self._take_screenshot("heartbeat_15MIN", full_screen=True)
                    last_heartbeat = time.time()

                overall = "PASS" if point_pass else "FAIL"

                diag_data = None
                if not point_pass:
                    try:
                        diag_data = FailureEvidenceCollector.collect(
                            session_dir=self.session_dir,
                            point_idx=i,
                            pt=pt,
                            point_results=point_results,
                            verifier=verifier,
                            trigger_time=trigger_time if 'trigger_time' in locals() else None,
                            expected_alarm=expected_alarm if 'expected_alarm' in locals() else {},
                            config=self.config,
                            reset_time=datetime.datetime.fromtimestamp(reset_ns / 1e9) if 'reset_ns' in locals() and reset_ns else None,
                            expected_norm=expected_norm if 'expected_norm' in locals() else None,
                        )
                    except Exception as fe:
                        logger.warning(f"FailureEvidenceCollector hook failed: {fe}")

                self.on_log(f"  [★] Point {identifier} execution: {overall}")
                self.on_log("  " + "═" * 50)
                for r in point_results:
                    icon  = "✓" if r.status == "PASS" else ("–" if r.status == "SKIP" else "✗")
                    phase = "TRIGGER  " if r.step.startswith("alarm_panel") else "NORMALIZE"
                    field = r.step.split("/")[-1].upper().ljust(12)
                    self.on_log(f"  │ [{icon}] {phase} {field} {r.msg}")
                self.on_log(f"  └{'─' * 40}")
                self.results.append({
                    "identifier": identifier, 
                    "overall": overall, 
                    "steps": [r.to_dict() for r in point_results],
                    "failure_diagnostics": diag_data
                })
                self.on_progress(i+1, total, identifier, overall)
                
            stopped = self._stop_event.is_set()
            self._take_screenshot("heartbeat_STOPPED" if stopped else "heartbeat_END", full_screen=True)

            # Track execution end
            end_time = datetime.datetime.now()

            if PANDAS_AVAILABLE: pd.DataFrame(self.results).to_csv(self.session_dir / "Report.csv", index=False)
            # Always write JSON alongside CSV for easy debugging
            with open(self.session_dir / "results.json", "w") as f:
                json.dump(self.results, f, indent=2)

            # Generate advanced visual and spreadsheet metrics — always runs even if stopped
            if stopped:
                self.on_log("  Test stopped — generating partial report for completed points...")
            if ReportManager is not None:
                try:
                    ReportManager.generate_reports(self.results, self.session_dir, start_time, end_time, title=self.card.name)
                except Exception as re:
                    self.on_log(f"Warning: Failed to generate reports: {re}")
                    logger.error(f"Report generation error: {re}", exc_info=True)
            else:
                self.on_log("Warning: ReportManager is unavailable. Skipping HTML/Excel reports.")

            self.on_done(self.session_dir, "", stopped)

        except Exception as e:
            logger.error(traceback.format_exc())
            self.on_done(getattr(self, 'session_dir', LOG_DIR), "", True)

class ClickEngine(threading.Thread):
    def __init__(self, mode, points, log_dir, delay, monitor, on_progress, on_paused, on_done):
        super().__init__(daemon=True)
        self.mode = mode
        self.points = points
        self.log_dir = log_dir
        self.delay = delay
        self.monitor = monitor
        self.on_progress = on_progress
        self.on_paused = on_paused
        self.on_done = on_done

        self._stop_event = threading.Event()
        self.name = "ClickEngineThread"
        self._pause_event = threading.Event(); self._pause_event.set()
        self.results = []
        self._pause_reason = ""
        self.session_dir = None
        self.current_idx = 0

    def stop(self): self._stop_event.set(); self._pause_event.set()
    def pause(self, reason="manual"): self._pause_reason = reason; self._pause_event.clear()
    def resume(self): self._pause_reason = ""; self._pause_event.set()
    @property
    def is_paused(self): return not self._pause_event.is_set()

    def _take_screenshot(self, prefix, pt_data=None, full_screen=False):
        if not PIL_AVAILABLE: return ""
        ss_name = f"{self.current_idx:04d}_{prefix}.png"
        ss_path = self.session_dir / ss_name

        if full_screen:
            bbox = (self.monitor.x, self.monitor.y, self.monitor.x + self.monitor.width, self.monitor.y + self.monitor.height)
        else:
            if self.mode == "sequence" and pt_data and pt_data.get("zone"):
                z = pt_data["zone"]
                x1 = max(self.monitor.x, z.x1 - WIDE_CROP_PAD)
                y1 = max(self.monitor.y, z.y1 - WIDE_CROP_PAD)
                x2 = min(self.monitor.x + self.monitor.width, z.x2 + WIDE_CROP_PAD)
                y2 = min(self.monitor.y + self.monitor.height, z.y2 + WIDE_CROP_PAD)
                bbox = (x1, y1, x2, y2)
            else:
                bbox = (self.monitor.x, self.monitor.y, self.monitor.x + self.monitor.width, self.monitor.y + self.monitor.height)

        try:
            img = ImageGrab.grab(bbox=bbox, all_screens=True)
            if pt_data:
                draw = ImageDraw.Draw(img)
                lx = pt_data["x"] - bbox[0]
                ly = pt_data["y"] - bbox[1]
                r = 8
                draw.ellipse((lx - r, ly - r, lx + r, ly + r), outline="#FF1744", width=3)
                draw.line((lx, ly - r - 6, lx, ly + r + 6), fill="#FF1744", width=3)
                draw.line((lx - r - 6, ly, lx + r + 6, ly), fill="#FF1744", width=3)
            img.save(str(ss_path))
            return str(ss_path)
        except Exception:
            return ""

    def run(self):
        try:
            logger.info("ClickEngine worker started.")
            self.log_dir.mkdir(parents=True, exist_ok=True)
            sid  = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            self.session_dir = self.log_dir / sid
            self.session_dir.mkdir()
            total = len(self.points)

            self._take_screenshot("heartbeat_START", full_screen=True)
            last_heartbeat = time.time()

            for i in range(self.current_idx, total):
                if self._stop_event.is_set(): break
                
                if not self._pause_event.is_set():
                    reason_clean = self._pause_reason.replace(" ", "_")
                    self._take_screenshot(f"heartbeat_PAUSE_{reason_clean}", full_screen=True)
                    self.on_paused(i, total, self._pause_reason)
                    self._pause_event.wait(timeout=0.2)
                    last_heartbeat = time.time() 
                    if self._stop_event.is_set(): break

                self.current_idx = i
                pt = self.points[i]
                x, y = pt["x"], pt["y"]
                label = pt["label"]
                result = {"x": x, "y": y, "label": label, "status": "ok", "screenshot": ""}
                
                try:
                    pyautogui.click(x, y)
                    time.sleep(0.06)
                    ax, ay = pyautogui.position()
                    if abs(ax - x) + abs(ay - y) > MOUSE_DRIFT_PX:
                        self.pause("mouse moved")
                    time.sleep(SCREENSHOT_DELAY)
                    result["screenshot"] = self._take_screenshot(f"click_{label}", pt, full_screen=False)
                except Exception as ex:
                    result["status"] = f"error: {ex}"
                    logger.error(f"Click logic error at point {i}: {ex}")

                self.results.append(result)
                self.on_progress(i + 1, total, x, y)
                self.current_idx += 1
                
                if time.time() - last_heartbeat > HEARTBEAT_SEC:
                    self._take_screenshot("heartbeat_15MIN", full_screen=True)
                    last_heartbeat = time.time()

                rem = self.delay - SCREENSHOT_DELAY - 0.06
                if rem > 0: time.sleep(rem)

            tag = "heartbeat_STOPPED" if self._stop_event.is_set() else "heartbeat_END"
            self._take_screenshot(tag, full_screen=True)

            log_path = self.session_dir / "results.json"
            with open(log_path, "w") as f:
                json.dump(self.results, f, indent=2)
            self.on_done(self.session_dir, log_path, self._stop_event.is_set())
            logger.info("ClickEngine worker finished.")
        except Exception:
            logger.error("CRITICAL: ClickEngine crashed!")
            logger.error(traceback.format_exc())

# ── Recording Settings Dialog ─────────────────────────────────────────────────
class RecordingSettingsDialog:
    """
    Small modal dialog for configuring recording options.
    Opened by the ⚙ button next to the recording toggle in the scenarios header.
    """

    @classmethod
    def show(cls, parent, current_settings) -> "RecorderSettings | None":
        if RecorderSettings is None:
            import tkinter.messagebox as mb
            mb.showwarning("Recording Unavailable",
                           "iscs_recorder.py is missing.\n"
                           "pip install imageio imageio-ffmpeg", parent=parent)
            return None

        dlg = tk.Toplevel(parent)
        dlg.title("Recording Settings")
        dlg.configure(bg="#0f0f0f")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.attributes("-topmost", True)

        # Centre over parent
        px = parent.winfo_rootx() + (parent.winfo_width()  - 380) // 2
        py = parent.winfo_rooty() + (parent.winfo_height() - 420) // 2
        dlg.geometry(f"380x420+{px}+{py}")

        cfg = current_settings if current_settings else RecorderSettings()
        result = [None]

        # ── Bulletproof Reference & Attribute Resolution ─────────────────────
        if hasattr(parent, "monitors"):
            app_ref = parent
        elif hasattr(parent, "app") and hasattr(parent.app, "monitors"):
            app_ref = parent.app
        else:
            app_ref = parent

        monitors_list = getattr(app_ref, "monitors", [])

        cfg_fps             = getattr(cfg, "fps", 5)
        cfg_display         = getattr(cfg, "capture_display", "Auto")
        cfg_resolution      = getattr(cfg, "capture_resolution", "Native (recommended)")
        cfg_show_timestamp  = getattr(cfg, "show_timestamp", True)
        cfg_show_remark     = getattr(cfg, "show_remark", True)
        cfg_warn_threshold  = getattr(cfg, "warn_threshold_gb", 2.0)

        BG  = "#0f0f0f"
        SBG = "#1a1a1a"
        FG  = "#ccc"
        ACT = "#2979FF"
        LBL = dict(bg=BG, fg=FG, font=("Consolas", 9))
        ENT = dict(bg=SBG, fg="#fff", insertbackground="#fff",
                   font=("Consolas", 10), relief="flat")

        pad = dict(padx=14, pady=5)

        # ── FPS ──────────────────────────────────────────────────────────────
        fps_frame = tk.Frame(dlg, bg=BG)
        fps_frame.pack(fill="x", **pad)
        tk.Label(fps_frame, text="Frame rate (fps):", **LBL).pack(side="left")
        fps_var = tk.StringVar(value=str(cfg_fps))
        
        fps_options = [str(f) for f in FPS_OPTIONS]
        fps_menu = tk.OptionMenu(fps_frame, fps_var, *fps_options)
        fps_menu.config(bg="#1a1a1a", fg="#fff", activebackground="#2979FF", activeforeground="#fff",
                        font=("Consolas", 10), relief="flat", bd=0, highlightthickness=0)
        fps_menu["menu"].config(bg="#1a1a1a", fg="#fff", font=("Consolas", 10), bd=0)
        fps_menu.pack(side="right")

        # fps hint label
        hint_var = tk.StringVar()
        _fps_hints = {
            "1":  "Minimal — alarm state capture only",
            "5":  "Recommended — smooth enough, light file",
            "10": "Good for client demos / audit",
            "15": "Lowest that feels like video",
            "24": "Film standard — overkill for SCADA",
            "30": "Broadcast standard — large files",
            "60": "Max — very large files, no SCADA benefit",
        }
        def _update_hint(*_):
            hint_var.set(_fps_hints.get(fps_var.get(), ""))
        fps_var.trace_add("write", _update_hint)
        _update_hint()
        tk.Label(dlg, textvariable=hint_var, bg=BG, fg="#666",
                 font=("Consolas", 8), anchor="e").pack(fill="x", padx=14)

        tk.Frame(dlg, bg="#2a2a2a", height=1).pack(fill="x", padx=14, pady=6)

        # ── Capture Display ──────────────────────────────────────────────────
        display_frame = tk.Frame(dlg, bg=BG)
        display_frame.pack(fill="x", **pad)
        tk.Label(display_frame, text="Capture Display:", **LBL).pack(side="left")
        
        display_options = ["Auto (follow scenario card)"] + [mon.label for mon in monitors_list]
        
        selected_disp = "Auto (follow scenario card)"
        if cfg_display != "Auto":
            for opt in display_options:
                if cfg_display in opt:
                    selected_disp = opt
                    break

        display_var = tk.StringVar(value=selected_disp)
        
        display_menu = tk.OptionMenu(display_frame, display_var, *display_options)
        display_menu.config(bg="#1a1a1a", fg="#fff", activebackground="#2979FF", activeforeground="#fff",
                            font=("Consolas", 10), relief="flat", bd=0, highlightthickness=0)
        display_menu["menu"].config(bg="#1a1a1a", fg="#fff", font=("Consolas", 10), bd=0)
        display_menu.pack(side="right")

        # ── Capture Resolution ───────────────────────────────────────────────
        resolution_frame = tk.Frame(dlg, bg=BG)
        resolution_frame.pack(fill="x", **pad)
        tk.Label(resolution_frame, text="Capture Resolution:", **LBL).pack(side="left")
        
        resolution_var = tk.StringVar(value=cfg_resolution)
        
        resolution_menu = tk.OptionMenu(resolution_frame, resolution_var, "Native (recommended)")
        resolution_menu.config(bg="#1a1a1a", fg="#fff", activebackground="#2979FF", activeforeground="#fff",
                               font=("Consolas", 10), relief="flat", bd=0, highlightthickness=0)
        resolution_menu["menu"].config(bg="#1a1a1a", fg="#fff", font=("Consolas", 10), bd=0)
        resolution_menu.pack(side="right")

        def update_resolution_options(*args):
            disp = display_var.get()
            
            selected_mon = None
            for mon in monitors_list:
                if mon.label == disp:
                    selected_mon = mon
                    break
            
            if selected_mon:
                mw, mh = selected_mon.width, selected_mon.height
                filtered_opts = ["Native (recommended)"]
                for opt in ["1280x720", "1024x768", "854x480"]:
                    w, h = map(int, opt.split('x'))
                    if w <= mw and h <= mh:
                        filtered_opts.append(opt)
            else:
                filtered_opts = ["Native (recommended)", "1280x720", "1024x768", "854x480"]
                
            menu = resolution_menu["menu"]
            menu.delete(0, "end")
            for opt in filtered_opts:
                menu.add_command(label=opt, command=lambda o=opt: resolution_var.set(o))
                
            if resolution_var.get() not in filtered_opts:
                resolution_var.set("Native (recommended)")

        display_var.trace_add("write", update_resolution_options)
        update_resolution_options()

        tk.Frame(dlg, bg="#2a2a2a", height=1).pack(fill="x", padx=14, pady=6)

        # ── Overlay Toggles ───────────────────────────────────────────────────
        ts_var  = tk.BooleanVar(value=cfg_show_timestamp)
        rmk_var = tk.BooleanVar(value=cfg_show_remark)

        ck_style = dict(bg=BG, fg=FG, selectcolor=SBG,
                        activebackground=BG, font=("Consolas", 9))
        tk.Checkbutton(dlg, text="Burn timestamp overlay onto frames",
                       variable=ts_var, **ck_style).pack(anchor="w", padx=14)
        tk.Checkbutton(dlg, text="Burn point remark overlay onto frames",
                       variable=rmk_var, **ck_style).pack(anchor="w", padx=14)

        tk.Frame(dlg, bg="#2a2a2a", height=1).pack(fill="x", padx=14, pady=6)

        # ── Storage warning threshold ─────────────────────────────────────────
        warn_frame = tk.Frame(dlg, bg=BG)
        warn_frame.pack(fill="x", padx=14, pady=(0, 6))
        tk.Label(warn_frame, text="Warn if free disk <", **LBL).pack(side="left")
        warn_var = tk.StringVar(value=str(cfg_warn_threshold))
        tk.Entry(warn_frame, textvariable=warn_var, width=5, **ENT).pack(side="left", padx=4)
        tk.Label(warn_frame, text="GB", **LBL).pack(side="left")

        # ── OK / Cancel ────────────────────────────────────────────────────────
        btn_f = tk.Frame(dlg, bg=BG)
        btn_f.pack(fill="x", padx=14, pady=(0, 12))

        def _ok():
            try:
                fps = int(fps_var.get())
                if fps not in FPS_OPTIONS:
                    fps = DEFAULT_FPS
            except ValueError:
                fps = DEFAULT_FPS
            try:
                warn = float(warn_var.get())
            except ValueError:
                warn = 2.0

            disp_val = "Auto"
            selected_disp_str = display_var.get()
            if selected_disp_str != "Auto (follow scenario card)":
                for mon in monitors_list:
                    if mon.label == selected_disp_str:
                        disp_val = mon.label
                        break

            s = RecorderSettings(
                enabled          = getattr(cfg, "enabled", False),
                fps              = fps,
                show_timestamp   = ts_var.get(),
                show_remark      = rmk_var.get(),
                warn_threshold_gb = warn,
                capture_display  = disp_val,
                capture_resolution = resolution_var.get()
            )
            result[0] = s
            dlg.destroy()

        def _cancel():
            dlg.destroy()

        tk.Button(btn_f, text="Save", bg=ACT, fg="#fff",
                  font=("Consolas", 9, "bold"), relief="flat",
                  padx=12, pady=4, cursor="hand2",
                  command=_ok).pack(side="right", padx=(4, 0))
        tk.Button(btn_f, text="Cancel", bg="#333", fg="#ccc",
                  font=("Consolas", 9), relief="flat",
                  padx=12, pady=4, cursor="hand2",
                  command=_cancel).pack(side="right")

        dlg.bind("<Return>", lambda e: _ok())
        dlg.bind("<Escape>", lambda e: _cancel())
        parent.wait_window(dlg)
        return result[0]


class SuitePanel(tk.Frame):
    def __init__(self, parent_frame, app):
        super().__init__(parent_frame, bg="#0f0f0f")
        self.pack(fill="both", expand=True)
        self.app = app
        self.scenarios = []
        self._selected_idx = None
        self._active_running_idx = None
        # ── Recording state ───────────────────────────────────────────────────
        self._rec_settings  = RecorderSettings() if RecorderSettings else None
        self._rec_enabled   = tk.BooleanVar(value=False)
        self._active_recorder: "Recorder | None" = None
        self._build()

    def _build(self):
        hdr = tk.Frame(self, bg="#0f0f0f")
        hdr.pack(fill="x", padx=12, pady=(12, 4))
        tk.Label(hdr, text="TEST SUITE", bg="#0f0f0f", fg="#fff", font=("Consolas", 13, "bold")).pack(side="left")

        # ── Bottom control panel — pack FIRST so it's always visible ─────────
        bot = tk.Frame(self, bg="#0f0f0f")
        bot.pack(side="bottom", fill="x", padx=12, pady=4)
        s_btn = dict(font=("Consolas", 9, "bold"), relief="flat", padx=10, pady=6, cursor="hand2")

        add_clear_f = tk.Frame(bot, bg="#0f0f0f")
        add_clear_f.pack(fill="x", pady=(0, 4))
        add_clear_f.columnconfigure(0, weight=3)  
        add_clear_f.columnconfigure(1, weight=1)  

        tk.Button(add_clear_f, text="➕ Add Current Scenario", bg=INCLUDE_COLOR, fg="#000", command=self._add_current, **s_btn).grid(row=0, column=0, sticky="ew", padx=(0, 2))
        tk.Button(add_clear_f, text="✕ Clear", bg="#8B0000", fg="#fff", command=self._clear_all, **s_btn).grid(row=0, column=1, sticky="ew", padx=(2, 0))

        cfg_f = tk.Frame(bot, bg="#161616", pady=6, padx=10)
        cfg_f.pack(fill="x", pady=(0, 4))
        cfg_f.columnconfigure(1, weight=1)

        # Row 0: Suite Title
        tk.Label(cfg_f, text="TITLE:", bg="#161616", fg="#aaa", font=("Consolas", 9, "bold")).grid(row=0, column=0, sticky="w")
        self.title_var = tk.StringVar(value="")
        self.title_entry = tk.Entry(cfg_f, textvariable=self.title_var, width=20, bg="#1a1a1a", fg="#fff", insertbackground="#fff", font=("Consolas", 10), relief="flat")
        self.title_entry.grid(row=0, column=1, columnspan=6, sticky="ew", padx=5)

        # Separator
        tk.Frame(cfg_f, bg="#2a2a2a", height=1).grid(row=1, column=0, columnspan=7, sticky="ew", pady=(6, 4))

        tk.Label(cfg_f, text="↺  RERUN FAILED:", bg="#161616", fg="#FF6F00", font=("Consolas", 9, "bold")).grid(row=2, column=0, sticky="w")
        
        self.rerun_var = tk.StringVar(value="0") # Default is 0 (Disabled)
        self.rerun_till_pass_var = tk.BooleanVar(value=False)
        
        # Numeric Entry Box: User inputs an integer (e.g., 2) for a strict retry limit
        self.rerun_entry = tk.Entry(cfg_f, textvariable=self.rerun_var, width=4, bg="#1a1a1a", fg="#fff", insertbackground="#fff", font=("Consolas", 10), relief="flat")
        self.rerun_entry.grid(row=2, column=1, padx=(8, 2), sticky="w")
        
        tk.Label(cfg_f, text="×", bg="#161616", fg="#555", font=("Consolas", 9)).grid(row=2, column=2, sticky="w", padx=(0, 6))
        
        # Checkbox: Overrides numeric entries to loop endlessly until all points pass
        self.rerun_chk = tk.Checkbutton(cfg_f, text="until pass", variable=self.rerun_till_pass_var, bg="#161616", fg="#888", selectcolor="#1a1a1a", activebackground="#161616", font=("Consolas", 9), command=self._on_rerun_toggle)
        self.rerun_chk.grid(row=2, column=3, sticky="w")

        btn_row = tk.Frame(bot, bg="#0f0f0f")
        btn_row.pack(fill="x")
        self.btn_run_suite = tk.Button(btn_row, text="▶ Run Suite", bg="#8AB5FF", fg="#fff", command=self._run_suite, state="disabled", **s_btn)
        self.btn_run_suite.pack(side="left", fill="x", expand=True, padx=(0, 2))
        tk.Button(btn_row, text="💾", bg="#222", fg="#ccc", command=self._save_suite, **s_btn).pack(side="left", padx=2)
        tk.Button(btn_row, text="📂", bg="#222", fg="#ccc", command=self._load_suite, **s_btn).pack(side="left", padx=2)
        tk.Button(btn_row, text="📊", bg="#222", fg="#ccc", command=self._open_report_picker, **s_btn).pack(side="left", padx=(2, 0))

        self.lbl_suite_progress = tk.Label(bot, text="", bg="#0f0f0f", fg="#aaa", font=("Consolas", 8), anchor="w")
        self.lbl_suite_progress.pack(fill="x", pady=(3, 0))

        # ── Scenario list — fills remaining space above the bottom panel ──────
        list_frame = tk.Frame(self, bg="#0f0f0f")
        list_frame.pack(fill="both", expand=True, padx=12, pady=4)

        # Header row: label  +  recording toggle  +  ⚙ settings
        hdr_row = tk.Frame(list_frame, bg="#0f0f0f")
        hdr_row.pack(fill="x", pady=(0, 4))
        tk.Label(hdr_row, text="SCENARIOS  (use arrows to reorder)",
                 bg="#0f0f0f", fg="#aaa", font=("Consolas", 8)).pack(side="left")

        # ⚙ recording settings button (far right)
        self.btn_rec_settings = tk.Button(
            hdr_row, text="⚙",
            bg="#0f0f0f", fg="#555",
            font=("Consolas", 9), relief="flat",
            padx=4, pady=0, cursor="hand2",
            command=self._open_rec_settings,
        )
        self.btn_rec_settings.pack(side="right", padx=(2, 0))

        # REC toggle button (right of label, left of ⚙)
        self._rec_btn_texts = {True: "⏺ REC ON", False: "⏺ REC OFF"}
        self._rec_btn_colors = {True: "#CC0000", False: "#333"}
        self.btn_rec_toggle = tk.Button(
            hdr_row, text=self._rec_btn_texts[False],
            bg=self._rec_btn_colors[False], fg="#888",
            font=("Consolas", 8, "bold"), relief="flat",
            padx=6, pady=1, cursor="hand2",
            command=self._toggle_recording,
        )
        self.btn_rec_toggle.pack(side="right", padx=(4, 2))

        self.cards_canvas = tk.Canvas(list_frame, bg="#0a0a0a", highlightthickness=1, highlightbackground="#222")
        sb = tk.Scrollbar(list_frame, orient="vertical", command=self.cards_canvas.yview)
        self.cards_canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self.cards_canvas.pack(fill="both", expand=True)

        self.cards_frame = tk.Frame(self.cards_canvas, bg="#0a0a0a")
        self.cards_window = self.cards_canvas.create_window((0, 0), window=self.cards_frame, anchor="nw")
        self.cards_frame.bind("<Configure>", lambda e: self.cards_canvas.configure(scrollregion=self.cards_canvas.bbox("all")))
        self.cards_canvas.bind("<Configure>", lambda e: self.cards_canvas.itemconfig(self.cards_window, width=e.width))

        self._rebuild_cards()
        self.after(100, self._bind_scroll)

    def _on_rerun_toggle(self):
        if self.rerun_till_pass_var.get():
            self.rerun_entry.config(state="disabled", fg="#444")
        else:
            self.rerun_entry.config(state="normal", fg="#fff")

    # ── Recording helpers ─────────────────────────────────────────────────────

    def _toggle_recording(self):
        """Toggle the recording-enabled flag and update button appearance."""
        if self._rec_settings is None:
            messagebox.showwarning(
                "Recording Unavailable",
                "iscs_recorder.py is missing.\n\nRun:\n  pip install imageio imageio-ffmpeg",
                parent=self.app,
            )
            return
        new_state = not self._rec_enabled.get()
        self._rec_enabled.set(new_state)
        self._rec_settings.enabled = new_state
        self.btn_rec_toggle.config(
            text=self._rec_btn_texts[new_state],
            bg=self._rec_btn_colors[new_state],
            fg="#fff" if new_state else "#888",
        )

    def _open_rec_settings(self):
        """Open the recording settings dialog."""
        if self._rec_settings is None:
            messagebox.showwarning(
                "Recording Unavailable",
                "iscs_recorder.py is missing.\n\nRun:\n  pip install imageio imageio-ffmpeg",
                parent=self.app,
            )
            return
        result = RecordingSettingsDialog.show(self.app, self._rec_settings)
        if result is not None:
            result.enabled = self._rec_enabled.get()
            self._rec_settings = result

    def _start_recorder_for_card(self, sc, evidence_dir: "Path"):
        """
        Called at the start of each card run.
        Performs pre-flight check, warns user if needed, then starts recorder.
        Returns the active Recorder instance or None.
        """
        if (not self._rec_enabled.get()
                or self._rec_settings is None
                or Recorder is None
                or not RECORDER_AVAILABLE):
            return None

        from pathlib import Path as _P
        ev_dir = _P(evidence_dir)
        ev_dir.mkdir(parents=True, exist_ok=True)

        ok, msg = pre_flight_check(self._rec_settings, ev_dir)
        if not ok and msg:
            proceed = messagebox.askyesno(
                "Recording Warning", msg, parent=self.app,
            )
            if not proceed:
                return None

        # Resolve the active monitor to capture based on settings configurations
        target_mon = None
        cfg_disp = getattr(self._rec_settings, "capture_display", "Auto")
        if cfg_disp == "Auto":
            target_mon = self.app._find_monitor_by_info(sc.monitor_info)
        else:
            for mon in self.app.monitors:
                if mon.label == cfg_disp:
                    target_mon = mon
                    break

        if target_mon is None:
            target_mon = self.app.active_mon or self.app.monitors[0]

        monitor_bbox = (target_mon.x, target_mon.y, target_mon.width, target_mon.height)

        # Map non-native output scale if requested
        target_resolution = None
        cfg_res = getattr(self._rec_settings, "capture_resolution", "Native (recommended)")
        if cfg_res != "Native (recommended)" and "x" in cfg_res:
            try:
                w, h = map(int, cfg_res.split('x'))
                target_resolution = (w, h)
            except ValueError:
                pass

        rec = Recorder(
            settings          = self._rec_settings,
            card_name         = sc.name,
            evidence_dir      = ev_dir,
            monitor_bbox      = monitor_bbox,
            target_resolution = target_resolution
        )
        started = rec.start()
        if not started:
            self.app._log("⚠ Recorder failed to start — check iscs_recorder logs.")
            return None
        self.app._log(f"⏺ Recording started for '{sc.name}' on {target_mon.name} @ {self._rec_settings.fps} fps")
        return rec

    def _stop_recorder(self, rec, card_name: str = ""):
        """Stop an active recorder and log completion."""
        if rec is None:
            return
        try:
            rec.stop()
            self.app._log(f"⏹ Recording stopped{' for ' + card_name if card_name else ''}.")
        except Exception as e:
            self.app._log(f"⚠ Recorder stop error: {e}")

    def _on_wheel(self, event):
        self.cards_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _bind_scroll(self):
        self.cards_canvas.bind("<MouseWheel>", self._on_wheel)
        self.cards_frame.bind("<MouseWheel>", self._on_wheel)

    def _ask_name(self, default=""):
        dlg = tk.Toplevel(self.app)
        dlg.title("Scenario Name")
        sx = self.winfo_rootx() + (self.winfo_width() - 320) // 2
        sy = self.winfo_rooty() + (self.winfo_height() - 120) // 2
        dlg.geometry(f"320x120+{sx}+{sy}")
        dlg.configure(bg="#0f0f0f")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.attributes("-topmost", True)
        result = [None]

        tk.Label(dlg, text="Enter scenario name:", bg="#0f0f0f", fg="#aaa",
                 font=("Consolas", 10)).pack(pady=(16, 4))
        var = tk.StringVar(value=default)
        entry = tk.Entry(dlg, textvariable=var, bg="#1a1a1a", fg="#fff",
                         insertbackground="#fff", font=("Consolas", 11),
                         relief="flat", width=28)
        entry.pack(padx=20)
        entry.select_range(0, "end")
        entry.focus_set()

        def ok(e=None):
            result[0] = var.get().strip() or default
            dlg.destroy()

        entry.bind("<Return>", ok)
        dlg.bind("<Escape>", lambda e: dlg.destroy())
        ok_btn = tk.Button(dlg, text="OK", bg="#2979FF", fg="#fff",
                  font=("Consolas", 10, "bold"), relief="flat",
                  padx=14, pady=4, command=ok, cursor="hand2")
        ok_btn.pack(pady=8)
        self.app.wait_window(dlg)
        return result[0]

    def _add_current(self):
        mon = self.app.active_mon
        mode = self.app.run_mode.get()

        if mode == "iscs":
            # For ISCS cards: open card config dialog first
            card_cfg = SuiteCardConfigDialog.ask(self.app)
            if card_cfg is None: return
            name = card_cfg.get("card_name", f"ISCS Card {len(self.scenarios)+1}")
            points = card_cfg.get("profile_points") or list(getattr(self.app, "iscs_excel_points", []))
        else:
            name = self._ask_name(f"Scenario {len(self.scenarios)+1}")
            if name is None: return
            card_cfg = None
            points = list(getattr(self.app, "iscs_excel_points", []))

        sc = Scenario(
            name,
            mode,
            [Zone.from_dict(z.to_dict()) for z in self.app.zones],
            {"x": mon.x, "y": mon.y, "width": mon.width, "height": mon.height, "name": mon.name},
            self.app.grid_spacing,
            points
        )
        # Attach card config to scenario for ISCS suite runs
        sc.card_cfg = card_cfg or {}
        # Merge template zones into zones_per_page under "Global"
        tmpl_zones = (card_cfg or {}).get("template_zones", {})
        if tmpl_zones:
            sc.zones_per_page.setdefault("Global", {}).update(tmpl_zones)
        # Also merge app's current zones_per_page
        for page, zt_dict in self.app.zones_per_page.items():
            sc.zones_per_page.setdefault(page, {}).update(zt_dict)
        self.scenarios.append(sc)
        self._rebuild_cards()

    def _edit_card_cfg(self, idx):
        sc = self.scenarios[idx]
        existing_cfg = getattr(sc, "card_cfg", {})
        result = SuiteCardConfigDialog.ask(self.app, existing_cfg)
        if result is None: return
        sc.card_cfg = result
        sc.name = result.get("card_name", sc.name)
        if result.get("profile_points"):
            sc.iscs_points = result["profile_points"]
        # Merge template zones if loaded
        tmpl_zones = result.get("template_zones", {})
        if tmpl_zones:
            sc.zones_per_page.setdefault("Global", {}).update(tmpl_zones)
        # Config changed — reset the flow so it's re-generated fresh on next run
        sc.procedure_flow = None
        self._rebuild_cards()

    def _open_flow_dialog(self, idx):
        """Open the ProcedureFlowDialog for the scenario at idx.
        If the scenario has no flow yet, auto-register defaults first so the
        dialog always shows something meaningful rather than an empty list.
        """
        if not WORKFLOW_AVAILABLE:
            return
        sc = self.scenarios[idx]
        # Ensure the flow is populated before opening the dialog
        if sc.procedure_flow is None:
            card_cfg   = getattr(sc, "card_cfg", {})
            nav        = card_cfg.get("navigation", {})
            zones_dict = {}
            for page_zones in sc.zones_per_page.values():
                for zt, z in page_zones.items():
                    if zt not in zones_dict:
                        zones_dict[zt] = z
            for z in sc.zones:
                if z.zone_type not in zones_dict:
                    zones_dict[z.zone_type] = z
            sc.procedure_flow = auto_register_procedures(sc, zones_dict, nav)
        _mon = (self.app._find_monitor_by_info(sc.monitor_info)
                if hasattr(self.app, '_find_monitor_by_info') and sc.monitor_info
                else None)
        open_procedure_flow_dialog(self.app, sc.procedure_flow,
                                   title=f"{sc.name} — Execution Flow",
                                   monitor=_mon)
        
    def _clear_all(self):
        if not self.scenarios:
            return
        if messagebox.askyesno("Clear All", "Are you sure you want to delete all scenarios from the suite?", parent=self.app):
            self.scenarios.clear()
            self._selected_idx = None
            self._rebuild_cards()

    def _rebuild_cards(self):
        for w in self.cards_frame.winfo_children():
            w.destroy()
        if not self.scenarios:
            tk.Label(self.cards_frame, text="No scenarios added.", bg="#0a0a0a", fg="#444", font=("Consolas", 9)).pack(pady=20)
            self.btn_run_suite.config(state="disabled")
            return
        self.btn_run_suite.config(state="normal")
        for i, sc in enumerate(self.scenarios):
            self._build_card(i, sc)

    def _rename_scenario(self, idx):
        name = self._ask_name(self.scenarios[idx].name)
        if name:
            self.scenarios[idx].name = name
            self._rebuild_cards()

    def _build_card(self, idx, sc):
        mode_map = {
            "sequence": (TARGET_COLOR, "RPA"), 
            "grid": (INCLUDE_COLOR, "FUZZER"), 
            "iscs": (ALARM_PANEL_COLOR, "ISCS")
        }
        active_color, m_label = mode_map.get(sc.mode, ("#888", sc.mode.upper()))
        
        is_selected = self._selected_idx == idx
        is_running = getattr(self, '_active_running_idx', None) == idx

        if is_running:
            hl_color = active_color
            hl_thick = 2
            bar_color = active_color
            lbl_fg = "#000" if sc.mode == "grid" else "#fff" 
        else:
            hl_color = "#666" if is_selected else "#2a2a2a"
            hl_thick = 1
            bar_color = "#333333" 
            lbl_fg = active_color 

        card = tk.Frame(self.cards_frame, bg="#161616", highlightthickness=hl_thick, highlightbackground=hl_color)
        card.pack(fill="x", padx=4, pady=2)

        tk.Frame(card, bg=bar_color, width=4).pack(side="left", fill="y")

        acts = tk.Frame(card, bg="#161616")
        acts.pack(side="right", padx=2)
        tk.Button(acts, text="↑", command=lambda i=idx: self._move(i, -1), bg="#222", fg="#aaa", font=("Consolas", 8), relief="flat").pack(pady=1)
        tk.Button(acts, text="↓", command=lambda i=idx: self._move(i, 1), bg="#222", fg="#aaa", font=("Consolas", 8), relief="flat").pack(pady=1)
        tk.Button(acts, text="🗑", command=lambda i=idx: self._remove(i), bg="#222", fg=EXCLUDE_COLOR, font=("Consolas", 8), relief="flat").pack(pady=1)
        if sc.mode == "iscs":
            tk.Button(acts, text="⚙", command=lambda i=idx: self._edit_card_cfg(i),
                      bg="#222", fg="#AA00FF", font=("Consolas", 9), relief="flat").pack(pady=1)
            if WORKFLOW_AVAILABLE:
                tk.Button(acts, text="⚡", command=lambda i=idx: self._open_flow_dialog(i),
                          bg="#1a1a2e", fg="#2979FF", font=("Consolas", 9), relief="flat",
                          cursor="hand2").pack(pady=1)

        info = tk.Frame(card, bg="#161616", cursor="hand2")
        info.pack(side="left", fill="both", expand=True, padx=6, pady=4)

        row1 = tk.Frame(info, bg="#161616")
        row1.pack(fill="x")
        
        tk.Label(row1, text=f"{idx+1}. {sc.name}", bg="#161616", fg="#fff", font=("Consolas", 10, "bold")).pack(side="left")
        rename_btn = tk.Button(row1, text="✏", command=lambda i=idx: self._rename_scenario(i),
                  bg="#161616", fg="#555", font=("Consolas", 8), relief="flat", padx=2, pady=0, cursor="hand2", bd=0)
        rename_btn.pack(side="left", padx=(4, 0))
        
        row2 = tk.Frame(info, bg="#161616")
        row2.pack(fill="x")
        
        tk.Label(row2, text=m_label, bg=bar_color, fg=lbl_fg, font=("Consolas", 7, "bold"), padx=3).pack(side="left", padx=(0, 5))
        row3 = tk.Frame(info, bg="#161616")
        row3.pack(fill="x", pady=(4, 2))
        if sc.mode == "iscs":
            tk.Button(row3, text="▶ Run Flow",
                      bg="#1a2a3a", fg="#60a5fa",
                      font=("Consolas", 8, "bold"), relief="flat",
                      padx=6, pady=2, cursor="hand2",
                      command=lambda i=idx: self._run_flow(i)).pack(side="left", padx=(0, 8))

        # ── Per-card loop controls ────────────────────────────────────────────
        loop_chip = tk.Frame(row3, bg="#1e1e1e", padx=6, pady=2)
        loop_chip.pack(side="left")

        tk.Label(loop_chip, text="LOOP", bg="#1e1e1e", fg="#666",
                 font=("Consolas", 7, "bold")).pack(side="left", padx=(0, 4))

        sc_loop_var = tk.StringVar(value=str(getattr(sc, "card_loop", 1)))
        sc_inf_var  = tk.BooleanVar(value=getattr(sc, "card_infinite", False))

        loop_entry = tk.Entry(loop_chip, textvariable=sc_loop_var, width=3,
                              bg="#111", fg="#fff", insertbackground="#fff",
                              font=("Consolas", 9), relief="flat",
                              disabledbackground="#111", disabledforeground="#444")
        loop_entry.pack(side="left", padx=(0, 4))

        def _toggle_inf(i=idx, lv=sc_loop_var, iv=sc_inf_var, le=loop_entry):
            sc = self.scenarios[i]
            sc.card_infinite = iv.get()
            if sc.card_infinite:
                le.config(state="disabled")
            else:
                le.config(state="normal")
                try: sc.card_loop = max(1, int(lv.get() or 1))
                except: sc.card_loop = 1

        def _save_loop(e=None, i=idx, lv=sc_loop_var, iv=sc_inf_var):
            sc = self.scenarios[i]
            if not iv.get():
                try: sc.card_loop = max(1, int(lv.get() or 1))
                except: sc.card_loop = 1; lv.set("1")

        loop_entry.bind("<FocusOut>", _save_loop)
        loop_entry.bind("<Return>",   _save_loop)
        if sc_inf_var.get():
            loop_entry.config(state="disabled")

        tk.Checkbutton(loop_chip, text="∞", variable=sc_inf_var,
                       bg="#1e1e1e", fg="#888", selectcolor="#111",
                       activebackground="#1e1e1e",
                       font=("Consolas", 9),
                       command=_toggle_inf).pack(side="left")
                      
        mon = self.app._find_monitor_by_info(sc.monitor_info)
        if mon and mon in self.app.monitors:
            mon_idx = self.app.monitors.index(mon)
            mon_color = SCREEN_COLORS[mon_idx % len(SCREEN_COLORS)]
            mon_text = f"Disp {mon.display_num}"
        else:
            mon_color = "#aaa"
            mon_text = "Disp ?"
            
        tk.Label(row2, text=mon_text, bg="#222", fg=mon_color, font=("Consolas", 7, "bold"), padx=3).pack(side="left", padx=(0, 5))
        
        zpp_count = sum(len(v) for v in sc.zones_per_page.values())
        z_count   = zpp_count if zpp_count > 0 else len(sc.zones)
        tk.Label(row2, text=f"{z_count} zones", bg="#222", fg="#aaa", font=("Consolas", 7), padx=3).pack(side="left")
        if sc.mode == "iscs":
            cfg = getattr(sc, "card_cfg", {})
            pt_count  = len(sc.iscs_points) if hasattr(sc, "iscs_points") else 0
            nav_pages = len(cfg.get("navigation", {}).get("pages", []))
            if pt_count:
                tk.Label(row2, text=f"{pt_count} pts", bg="#1a2a1a", fg="#00C853",
                         font=("Consolas", 7), padx=3).pack(side="left", padx=(4, 0))
            if nav_pages:
                tk.Label(row2, text=f"{nav_pages} pages", bg="#1a1a2a", fg="#2979FF",
                         font=("Consolas", 7), padx=3).pack(side="left", padx=(2, 0))
            proto_host = cfg.get("protocol", {}).get("host", "")
            if proto_host:
                tk.Label(row2, text=proto_host, bg="#222", fg="#888",
                         font=("Consolas", 7), padx=3).pack(side="left", padx=(2, 0))

        clickable = [card, info, row1, row2] + [w for w in row1.winfo_children() if w != rename_btn] + list(row2.winfo_children())
        for widget in clickable:
            widget.bind("<Button-1>", lambda e, i=idx: self._select_scenario(i))

        def _bind_tree(w):
            w.bind("<MouseWheel>", self._on_wheel)
            for child in w.winfo_children():
                _bind_tree(child)
                
        _bind_tree(card)

    def _select_scenario(self, idx):
        self._selected_idx = idx
        sc = self.scenarios[idx]
        mon = self.app._find_monitor_by_info(sc.monitor_info)
        if mon:
            self.app.active_mon = mon
            self.app.screen_selector._highlight(self.app.monitors.index(mon))
        self.app.run_mode.set(sc.mode)
        # Load zones FROM the scenario into the workspace (was incorrectly copying app.zones into itself)
        self.app.zones = [Zone.from_dict(z.to_dict()) for z in sc.zones]
        # Also load zones_per_page from the scenario
        zpp = {}
        for page, zt_dict in sc.zones_per_page.items():
            zpp[page] = {zt: Zone.from_dict(z.to_dict()) for zt, z in zt_dict.items()}
        self.app.zones_per_page = zpp
        self.app.grid_spacing = sc.grid_spacing
        self.app._update_mode_buttons()
        self.app._refresh_stats_only()
        self.app._draw_minimap()
        self.app._update_overlay_btn()
        self._rebuild_cards()

    def _remove(self, idx): self.scenarios.pop(idx); self._rebuild_cards()

    def _move(self, idx, dir):
        new_idx = idx + dir
        if 0 <= new_idx < len(self.scenarios):
            self.scenarios[idx], self.scenarios[new_idx] = self.scenarios[new_idx], self.scenarios[idx]
            if self._selected_idx == idx: self._selected_idx = new_idx
            elif self._selected_idx == new_idx: self._selected_idx = idx
            self._rebuild_cards()

    @staticmethod
    def _json_safe(obj):
        """
        Recursively convert any non-JSON-serialisable value to a safe type.
        Handles datetime, numpy scalars, openpyxl types, NaN/Inf floats, etc.
        """
        import math
        if obj is None:
            return None
        if isinstance(obj, bool):
            return obj
        if isinstance(obj, int):
            return obj
        if isinstance(obj, float):
            if math.isnan(obj) or math.isinf(obj):
                return None
            return obj
        if isinstance(obj, str):
            return obj
        if isinstance(obj, dict):
            return {str(k): SuitePanel._json_safe(v) for k, v in obj.items()}
        if isinstance(obj, (list, tuple)):
            return [SuitePanel._json_safe(v) for v in obj]
        # datetime / date
        try:
            import datetime as _dt
            if isinstance(obj, (_dt.datetime, _dt.date)):
                return obj.isoformat()
        except Exception:
            pass
        # numpy scalars
        try:
            import numpy as np
            if isinstance(obj, np.integer):
                return int(obj)
            if isinstance(obj, np.floating):
                v = float(obj)
                return None if (math.isnan(v) or math.isinf(v)) else v
            if isinstance(obj, np.ndarray):
                return obj.tolist()
        except Exception:
            pass
        # fallback — convert to string
        return str(obj)

    def _save_suite(self):
        path = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("Suite config", "*.json")], parent=self.app)
        if not path: return
        data = {
            "rerun_count": int(self.rerun_var.get() or 0),
            "rerun_till_pass": self.rerun_till_pass_var.get(),
            "scenarios": [s.to_dict() for s in self.scenarios]
        }
        data = SuitePanel._json_safe(data)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        self.app._log(f"Suite saved → {path}")

    def _load_suite(self):
        path = filedialog.askopenfilename(filetypes=[("Suite config", "*.json")], parent=self.app)
        if not path: return
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            messagebox.showerror(
                "Load Failed",
                f"Could not read suite file:\n{e}\n\n"
                "The file may be corrupted. Re-save the suite to fix this.",
                parent=self.app)
            return
        self.scenarios = [Scenario.from_dict(s) for s in data.get("scenarios", [])]
        self.rerun_var.set(str(data.get("rerun_count", 0)))
        self.rerun_till_pass_var.set(data.get("rerun_till_pass", False))
        self._on_rerun_toggle()
        self._rebuild_cards()
        self.app._log(f"Suite loaded ← {path}")

    def _run_suite(self):
        try:
            # -1 represents infinite (till pass), otherwise parse the numeric text box value (e.g., 2)
            rerun_count = -1 if self.rerun_till_pass_var.get() else int(self.rerun_var.get() or 0)
        except ValueError:
            rerun_count = 0  # Fallback to disabled on parsing errors

        self.app.set_execution_state("running")
        self.rerun_entry.config(state="disabled", fg="#444")
        self.rerun_chk.config(state="disabled")
        self.title_entry.config(state="disabled", fg="#444")

        # ── Pass a recorder factory so SuiteRunner can start/stop per card ──
        rec_panel = self  # captured for closure

        def _rec_start(sc, evidence_dir):
            return rec_panel._start_recorder_for_card(sc, evidence_dir)

        def _rec_stop(rec, card_name=""):
            rec_panel._stop_recorder(rec, card_name)

        def _rec_update(rec, point_id, equip_desc, attr_desc):
            if rec is not None:
                rec.update_point(point_id, equip_desc, attr_desc)

        # M3.4: HUD protocol-wait updates via callback (was a __main__.app.hud poke
        # inside SuiteRunner). Preserves the original marshalling: waiting → direct,
        # online → app.after(0, …).
        _app = self.app
        def _cb_proto_status(state, done, total, msg):
            hud = getattr(_app, "hud", None)
            if hud is None or not hud.winfo_exists():
                return
            if state == "waiting":
                hud.update_waiting(done, total, msg)
            else:  # "online"
                _app.after(0, lambda: hud.update_running(done, total, msg))

        self.suite_runner = SuiteRunner(
            self.scenarios, self.app.monitors, self.app.protocols, APP_CONFIG,
            self._cb_start, self._cb_prog, self._cb_pause, lambda p, tp, path: None,
            self._cb_done, self.app._log, suite_title=self.title_var.get().strip(),
            rerun_failed_count=rerun_count,
            on_rec_start=_rec_start,
            on_rec_stop=_rec_stop,
            on_rec_update=_rec_update,
            on_proto_status=_cb_proto_status,
        )
        self.suite_runner.start()
        self.app.click_engine = self.suite_runner

    def _run_flow(self, idx):
        sc = self.scenarios[idx]
        if not WORKFLOW_AVAILABLE:
            messagebox.showerror("Workflow Error", "iscs_workflow.py not available.")
            return

        enabled_steps = []
        if sc.procedure_flow:
            enabled_steps = [p.name for p in sc.procedure_flow.procedures if p.enabled]
        
        if not enabled_steps:
            messagebox.showinfo("Empty Flow", "The procedure flow for this card contains no enabled steps.")
            return

        steps_list = "\n".join(f"  - {step}" for step in enabled_steps)
        if not messagebox.askyesno("Confirm Flow Execution", 
                                   f"Do you want to execute this standalone flow run?\n\nSteps to run:\n{steps_list}"):
            return

        self.app.set_execution_state("running")
        self.app.stats.set("state", "RUNNING FLOW", "#2979FF")
        self.app.run_progress.set_fraction(0)

        # Mock standard execution engine interfaces for Space/Esc hooks
        class MockEngine:
            def __init__(self, stop_evt, pause_evt):
                self._stop_event = stop_evt
                self._pause_event = pause_evt
                self.results = []
            def stop(self): self._stop_event.set(); self._pause_event.set()
            def pause(self, r="manual"): self._pause_event.clear()
            def resume(self): self._pause_event.set()
            @property
            def is_paused(self): return not self._pause_event.is_set()
            def is_alive(self): return True

        self.app.click_engine = MockEngine(threading.Event(), threading.Event())
        self.app.click_engine._pause_event.set()

        def worker():
            try:
                ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                sc_dir = LOG_DIR / f"standalone_{sc.name.replace(' ', '_')}_{ts}"
                sc_dir.mkdir(parents=True, exist_ok=True)
                
                self.app._log(f"[Flow] Standalone run started for card: {sc.name}")
                
                # Resolve active verifier and protocol targets
                zones_dict = {}
                for page_zones in sc.zones_per_page.values():
                    for zt, z in page_zones.items():
                        zones_dict[zt] = z
                for z in sc.zones:
                    zones_dict[z.zone_type] = z
                
                verifier = ISCSVerifier(zones_dict, APP_CONFIG)
                handler = self.app.protocols.get_protocol(sc.card_cfg.get("protocol", {}).get("type", "MODBUS"))

                runner = ProcedureRunner(
                    flow=sc.procedure_flow,
                    verifier=verifier,
                    handler=handler,
                    config=APP_CONFIG,
                    on_log=lambda msg: self.app._log(f"[Flow] {msg}"),
                    stop_event=self.app.click_engine._stop_event,
                    pause_event=self.app.click_engine._pause_event
                )

                def update_progress(step_name, done, total):
                    pct = (done / total) * 100
                    self.app.run_progress.set_fraction(pct)
                    self.app.run_progress.set_text(f"[Flow] Executing: {step_name} ({done}/{total})")

                runner.run_standalone(sc, sc_dir, on_progress=update_progress)
                self.app.after(0, lambda: self.app._log(f"[Flow] Standalone run complete."))
            except Exception as ex:
                self.app.after(0, lambda: self.app._log(f"[Flow] Standalone run crashed: {ex}"))
            finally:
                self.app.after(0, lambda: self.app.set_execution_state("idle"))
                self.app.stats.set("state", "IDLE", "#444")
                self.app.run_progress.set_text("Ready.")

        threading.Thread(target=worker, daemon=True).start()
        
    def _cb_start(self, p, tp, sn, ts, sc): 
        def update_ui():
            if getattr(self, 'suite_runner', None) and self.suite_runner._stop_event.is_set():
                return

            self.lbl_suite_progress.config(text=f"Scenario {sn}/{ts}: {sc.name}")
            self._active_running_idx = sn - 1  
            self._rebuild_cards()
            
            if self.app.hud: 
                try: self.app.hud.destroy()
                except: pass
                
            mon = self.app._find_monitor_by_info(sc.monitor_info)
            if sc.mode == "iscs": 
                total = len(sc.iscs_points)
            else: 
                total = len(generate_points(sc.mode, mon, sc.grid_spacing, sc.zones)[0])
            self.app.hud = HudOverlay(self.app, total, mon, name=sc.name)
            self.app.hud.set_suite_info(sn, ts, p, tp, sc.name)
            
        self.after(0, update_ui)

    def _cb_prog(self, p, tp, sn, ts, d, t, v1, v2):
        def _update():
            attempt = getattr(self.suite_runner, "current_rerun_attempt", 0) if getattr(self, "suite_runner", None) else 0
            if attempt > 0:
                if self.app.hud and self.app.hud.winfo_exists():
                    self.app.hud.update_rerun(attempt, d, t)
            else:
                if d < 0:
                    if self.app.hud and self.app.hud.winfo_exists():
                        self.app.hud.update_rerun(abs(d), 0, t)
                else:
                    self.app._cb_progress(d, t, v1, v2)
        self.app.after(0, _update)
    def _cb_pause(self, r): self.app.after(0, lambda: self.app._cb_paused(0, 0, r))
    
    def _cb_done(self, sd, log, stp):
        self._last_suite_dir = sd          # remembered for the report picker (P5)
        self.after(0, lambda: self._finish(stp, error_msg=log))

    def _open_report_picker(self):
        """P5 UI picker — generate any registered report template from a finished
        suite's saved results (suite_results.json), on demand."""
        try:
            import iscs_report_templates as rpt
        except Exception as e:
            messagebox.showerror("Reports", f"Report templates unavailable: {e}", parent=self.app)
            return

        # Resolve the results source: the last run, else let the user pick one.
        results_path = None
        last = getattr(self, "_last_suite_dir", None)
        if last:
            cand = Path(last) / "suite_results.json"
            if cand.exists():
                results_path = cand
        if results_path is None:
            picked = filedialog.askopenfilename(
                title="Pick a suite_results.json",
                filetypes=[("Suite results", "suite_results.json"), ("JSON files", "*.json")],
                parent=self.app)
            if not picked:
                return
            results_path = Path(picked)
        if not results_path.exists():
            messagebox.showinfo("Reports",
                                "No suite_results.json found yet — run a suite first.",
                                parent=self.app)
            return

        dlg = tk.Toplevel(self.app)
        dlg.title("📊 Generate Report")
        dlg.configure(bg="#0f0f0f"); dlg.geometry("440x340"); dlg.attributes("-topmost", True)
        tk.Label(dlg, text="Generate Report As…", bg="#0f0f0f", fg="#69ff9a",
                 font=("Consolas", 13, "bold")).pack(anchor="w", padx=14, pady=(12, 2))
        tk.Label(dlg, text=f"Source: {results_path.parent.name}/{results_path.name}",
                 bg="#0f0f0f", fg="#777", font=("Consolas", 8)).pack(anchor="w", padx=14, pady=(0, 8))

        templates = rpt.list_templates()
        choice = tk.StringVar(value=templates[0]["key"])
        for t in templates:
            tk.Radiobutton(dlg, text=f"{t['name']}   ({t['audience']})", value=t["key"],
                           variable=choice, bg="#0f0f0f", fg="#ddd", selectcolor="#1a1a1a",
                           activebackground="#0f0f0f", font=("Consolas", 10), anchor="w").pack(
                           fill="x", padx=20, pady=2)

        def _gen():
            try:
                key = choice.get()
                # Legacy = the original Suite_Report.html. If the run already
                # produced one, just open it (it has the real run times/evidence);
                # only re-generate if it's missing.
                if key == "legacy":
                    existing = results_path.parent / "Suite_Report.html"
                    if existing.exists():
                        out = existing
                    else:
                        raw = json.loads(results_path.read_text(encoding="utf-8"))
                        out = rpt.generate_template_report(
                            key, raw, results_path.parent,
                            title=self.title_var.get().strip() or "Test Run")
                else:
                    raw = json.loads(results_path.read_text(encoding="utf-8"))
                    out = rpt.generate_template_report(
                        key, raw, results_path.parent,
                        title=self.title_var.get().strip() or "Test Run")
                self.app._log(f"📊 Report generated: {out.name}")
                dlg.destroy()
                try:
                    os.startfile(str(out))    # Windows — open in browser/editor
                except Exception:
                    pass
            except Exception as e:
                messagebox.showerror("Reports", f"Failed to generate report: {e}", parent=dlg)

        bf = tk.Frame(dlg, bg="#0f0f0f"); bf.pack(fill="x", padx=14, pady=14)
        tk.Button(bf, text="Generate & Open", bg="#1a3a1a", fg="#69ff9a",
                  font=("Consolas", 10, "bold"), relief="flat", padx=12, pady=6,
                  cursor="hand2", command=_gen).pack(side="left")
        tk.Button(bf, text="Cancel", bg="#222", fg="#ccc", relief="flat", padx=12, pady=6,
                  cursor="hand2", command=dlg.destroy).pack(side="left", padx=6)

    def _finish(self, was_stopped=False, error_msg=""):
        if not was_stopped:
            self.app._log("✅ Suite testing finished.")
        else:
            self.app._log("⏹ Suite testing stopped.")
        self.app.set_execution_state("idle")
        self._on_rerun_toggle()   # restores rerun_entry correctly
        self.rerun_chk.config(state="normal")
        self.title_entry.config(state="normal", fg="#fff")
        self.lbl_suite_progress.config(text="Suite Stopped." if was_stopped else "Suite Done.")
        self._active_running_idx = None  
        self._rebuild_cards()
        if error_msg:
            messagebox.showerror("Execution Error", error_msg, parent=self.app)
            
        if self.app.hud and self.app.hud.winfo_exists():
            self.app.hud.blink_and_destroy()

# ── Overlays ──────────────────────────────────────────────────────────────────
class CrosshairOverlay(tk.Toplevel):
    def __init__(self, master, valid_pts, all_pts, monitor: Monitor, mode):
        super().__init__(master)
        self.monitor = monitor
        self.valid_pts = valid_pts
        self.all_pts = all_pts
        self.mode = mode

        self.geometry(f"{monitor.width}x{monitor.height}+{monitor.x}+{monitor.y}")
        self.attributes("-topmost", True); self.attributes("-alpha", 0.55); self.overrideredirect(True)
        self.configure(bg="#010101"); self.wm_attributes("-transparentcolor", "#010101")
        self._make_clickthrough()
        self._build()

    def _make_clickthrough(self):
        try:
            self.update_idletasks()
            hwnd = ctypes.windll.user32.GetParent(self.winfo_id()) or self.winfo_id()
            style = ctypes.windll.user32.GetWindowLongW(hwnd, -20)
            ctypes.windll.user32.SetWindowLongW(hwnd, -20, style | 0x80000 | 0x20)
        except Exception: pass

    def _build(self):
        c = tk.Canvas(self, bg="#010101", highlightthickness=0, width=self.monitor.width, height=self.monitor.height)
        c.pack(fill="both", expand=True)
        self.c = c
        self._draw_crosshairs()

    def _draw_crosshairs(self):
        c, mon, r, arm = self.c, self.monitor, CROSSHAIR_R, CROSSHAIR_ARM
        valid_coords = [(pt["x"], pt["y"]) for pt in self.valid_pts]
        
        if self.mode == "grid":
            for (ax, ay) in self.all_pts:
                if (ax, ay) not in valid_coords:
                    cx, cy = ax - mon.x, ay - mon.y
                    c.create_oval(cx - r, cy - r, cx + r, cy + r, outline=FILTERED_COLOR, width=2)
                    c.create_oval(cx - 2, cy - 2, cx + 2, cy + 2, fill=FILTERED_COLOR, outline="")

        for pt in self.valid_pts:
            cx, cy = pt["x"] - mon.x, pt["y"] - mon.y
            color = TARGET_COLOR if "Target" in pt["label"] else INCLUDE_COLOR
            c.create_oval(cx - r, cy - r, cx + r, cy + r, outline=color, width=2)
            c.create_line(cx - r - arm, cy, cx - r, cy, fill=color, width=2)
            c.create_line(cx + r, cy, cx + r + arm, cy, fill=color, width=2)
            c.create_line(cx, cy - r - arm, cx, cy - r, fill=color, width=2)
            c.create_line(cx, cy + r, cx, cy + r + arm, fill=color, width=2)
            c.create_oval(cx - 2, cy - 2, cx + 2, cy + 2, fill=color, outline="")
            
            if self.mode == "sequence":
                c.create_text(cx + 15, cy - 15, text=pt["label"].replace("_", " "), fill=color, font=("Consolas", 10, "bold"), anchor="w")

        c.create_text(mon.width - 10, mon.height - 10, text=f"PREVIEW — {len(self.valid_pts)} click pts", fill="#333", font=("Consolas", 9), anchor="se")

class HudOverlay(tk.Toplevel):
    def __init__(self, master, total, monitor, name="Manual Run"):
        super().__init__(master)
        self.monitor = monitor
        self.name = name 
        x = monitor.x + monitor.width - HUD_W - HUD_MARGIN
        y = monitor.y + monitor.height - HUD_H - HUD_MARGIN
        self.geometry(f"{HUD_W}x{HUD_H}+{x}+{y}")
        self.attributes("-topmost", True); self.attributes("-alpha", HUD_ALPHA); self.overrideredirect(True)
        self.configure(bg="#0a0a0a")
        self._make_clickthrough()
        self._build()

    def _make_clickthrough(self):
        try:
            self.update_idletasks()
            hwnd = ctypes.windll.user32.GetParent(self.winfo_id()) or self.winfo_id()
            style = ctypes.windll.user32.GetWindowLongW(hwnd, -20)
            ctypes.windll.user32.SetWindowLongW(hwnd, -20, style | 0x80000 | 0x20)
        except Exception: pass

    def blink_and_destroy(self, blink_count=4, blink_period_ms=500):
        """Show TEST COMPLETE, blink, then destroy."""
        try:
            self.c.itemconfig(self.dot,        fill="#69ff9a")
            self.c.itemconfig(self.lbl_status, text="COMPLETE",     fill="#69ff9a")
            self.c.itemconfig(self.lbl_main,   text="TEST COMPLETE", fill="#69ff9a")
            self.c.itemconfig(self.lbl_target, text="",              fill="#69ff9a")
            self.attributes("-alpha", HUD_ALPHA)
        except Exception:
            pass
        def _toggle(step=0):
            if step >= blink_count * 2:
                try:
                    self.destroy()
                except Exception:
                    pass
                return
            try:
                alpha = HUD_ALPHA if step % 2 == 0 else 0.0
                self.attributes("-alpha", alpha)
            except Exception:
                pass
            self.after(blink_period_ms // 2, _toggle, step + 1)
        self.after(700, lambda: _toggle(0))
        
    def _build(self):
        c = tk.Canvas(self, bg="#0a0a0a", highlightthickness=0, width=HUD_W, height=HUD_H)
        c.pack(fill="both", expand=True)
        self.c = c
        c.create_rectangle(0, 0, HUD_W, 3, fill=INCLUDE_COLOR, outline="")
        c.create_rectangle(0, 0, HUD_W-1, HUD_H-1, outline="#1f1f1f", fill="")
        
        self.dot = c.create_oval(12, 14, 24, 26, fill=INCLUDE_COLOR, outline="")
        self.lbl_status = c.create_text(32, 20, text="RUNNING", fill=INCLUDE_COLOR, font=("Consolas", 9, "bold"), anchor="w")
        self.lbl_suite  = c.create_text(HUD_W - 12, 20, text="", fill="#aaa", font=("Consolas", 9, "bold"), anchor="e")
        
        self.lbl_scen_name = c.create_text(HUD_W//2, 45, text=self.name, fill="#f9e2af", font=("Consolas", 10, "bold"), anchor="center")
        self.lbl_main   = c.create_text(HUD_W//2, 75, text="Starting…", fill="#ffffff", font=("Consolas", 15, "bold"), anchor="center")
        
        self.lbl_target = c.create_text(HUD_W//2, 102, text="", fill="#888", font=("Consolas", 9), anchor="center")
        
        c.create_text(HUD_W//2, 132, text="Space=Pause   Esc / Ctrl+F12=Stop", fill="#444", font=("Consolas", 8), anchor="center")

    def set_suite_info(self, sn, ts, pass_num, total_passes, scen_name):
        loop_str = "∞" if total_passes == -1 else str(total_passes)
        pass_str = f"  |  LOOP {pass_num}/{loop_str}" if total_passes != 1 else ""
        self.c.itemconfig(self.lbl_suite, text=f"SCEN: {sn}/{ts}{pass_str}")
        self.c.itemconfig(self.lbl_scen_name, text=scen_name)

    def update_running(self, done, total, target_text=""):
        self.c.itemconfig(self.dot, fill=INCLUDE_COLOR)
        self.c.itemconfig(self.lbl_status, text="RUNNING", fill=INCLUDE_COLOR)
        self.c.itemconfig(self.lbl_main, text=f"{done} / {total}   ({done/total*100:.0f}%)", fill="#ffffff")
        self.c.itemconfig(self.lbl_target, text=target_text, fill="#888")
        
    def update_rerun(self, attempt, done, total):
        self.c.itemconfig(self.dot,       fill="#FFD600")
        self.c.itemconfig(self.lbl_status,
                          text="↺ RE-RUNNING TEST",
                          fill="#FFD600")
        pct = f"   ({done/total*100:.0f}%)" if total else ""
        self.c.itemconfig(self.lbl_main,
                          text=f"Rerun #{attempt}: {done} / {total}{pct}",
                          fill="#FFD600")
        self.c.itemconfig(self.lbl_target,
                          text=f"Retrying {total} failed point(s)",
                          fill="#FFD600")
        
    def update_scenario_name(self, name):
        try: self.c.itemconfig(self.lbl_scen_name, text=name)
        except Exception: pass

    def update_paused(self, done, total, reason=""):
        self.c.itemconfig(self.dot, fill=PAUSE_COLOR)
        self.c.itemconfig(self.lbl_status, text=f"PAUSED — {reason}" if reason else "PAUSED", fill=PAUSE_COLOR)
        self.c.itemconfig(self.lbl_main, text=f"{done} / {total}   ({done/total*100:.0f}%)", fill=PAUSE_COLOR)
        self.c.itemconfig(self.lbl_target, text="Press Space to resume", fill=PAUSE_COLOR)

    def update_waiting(self, done, total, reason=""):
        self.c.itemconfig(self.dot, fill=WARN_COLOR)
        self.c.itemconfig(self.lbl_status, text=f"WAITING — {reason}", fill=WARN_COLOR)
        self.c.itemconfig(self.lbl_main, text=f"{done} / {total}   ({done/total*100:.0f}%)", fill=WARN_COLOR)
        self.c.itemconfig(self.lbl_target, text="Waiting for connection to establish...", fill=WARN_COLOR)
    def update_stopped(self):
        self.c.itemconfig(self.dot, fill=EXCLUDE_COLOR)
        self.c.itemconfig(self.lbl_status, text="STOPPED", fill=EXCLUDE_COLOR)
        self.c.itemconfig(self.lbl_main, text="Aborted.", fill=EXCLUDE_COLOR)
        self.c.itemconfig(self.lbl_target, text="", fill="#555")

class ConfirmDialog(tk.Toplevel):
    def __init__(self, master, mode, n_points, monitor_name):
        super().__init__(master)
        self.result = False
        self.title("Start Test")
        self.resizable(False, False); self.configure(bg="#0f0f0f"); self.grab_set(); self.attributes("-topmost", True)
        w, h = 440, 240
        x = master.winfo_x() + (master.winfo_width() // 2) - (w // 2)
        y = master.winfo_y() + (master.winfo_height() // 2) - (h // 2)
        self.geometry(f"{w}x{h}+{x}+{y}")
        
        self._build(mode, n_points, monitor_name)
        self.bind("<Return>", self._yes)
        self.bind("<Escape>", self._no)

    def _build(self, mode, n, mon):
        tk.Label(self, text=f"Start {mode.title()} Test", bg="#0f0f0f", fg="#fff", font=("Consolas", 13, "bold")).pack(pady=(20, 4))
        tk.Label(self, text=f"{n} targets  ·  {mon}", bg="#0f0f0f", fg="#888", font=("Consolas", 10)).pack()
        info = ("  Space      →  Pause / Resume\n  Esc        →  Stop immediately\n  Move mouse →  Auto-pause drift safety")
        tk.Label(self, text=info, bg="#0f0f0f", fg="#555", font=("Consolas", 9), justify="left").pack(pady=10)
        row = tk.Frame(self, bg="#0f0f0f")
        row.pack(pady=(0, 18))
        self.yes_btn = tk.Button(row, text="✓  Yes, Start", bg=INCLUDE_COLOR, fg="#000", font=("Consolas", 11, "bold"), relief="flat", padx=18, pady=8, cursor="hand2", command=self._yes)
        self.yes_btn.pack(side="left", padx=8)
        self.yes_btn.focus_set()
        tk.Button(row, text="✕  Cancel", bg="#222", fg="#888", font=("Consolas", 11), relief="flat", padx=18, pady=8, cursor="hand2", command=self._no).pack(side="left", padx=8)

    def _yes(self, event=None): self.result = True; self.destroy()
    def _no(self, event=None): self.result = False; self.destroy()

    @classmethod
    def ask(cls, master, mode, n_points, monitor_name):
        dlg = cls(master, mode, n_points, monitor_name)
        master.wait_window(dlg)
        return dlg.result

# ── Live OCR Monitor ──────────────────────────────────────────────────────────

class OcrOverlay(tk.Toplevel):
    """
    Full-screen transparent overlay on the selected monitor.
    User drags to draw ONE rectangle; on release the bbox is returned
    via on_done(x1, y1, x2, y2) and the overlay closes.
    """
    MIN_PX = 10

    def __init__(self, master, monitor: Monitor, on_done):
        super().__init__(master)
        self.monitor  = monitor
        self.on_done  = on_done
        self.start_x  = self.start_y = 0
        self.rect_id  = None
        self._drawing = False

        self.geometry(f"{monitor.width}x{monitor.height}+{monitor.x}+{monitor.y}")
        self.attributes("-topmost", True)
        self.attributes("-alpha", 0.35)
        self.overrideredirect(True)
        self.configure(bg="#0a0a0a")
        self._make_clickthrough_off()

        self.canvas = tk.Canvas(self, bg="#0a0a0a", highlightthickness=0,
                                cursor="crosshair",
                                width=monitor.width, height=monitor.height)
        self.canvas.pack(fill="both", expand=True)

        self.canvas.create_text(
            monitor.width // 2, 28,
            text="🔍 OCR Monitor  —  Drag to draw capture zone  |  Esc = cancel",
            fill="#aaa", font=("Consolas", 11)
        )

        self.canvas.bind("<ButtonPress-1>",   self._press)
        self.canvas.bind("<B1-Motion>",        self._drag)
        self.canvas.bind("<ButtonRelease-1>",  self._release)
        self.bind("<Escape>", lambda e: self.destroy())
        self.focus_force()

    def _make_clickthrough_off(self):
        # Make sure clicks are captured (NOT click-through)
        try:
            hwnd = ctypes.windll.user32.GetParent(self.winfo_id()) or self.winfo_id()
            style = ctypes.windll.user32.GetWindowLongW(hwnd, -20)
            # Remove WS_EX_TRANSPARENT so we receive clicks
            style = style & ~0x20
            ctypes.windll.user32.SetWindowLongW(hwnd, -20, style)
        except Exception:
            pass

    def _press(self, ev):
        self._drawing = True
        self.start_x, self.start_y = ev.x, ev.y
        if self.rect_id:
            self.canvas.delete(self.rect_id)
            self.rect_id = None

    def _drag(self, ev):
        if not self._drawing: return
        if self.rect_id:
            self.canvas.delete(self.rect_id)
        self.rect_id = self.canvas.create_rectangle(
            self.start_x, self.start_y, ev.x, ev.y,
            outline="#00BCD4", width=2, dash=(4, 3)
        )

    def _release(self, ev):
        if not self._drawing: return
        self._drawing = False
        x1 = min(self.start_x, ev.x) + self.monitor.x
        y1 = min(self.start_y, ev.y) + self.monitor.y
        x2 = max(self.start_x, ev.x) + self.monitor.x
        y2 = max(self.start_y, ev.y) + self.monitor.y
        if (x2 - x1) < self.MIN_PX or (y2 - y1) < self.MIN_PX:
            return  # too small, let user redraw
        self.destroy()
        self.on_done(x1, y1, x2, y2)


class OcrMonitorPanel(tk.Toplevel):
    """
    Floating live-OCR result panel.
    Continuously grabs bbox, runs OCR, displays raw text.
    Controls: Start/Stop toggle, Redraw (re-opens OcrOverlay), Copy, Save .txt, Open last saved.
    """
    POLL_MS    = 800   # refresh interval in milliseconds
    PANEL_W    = 540
    PANEL_H    = 420

    def __init__(self, master, monitor: Monitor):
        super().__init__(master)
        self.master_app = master
        self.monitor    = monitor
        self.bbox       = None        # (x1,y1,x2,y2) screen coords
        self._running   = False
        self._after_id  = None
        self._last_text = ""
        self._last_saved_path = ""

        self.title("🔍 Live OCR Monitor")
        self.configure(bg="#0f0f0f")
        self.resizable(True, True)
        self.attributes("-topmost", True)

        # Position: top-right of master window
        mx = master.winfo_x() + master.winfo_width() - self.PANEL_W - 20
        my = master.winfo_y() + 60
        self.geometry(f"{self.PANEL_W}x{self.PANEL_H}+{max(0,mx)}+{max(0,my)}")

        self._build()
        self.protocol("WM_DELETE_WINDOW", self._on_close)

        # Auto-open draw overlay on first open
        self.after(150, self._open_draw_overlay)

    def _build(self):
        hdr = tk.Frame(self, bg="#0f0f0f")
        hdr.pack(fill="x", padx=12, pady=(10, 4))

        tk.Label(hdr, text="🔍 LIVE OCR MONITOR", bg="#0f0f0f", fg="#00BCD4",
                 font=("Consolas", 11, "bold")).pack(side="left")

        self.lbl_status = tk.Label(hdr, text="● IDLE", bg="#0f0f0f", fg="#444",
                                   font=("Consolas", 9, "bold"))
        self.lbl_status.pack(side="right")

        self.lbl_zone = tk.Label(self, text="No zone drawn yet — click Redraw",
                                 bg="#161616", fg="#888", font=("Consolas", 8),
                                 anchor="w", padx=8, pady=4)
        self.lbl_zone.pack(fill="x", padx=12)

        bar = tk.Frame(self, bg="#111", pady=6)
        bar.pack(fill="x", padx=12, pady=(4, 6))

        bs = dict(font=("Consolas", 9, "bold"), relief="flat",
                  padx=10, pady=5, cursor="hand2")

        self.btn_toggle = tk.Button(bar, text="▶ Start", bg="#00C853", fg="#000",
                                    command=self._toggle_live, **bs)
        self.btn_toggle.pack(side="left", padx=(0, 4))

        tk.Button(bar, text="✏ Redraw", bg="#00BCD4", fg="#000",
                  command=self._open_draw_overlay, **bs).pack(side="left", padx=4)

        # UPGRADE: UI Dropdown to switch layout logic on the fly
        tk.Label(bar, text="Mode:", bg="#111", fg="#888", font=("Consolas", 9)).pack(side="left", padx=(6, 2))
        self.v_layout = tk.StringVar(value="tabular")  # Defaulting to tabular for SCADA
        self.opt_mode = tk.OptionMenu(bar, self.v_layout, "tabular", "block", "single_line", "sparse")
        self.opt_mode.config(bg="#222", fg="#fff", font=("Consolas", 9), relief="flat", bd=0, highlightthickness=0)
        self.opt_mode["menu"].config(bg="#222", fg="#fff", font=("Consolas", 9))
        self.opt_mode.pack(side="left", padx=2)

        tk.Button(bar, text="⎘ Copy", bg="#333", fg="#ccc",
                  command=self._copy_text, **bs).pack(side="left", padx=4)

        tk.Button(bar, text="💾 Save .txt", bg="#333", fg="#ccc",
                  command=self._save_txt, **bs).pack(side="left", padx=4)

        self.btn_open = tk.Button(bar, text="📄 Open", bg="#222", fg="#555",
                                  command=self._open_last, state="disabled", **bs)
        self.btn_open.pack(side="left", padx=4)

        tk.Label(bar, text=f"~{self.POLL_MS}ms", bg="#111", fg="#444",
                 font=("Consolas", 8)).pack(side="right")

        txt_frame = tk.Frame(self, bg="#0f0f0f")
        txt_frame.pack(fill="both", expand=True, padx=12, pady=(0, 10))

        self.txt = tk.Text(
            txt_frame, bg="#080808", fg="#00FF9C",
            font=("Consolas", 11), relief="flat",
            wrap="word", state="disabled",
            insertbackground="#00FF9C",
            selectbackground="#1a3a2a", selectforeground="#fff"
        )
        sb = tk.Scrollbar(txt_frame, command=self.txt.yview)
        self.txt.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self.txt.pack(fill="both", expand=True)

    # ── Draw overlay ───────────────────────────────────────────────────────────
    def _open_draw_overlay(self):
        was_running = self._running
        self._stop_live()
        self.withdraw()

        def on_bbox_done(x1, y1, x2, y2):
            self.bbox = (x1, y1, x2, y2)
            self.lbl_zone.config(
                text=f"Zone: ({x1}, {y1}) → ({x2}, {y2})   "
                     f"{x2-x1}×{y2-y1} px",
                fg="#00BCD4"
            )
            self.deiconify()
            self.lift()
            self._start_live()  # auto-start immediately after drawing

        # Hide master briefly so overlay can see the screen cleanly
        try: self.master_app.withdraw()
        except Exception: pass
        self.after(180, lambda: self._launch_overlay(on_bbox_done))

    def _launch_overlay(self, on_bbox_done):
        # Always use the currently selected monitor, not the one at panel-open time
        try:
            mon = self.master_app.active_mon
        except Exception:
            mon = self.monitor

        def _done(x1, y1, x2, y2):
            try: self.master_app.deiconify()
            except Exception: pass
            on_bbox_done(x1, y1, x2, y2)

        def _cancelled():
            try: self.master_app.deiconify()
            except Exception: pass
            self.deiconify()

        ov = OcrOverlay(self, mon, _done)
        ov.protocol("WM_DELETE_WINDOW", lambda: (_cancelled(), ov.destroy()))

    # ── Live polling ───────────────────────────────────────────────────────────
    def _toggle_live(self):
        if self._running:
            self._stop_live()
        else:
            self._start_live()

    def _start_live(self):
        if self.bbox is None:
            self._set_text("⚠  Draw a zone first (click ✏ Redraw).")
            return
        if not PIL_AVAILABLE:
            self._set_text("⚠  Pillow (PIL) not installed.")
            return
        if not TESSERACT_AVAILABLE:
            self._set_text("⚠  Tesseract not available. Check Settings → Tesseract Path.")
            return
        self._running = True
        self.btn_toggle.config(text="⏹ Stop", bg=EXCLUDE_COLOR, fg="#fff")
        self.lbl_status.config(text="● LIVE", fg="#00C853")
        self._poll()

    def _stop_live(self):
        self._running = False
        if self._after_id:
            try: self.after_cancel(self._after_id)
            except Exception: pass
            self._after_id = None
        self.btn_toggle.config(text="▶ Start", bg="#00C853", fg="#000")
        self.lbl_status.config(text="● IDLE", fg="#444")

    def _poll(self):
        if not self._running: return
        try:
            img  = ImageGrab.grab(bbox=self.bbox, all_screens=True)
            lang = APP_CONFIG.get("tesseract_lang", "eng")
            
            # UPGRADE: Pass selected layout directly into the poll execution
            current_layout = self.v_layout.get()
            raw  = ocr_run(img, lang=lang, layout=current_layout)
            
            self._last_text = raw
            self._set_text(raw if raw.strip() else "(no text detected in zone)")
            self.lbl_status.config(text="● LIVE", fg="#00C853")
        except Exception as e:
            self.lbl_status.config(text="● ERR", fg=EXCLUDE_COLOR)
            self._set_text(f"OCR error:\n{e}")

        if self._running:
            self._after_id = self.after(self.POLL_MS, self._poll)

    # ── Text helpers ───────────────────────────────────────────────────────────
    def _set_text(self, text: str):
        self.txt.configure(state="normal")
        self.txt.delete("1.0", "end")
        self.txt.insert("1.0", text)
        self.txt.configure(state="disabled")

    def _copy_text(self):
        text = self._last_text or self.txt.get("1.0", "end").strip()
        if text:
            self.clipboard_clear()
            self.clipboard_append(text)
            self.lbl_status.config(text="✓ Copied!", fg="#00C853")
            self.after(1200, lambda: self.lbl_status.config(
                text="● LIVE" if self._running else "● IDLE",
                fg="#00C853" if self._running else "#444"
            ))

    def _save_txt(self):
        text = self._last_text or self.txt.get("1.0", "end").strip()
        if not text:
            return
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        path = LOG_DIR / f"ocr_capture_{ts}.txt"
        try:
            with open(path, "w", encoding="utf-8") as f:
                if self.bbox:
                    f.write(f"# OCR Capture — {ts}\n")
                    f.write(f"# Zone: {self.bbox}\n\n")
                f.write(text)
            self._last_saved_path = str(path)
            self.btn_open.config(state="normal", fg="#ccc")
            self.lbl_status.config(text="✓ Saved!", fg="#00C853")
            self.after(1500, lambda: self.lbl_status.config(
                text="● LIVE" if self._running else "● IDLE",
                fg="#00C853" if self._running else "#444"
            ))
        except Exception as e:
            self.lbl_status.config(text=f"Save failed: {e}", fg=EXCLUDE_COLOR)

    def _open_last(self):
        if self._last_saved_path and os.path.exists(self._last_saved_path):
            try:
                os.startfile(self._last_saved_path)   # Windows — opens in Notepad
            except Exception:
                import subprocess
                subprocess.Popen(["notepad", self._last_saved_path])

    # ── Cleanup ────────────────────────────────────────────────────────────────
    def _on_close(self):
        self._stop_live()
        self.destroy()


class OverlayWindow(tk.Toplevel):
    EDGE_HIT = 10
    ISCS_TYPES = ["alarm_panel", "equipment_page", "alarm_list", "event_list", "anchor"]

    def __init__(self, master, app_mode, zones, monitor: Monitor, grid_spacing: int, on_done, pages: list = None, zones_per_page: dict = None):
        super().__init__(master)
        self.app_mode = app_mode
        self.zones = zones
        self.monitor = monitor
        self.grid_spacing = grid_spacing
        self.on_done = on_done
        self.current_type = "target" if app_mode == "sequence" else ("alarm_panel" if app_mode == "iscs" else "include")

        # Per-page zone storage for ISCS mode
        self.pages = pages or []  # list of {"name": ..., "x": ..., "y": ...}
        self.zones_per_page = zones_per_page or {}  # {"Page Name": {"alarm_panel": Zone, ...}}
        self._current_page = self.pages[0]["name"] if self.pages else "Global"
        # Ensure current page exists in dict
        if self._current_page not in self.zones_per_page:
            self.zones_per_page[self._current_page] = {}

        phys_rects = get_physical_monitor_rects()
        phys = match_physical_rect(monitor, phys_rects)
        self.phys_x, self.phys_y = (phys[0], phys[1]) if phys else (monitor.x, monitor.y)

        self.drawing = False
        self.start_x = self.start_y = 0
        self.new_rect = None
        self.drag_zone = self.drag_mode = None
        self.drag_start_state = None
        self.last_touched_zone = None
        self.zone_items = {}
        self._undo_stack = []

        self.geometry(f"{monitor.width}x{monitor.height}+{monitor.x}+{monitor.y}")
        self.attributes("-topmost", True); self.attributes("-alpha", 0.82); self.overrideredirect(True)
        self.configure(bg="#0a0a0a")

        self._build_canvas()
        self._build_toolbar()
        self._bind_events()
        self.after(50, self._redraw_existing_zones)
        self.after(60, self._sync_undo_btn)
        self.focus_force()

    def _build_canvas(self):
        self.canvas = tk.Canvas(self, bg="#0a0a0a", cursor="crosshair", highlightthickness=0)
        self.canvas.pack(fill=tk.BOTH, expand=True)
        mode_str = "Targeted Sequence" if self.app_mode == "sequence" else ("ISCS Modbus" if self.app_mode == "iscs" else "Grid Scan")
        self.canvas.create_text(
            self.monitor.width // 2, 22,
            text=f"Mode: {mode_str}  │  Drag to draw  │  Drag edge=Resize  Drag body=Move  Right-click=Options  │  Display {self.monitor.display_num}",
            fill="#555", font=("Consolas", 10)
        )

    def _build_toolbar(self):
        bar = tk.Frame(self, bg="#111111", pady=6)
        bar.place(relx=0.5, rely=0.97, anchor="s")
        s = dict(font=("Consolas", 11, "bold"), relief="flat", padx=14, pady=6, cursor="hand2", bd=0)
        
        if self.app_mode == "grid":
            tk.Button(bar, text="● INCLUDE", bg=INCLUDE_COLOR, fg="#000", command=lambda: self._set_type("include"), **s).pack(side="left", padx=4)
            tk.Button(bar, text="● EXCLUDE", bg=EXCLUDE_COLOR, fg="#fff", command=lambda: self._set_type("exclude"), **s).pack(side="left", padx=4)
        elif self.app_mode == "sequence":
            tk.Button(bar, text="🎯 TARGET BOX", bg=TARGET_COLOR, fg="#fff", command=lambda: self._set_type("target"), **s).pack(side="left", padx=4)
        elif self.app_mode == "iscs":
            # ── Page selector ─────────────────────────────────────────────────
            page_names = [p["name"] for p in self.pages] if self.pages else ["Global"]
            if "Global" not in page_names:
                page_names = ["Global"] + page_names

            tk.Label(bar, text="Page:", bg="#111111", fg="#888",
                     font=("Consolas", 9)).pack(side="left", padx=(8, 2))
            self._page_var = tk.StringVar(value=self._current_page)
            page_menu = tk.OptionMenu(bar, self._page_var, *page_names, command=self._on_page_change)
            page_menu.config(bg="#222", fg="#fff", font=("Consolas", 9, "bold"),
                             relief="flat", bd=0, highlightthickness=0,
                             activebackground="#333", activeforeground="#fff")
            page_menu["menu"].config(bg="#222", fg="#fff", font=("Consolas", 9))
            page_menu.pack(side="left", padx=(0, 10))

            # ── Zone type buttons with Save/Show ──────────────────────────────
            iscs_btns = [
                ("alarm_panel",    "🚨 ALARM PANEL",    ALARM_PANEL_COLOR, "#fff"),
                ("equipment_page", "🖥 EQUIP PAGE",     EQUIP_ZONE_COLOR,  "#000"),
                ("alarm_list",     "📋 ALARM LIST",     ALARM_LIST_COLOR,  "#000"),
                ("event_list",     "📅 EVENT LIST",     EVENT_LIST_COLOR,  "#000"),
                ("anchor",         "⚓ ANCHOR",          ANCHOR_COLOR,      "#000"),
            ]
            self._iscs_btns = {}
            for zt, lbl, color, fg in iscs_btns:
                # Container frame per zone type
                grp = tk.Frame(bar, bg="#1a1a1a", padx=2, pady=2)
                grp.pack(side="left", padx=4)

                # Main zone type button
                btn = tk.Button(grp, text=lbl,
                                bg=color if self.current_type == zt else "#333",
                                fg=fg if self.current_type == zt else color,
                                font=("Consolas", 10, "bold"), relief="flat",
                                padx=10, pady=5, cursor="hand2",
                                command=lambda t=zt: self._set_type(t))
                btn.pack(side="top", fill="x")

                # Save / Show mini buttons
                mini_row = tk.Frame(grp, bg="#1a1a1a")
                mini_row.pack(side="top", fill="x")
                ms = dict(font=("Consolas", 8, "bold"), relief="flat",
                          padx=6, pady=2, cursor="hand2", bd=0)

                saved_indicator = tk.Label(mini_row, text="", bg="#1a1a1a",
                                           fg="#00C853", font=("Consolas", 8))
                saved_indicator.pack(side="right", padx=2)

                tk.Button(mini_row, text="💾 Save", bg="#1e3a1e", fg="#00C853",
                          command=lambda t=zt, ind=saved_indicator: self._save_zone_type(t, ind),
                          **ms).pack(side="left", padx=(0, 1))
                show_btn = tk.Button(mini_row, text="👁 Show", bg="#1a1a2e", fg="#5599ff",
                          **ms)
                show_btn.config(command=lambda t=zt, b=show_btn: self._show_zone_type(t, b))
                show_btn.pack(side="left")

                self._iscs_btns[zt] = (btn, color, fg, saved_indicator)

            # Update saved indicators for already-saved zones
            self.after(100, self._refresh_saved_indicators)

        self.btn_undo = tk.Button(bar, text="↩ Undo", bg="#333", fg="#ccc", command=self._undo, **s)
        self.btn_undo.pack(side="left", padx=4)
        tk.Button(bar, text="🗑 Delete", bg="#8B0000", fg="#fff", command=self._delete_last, **s).pack(side="left", padx=4)
        tk.Button(bar, text="✕ Clear", bg="#333", fg="#ccc", command=self._clear_all,      **s).pack(side="left", padx=4)
        tk.Button(bar, text="✓ Done",  bg="#2979FF", fg="#fff", command=self._finish,       **s).pack(side="left", padx=4)

    def _set_type(self, t):
        self.current_type = t
        if hasattr(self, '_iscs_btns'):
            for zt, entry in self._iscs_btns.items():
                btn, color, fg = entry[0], entry[1], entry[2]
                if zt == t: btn.config(bg=color, fg=fg)
                else: btn.config(bg="#333", fg=color)

    def _on_page_change(self, page_name):
        """Switch active page — clears canvas of current page zones, loads new page."""
        # Save any unsaved drawn zones for current page before switching
        self._current_page = page_name
        if page_name not in self.zones_per_page:
            self.zones_per_page[page_name] = {}
        # Clear all zones from canvas
        for z in list(self.zones):
            self._erase_zone_items(z)
        self.zones.clear()
        self._undo_stack.clear()
        self._sync_undo_btn()
        self._refresh_saved_indicators()

    def _save_zone_type(self, zone_type: str, indicator: tk.Label):
        """Save the currently drawn zone of this type for the current page, then clear it from canvas."""
        # Find the zone of this type on canvas
        matching = [z for z in self.zones if z.zone_type == zone_type]
        if not matching:
            indicator.config(text="nothing drawn")
            self.after(1500, lambda: indicator.config(text="✓ saved" if zone_type in self.zones_per_page.get(self._current_page, {}) else ""))
            return
        # Use the last drawn one if multiple
        zone = matching[-1]
        # Store in zones_per_page
        if self._current_page not in self.zones_per_page:
            self.zones_per_page[self._current_page] = {}
        self.zones_per_page[self._current_page][zone_type] = zone
        # Auto-save to global template
        _save_template({"zones": {zone_type: zone.to_dict()}})
        # Remove from active canvas zones
        for z in matching:
            self._erase_zone_items(z)
            if z in self.zones:
                self.zones.remove(z)
        self._undo_stack = [a for a in self._undo_stack if not (a[0] in ("draw", "delete") and a[1].zone_type == zone_type)]
        self._sync_undo_btn()
        indicator.config(text="✓ saved")

    def _show_zone_type(self, zone_type: str, btn: tk.Button):
        """Toggle saved zone for this type/page on/off on canvas."""
        # If already showing on canvas, hide it (save back)
        on_canvas = [z for z in self.zones if z.zone_type == zone_type]
        if on_canvas:
            # Currently showing — hide it (save back to zones_per_page)
            zone = on_canvas[-1]
            if self._current_page not in self.zones_per_page:
                self.zones_per_page[self._current_page] = {}
            self.zones_per_page[self._current_page][zone_type] = zone
            for z in on_canvas:
                self._erase_zone_items(z)
                if z in self.zones:
                    self.zones.remove(z)
            self._undo_stack = [a for a in self._undo_stack
                                if not (a[0] in ("draw", "delete") and a[1].zone_type == zone_type)]
            self._sync_undo_btn()
            btn.config(text="👁 Show", bg="#1a1a2e", fg="#5599ff")
            self._refresh_saved_indicators()
            return

        # Not on canvas — load from saved
        page_zones = self.zones_per_page.get(self._current_page, {})
        saved_zone = page_zones.get(zone_type)
        if saved_zone is None:
            return
        # Remove from saved store — now live on canvas
        del self.zones_per_page[self._current_page][zone_type]
        self.zones.append(saved_zone)
        self._draw_zone(saved_zone)
        self._push_undo(("draw", saved_zone))
        btn.config(text="🙈 Hide", bg="#2e1a2e", fg="#cc88ff")
        self._refresh_saved_indicators()

    def _refresh_saved_indicators(self):
        """Update the ✓ saved indicators and Show/Hide button state."""
        if not hasattr(self, '_iscs_btns'):
            return
        page_zones = self.zones_per_page.get(self._current_page, {})
        on_canvas_types = {z.zone_type for z in self.zones}
        for zt, entry in self._iscs_btns.items():
            indicator = entry[3]
            indicator.config(text="✓ saved" if zt in page_zones else "")

    def _redraw_existing_zones(self):
        for z in self.zones: self._draw_zone(z)

    def _bind_events(self):
        c = self.canvas
        c.bind("<ButtonPress-1>", self._on_press)
        c.bind("<B1-Motion>", self._on_drag)
        c.bind("<ButtonRelease-1>", self._on_release)
        c.bind("<ButtonPress-3>", self._on_right_click)
        c.bind("<Motion>", self._on_hover)
        self.bind("<Escape>",    lambda e: self._finish())
        self.bind("<z>",         lambda e: self._undo())
        self.bind("<Delete>",    lambda e: self._delete_last())
        self.bind("<BackSpace>", lambda e: self._delete_last())

    def _canvas_to_abs(self, cx, cy): return cx + self.winfo_rootx(), cy + self.winfo_rooty()
    def _abs_to_canvas(self, ax, ay): return ax - self.winfo_rootx(), ay - self.winfo_rooty()

    def _hit_zone(self, cx, cy):
        eh = self.EDGE_HIT
        for zone in reversed(self.zones):
            zx1, zy1 = self._abs_to_canvas(zone.x1, zone.y1)
            zx2, zy2 = self._abs_to_canvas(zone.x2, zone.y2)
            on_left, on_right = abs(cx - zx1) < eh, abs(cx - zx2) < eh
            on_top, on_bottom = abs(cy - zy1) < eh, abs(cy - zy2) < eh
            in_x, in_y = zx1 - eh < cx < zx2 + eh, zy1 - eh < cy < zy2 + eh
            if not (in_x and in_y): continue
            if on_top and on_left: return zone, "nw"
            if on_top and on_right: return zone, "ne"
            if on_bottom and on_left: return zone, "sw"
            if on_bottom and on_right: return zone, "se"
            if on_top and in_x: return zone, "n"
            if on_bottom and in_x: return zone, "s"
            if on_left and in_y: return zone, "w"
            if on_right and in_y: return zone, "e"
            if zx1 + eh < cx < zx2 - eh and zy1 + eh < cy < zy2 - eh: return zone, "move"
        return None, None

    def _cursor_for_mode(self, mode):
        cursors = {"move": "fleur", "n": "sb_v_double_arrow", "s": "sb_v_double_arrow", "e": "sb_h_double_arrow", "w": "sb_h_double_arrow", "ne": "size_ne_sw", "sw": "size_ne_sw", "nw": "size_nw_se", "se": "size_nw_se"}
        return cursors.get(mode, "crosshair")

    def _on_hover(self, ev):
        zone, mode = self._hit_zone(ev.x, ev.y)
        self.canvas.config(cursor=self._cursor_for_mode(mode) if zone else "crosshair")

    def _on_press(self, ev):
        zone, mode = self._hit_zone(ev.x, ev.y)
        if zone and mode:
            self.drag_zone, self.drag_mode, self.drag_ox, self.drag_oy = zone, mode, ev.x, ev.y
            self.last_touched_zone = zone
            self.drag_start_state = (zone.x1, zone.y1, zone.x2, zone.y2)
        else:
            self.drag_start_state = None
            self.drawing = True
            self.start_x, self.start_y = self._canvas_to_abs(ev.x, ev.y)
            color = TARGET_COLOR if self.current_type == "target" else (ALARM_PANEL_COLOR if self.current_type == "alarm_panel" else (EQUIP_ZONE_COLOR if self.current_type == "equipment_page" else (ALARM_LIST_COLOR if self.current_type == "alarm_list" else (EVENT_LIST_COLOR if self.current_type == "event_list" else (ANCHOR_COLOR if self.current_type == "anchor" else (INCLUDE_COLOR if self.current_type == "include" else EXCLUDE_COLOR))))))
            self.new_rect = self.canvas.create_rectangle(ev.x, ev.y, ev.x, ev.y, outline=color, width=2, dash=(4, 4))

    def _on_drag(self, ev):
        if self.drag_zone:
            dx, dy = ev.x - self.drag_ox, ev.y - self.drag_oy
            self.drag_ox, self.drag_oy = ev.x, ev.y
            z, m = self.drag_zone, self.drag_mode
            if m == "move": z.x1 += dx; z.x2 += dx; z.y1 += dy; z.y2 += dy
            elif m == "n": z.y1 += dy
            elif m == "s": z.y2 += dy
            elif m == "w": z.x1 += dx
            elif m == "e": z.x2 += dx
            elif m == "nw": z.x1 += dx; z.y1 += dy
            elif m == "ne": z.x2 += dx; z.y1 += dy
            elif m == "sw": z.x1 += dx; z.y2 += dy
            elif m == "se": z.x2 += dx; z.y2 += dy
            if z.x1 > z.x2: z.x1, z.x2 = z.x2, z.x1
            if z.y1 > z.y2: z.y1, z.y2 = z.y2, z.y1
            self._erase_zone_items(z); self._draw_zone(z); self._check_zone_size(z)
        elif self.drawing and self.new_rect:
            sx, sy = self._abs_to_canvas(self.start_x, self.start_y)
            self.canvas.coords(self.new_rect, sx, sy, ev.x, ev.y)

    def _on_release(self, ev):
        if self.drag_zone:
            self._erase_zone_items(self.drag_zone); self._draw_zone(self.drag_zone); self._check_zone_size(self.drag_zone)
            if self.drag_start_state:
                ox1, oy1, ox2, oy2 = self.drag_start_state
                z = self.drag_zone
                if (z.x1, z.y1, z.x2, z.y2) != (ox1, oy1, ox2, oy2):
                    self._push_undo(("move", z, ox1, oy1, ox2, oy2))
            self.drag_zone = None; self.drag_start_state = None; return
        if not self.drawing: return
        self.drawing = False
        x2, y2 = self._canvas_to_abs(ev.x, ev.y)
        if abs(x2 - self.start_x) < MIN_ZONE_PX or abs(y2 - self.start_y) < MIN_ZONE_PX:
            if self.new_rect: self.canvas.delete(self.new_rect)
            return
        mon_idx = getattr(self.monitor, 'index', 0)
        z = Zone(self.start_x, self.start_y, x2, y2, self.current_type, monitor_index=mon_idx)
        self.zones.append(z)
        if self.new_rect: self.canvas.delete(self.new_rect)
        self._draw_zone(z); self._check_zone_size(z)
        self._push_undo(("draw", z))

        # ── Feature 1: Visual Anchoring — capture anchor crop on draw ─────────
        if z.zone_type == "anchor" and UPGRADES_AVAILABLE:
            if not hasattr(self, '_anchor_mgr') or self._anchor_mgr is None:
                # Use BASE_DIR as fallback if no session dir available yet
                anchor_dir = getattr(self, '_scenario_dir', BASE_DIR / "anchors")
                self._anchor_mgr = AnchorManager(Path(anchor_dir))
                self._anchor_mgr.load()
            anchor_name = f"anchor_{sum(1 for zz in self.zones if zz.zone_type == 'anchor')}"
            anchor = VisualAnchor.create_from_zone(z, Path(getattr(self, '_scenario_dir', BASE_DIR / "anchors")), name=anchor_name)
            if anchor:
                self._anchor_mgr.register_anchor(anchor)
                self._anchor_mgr.save()
                z.label = anchor_name   # store name on zone for reference
                self._erase_zone_items(z); self._draw_zone(z)
                logger.info(f"OverlayWindow: captured anchor '{anchor_name}' at ({z.cx},{z.cy})")

    def _on_right_click(self, ev):
        zone, _ = self._hit_zone(ev.x, ev.y)
        if not zone: return
        menu = tk.Menu(self, tearoff=0, bg="#1a1a1a", fg="#ccc", font=("Consolas", 10))
        if self.app_mode == "grid":
            other = "exclude" if zone.zone_type == "include" else "include"
            menu.add_command(label=f"Switch to {other.upper()}", command=lambda z=zone: self._toggle_zone_type(z))
            menu.add_separator()
        elif self.app_mode == "iscs":
            for zt in self.ISCS_TYPES:
                if zt != zone.zone_type:
                    label = zt.replace("_", " ").title()
                    menu.add_command(label=f"→ Change to {label}", command=lambda z=zone, t=zt: self._change_zone_type(z, t))
            menu.add_separator()
            # ── Feature 1: Link zone to an anchor ─────────────────────────────
            if zone.zone_type != "anchor" and UPGRADES_AVAILABLE:
                anchors_available = [zz for zz in self.zones if zz.zone_type == "anchor"]
                if anchors_available:
                    menu.add_command(
                        label="⚓ Link to anchor…",
                        command=lambda z=zone: self._link_zone_to_anchor(z)
                    )
            menu.add_separator()
            
        menu.add_command(label="🗑 Delete zone", command=lambda z=zone: self._delete_zone(z))
        menu.tk_popup(ev.x_root, ev.y_root)

    def _toggle_zone_type(self, zone):
        zone.zone_type = "exclude" if zone.zone_type == "include" else "include"
        self._erase_zone_items(zone); self._draw_zone(zone); self._check_zone_size(zone)
        
    def _change_zone_type(self, zone, new_type):
        zone.zone_type = new_type
        self._erase_zone_items(zone); self._draw_zone(zone)

    def _link_zone_to_anchor(self, zone):
        """Feature 1: Show a picker to choose which anchor this zone tracks."""
        if not UPGRADES_AVAILABLE:
            return
        anchor_zones = [z for z in self.zones if z.zone_type == "anchor"]
        if not anchor_zones:
            from tkinter import messagebox as _mb
            _mb.showinfo("No Anchors", "Draw an ⚓ ANCHOR zone first.", parent=self)
            return

        if not hasattr(self, '_anchor_mgr') or self._anchor_mgr is None:
            from tkinter import messagebox as _mb
            _mb.showinfo("No Anchors", "No anchor manager available. Draw an anchor zone first.", parent=self)
            return

        # If only one anchor, use it automatically
        if len(anchor_zones) == 1:
            az = anchor_zones[0]
            anchor_name = az.label or "anchor_1"
            anchor = self._anchor_mgr._anchors.get(anchor_name)
            if anchor:
                self._anchor_mgr.link_zone(anchor, zone)
                self._anchor_mgr.save()
                from tkinter import messagebox as _mb
                _mb.showinfo(
                    "Anchor Linked",
                    f"Zone '{zone.zone_type}' is now linked to anchor '{anchor_name}'.\n"
                    f"Offset: dx={anchor.abs_cx - zone.x1:+d}  dy={anchor.abs_cy - zone.y1:+d}",
                    parent=self
                )
            return

        # Multiple anchors — show a simple picker dialog
        dlg = tk.Toplevel(self)
        dlg.title("Choose Anchor")
        dlg.configure(bg="#0f0f0f")
        dlg.attributes("-topmost", True)
        dlg.grab_set()
        dlg.geometry("320x200")
        tk.Label(dlg, text="Select anchor to link this zone to:",
                 bg="#0f0f0f", fg="#ccc", font=("Consolas", 10)).pack(pady=(14, 6))
        lb = tk.Listbox(dlg, bg="#1a1a1a", fg="#fff", font=("Consolas", 10),
                        selectbackground="#2979FF", relief="flat")
        lb.pack(fill="both", expand=True, padx=16, pady=4)
        for az in anchor_zones:
            lb.insert("end", az.label or f"anchor_{anchor_zones.index(az)+1}")
        lb.selection_set(0)

        def _confirm():
            sel = lb.curselection()
            if not sel:
                dlg.destroy(); return
            chosen_name = lb.get(sel[0])
            anchor = self._anchor_mgr._anchors.get(chosen_name)
            if anchor:
                self._anchor_mgr.link_zone(anchor, zone)
                self._anchor_mgr.save()
            dlg.destroy()

        tk.Button(dlg, text="✓ Link", bg="#2979FF", fg="#fff",
                  font=("Consolas", 10, "bold"), relief="flat",
                  padx=14, pady=6, command=_confirm, cursor="hand2").pack(pady=8)

    def _delete_zone(self, zone):
        idx = self.zones.index(zone)
        self._erase_zone_items(zone)
        if zone in self.zones:
            self.zones.remove(zone)
        self._push_undo(("delete", zone, idx))
        if zone is self.last_touched_zone:
            self.last_touched_zone = None
        if self.app_mode == "sequence":
            for z in self.zones: self._erase_zone_items(z)
            for z in self.zones: self._draw_zone(z)
        self._sync_undo_btn()

    def _delete_last(self):
        if self.last_touched_zone and self.last_touched_zone in self.zones:
            self._delete_zone(self.last_touched_zone)
            return
        if self.zones:
            self._delete_zone(self.zones[-1])

    def _push_undo(self, action):
        self._undo_stack.append(action)
        if len(self._undo_stack) > UNDO_LIMIT: self._undo_stack.pop(0)
        self._sync_undo_btn()

    def _sync_undo_btn(self):
        if hasattr(self, "btn_undo"):
            self.btn_undo.config(state="normal" if self._undo_stack else "disabled", fg="#ccc" if self._undo_stack else "#444")

    def _undo(self):
        if not self._undo_stack: return
        action = self._undo_stack.pop()

        if action[0] == "draw":
            zone = action[1]
            self._erase_zone_items(zone)
            if zone in self.zones: self.zones.remove(zone)
        elif action[0] == "delete":
            _, zone, idx = action
            self.zones.insert(idx, zone)
            self._draw_zone(zone); self._check_zone_size(zone)
            if self.app_mode == "sequence":
                for z in self.zones: self._erase_zone_items(z)
                for z in self.zones: self._draw_zone(z)
        elif action[0] == "move":
            _, zone, ox1, oy1, ox2, oy2 = action
            zone.x1, zone.y1, zone.x2, zone.y2 = ox1, oy1, ox2, oy2
            self._erase_zone_items(zone); self._draw_zone(zone); self._check_zone_size(zone)

        self._sync_undo_btn()

    def _erase_zone_items(self, zone):
        user_items = self.zone_items.pop(id(zone), {})
        for k in ("fill", "border", "label", "dot", "warn"):
            if user_items.get(k): self.canvas.delete(user_items[k])
        for hid in user_items.get("handles", []): self.canvas.delete(hid)

    def _draw_zone(self, zone):
        old = self.zone_items.pop(id(zone), {})
        for k in ("fill", "border", "label", "dot", "warn"):
            if old.get(k): self.canvas.delete(old[k])
        for hid in old.get("handles", []): self.canvas.delete(hid)

        c = self.canvas
        if zone.zone_type == "target":
            color = TARGET_COLOR
            targets = [z for z in self.zones if z.zone_type == "target"]
            idx = targets.index(zone) + 1 if zone in targets else "?"
            lbl_text = f"TARGET {idx}"
        elif zone.zone_type == "alarm_panel":
            color = ALARM_PANEL_COLOR
            lbl_text = "ALARM PANEL"
        elif zone.zone_type == "equipment_page":
            color = EQUIP_ZONE_COLOR
            lbl_text = "EQUIPMENT PAGE"
        elif zone.zone_type == "alarm_list":
            color = ALARM_LIST_COLOR
            lbl_text = "ALARM LIST"
        elif zone.zone_type == "event_list":
            color = EVENT_LIST_COLOR
            lbl_text = "EVENT LIST"
        elif zone.zone_type == "anchor":
            color = ANCHOR_COLOR
            lbl_text = f"⚓ ANCHOR{(' — ' + zone.label) if zone.label else ''}"
        else:
            color    = INCLUDE_COLOR if zone.zone_type == "include" else EXCLUDE_COLOR
            lbl_text = zone.zone_type.upper()

        cx1, cy1 = self._abs_to_canvas(zone.x1, zone.y1)
        cx2, cy2 = self._abs_to_canvas(zone.x2, zone.y2)
        mx, my   = (cx1 + cx2) // 2, (cy1 + cy2) // 2

        i = {}
        i["fill"]   = c.create_rectangle(cx1, cy1, cx2, cy2, fill=color, outline="", stipple="gray25")
        i["border"] = c.create_rectangle(cx1, cy1, cx2, cy2, fill="", outline=color, width=2)
        i["label"]  = c.create_text(mx, my - (8 if zone.zone_type == "target" else 0), text=lbl_text, fill=color, font=("Consolas", 10, "bold"))
        if zone.zone_type == "target":
            i["dot"] = c.create_oval(mx-4, my-4, mx+4, my+4, fill="#fff", outline=color, width=2)

        h_ids = []
        for hx, hy in [(cx1, cy1), (mx, cy1), (cx2, cy1), (cx1, my), (cx2, my), (cx1, cy2), (mx, cy2), (cx2, cy2)]:
            h_ids.append(c.create_rectangle(hx - HANDLE_SIZE//2, hy - HANDLE_SIZE//2, hx + HANDLE_SIZE//2, hy + HANDLE_SIZE//2, fill=color, outline="#000", width=1))
        i["handles"] = h_ids
        i["warn"]    = None
        self.zone_items[id(zone)] = i

    def _check_zone_size(self, zone):
        if self.app_mode in ("sequence", "iscs"): return
        i = self.zone_items.get(id(zone))
        if not i: return
        has_pts = zone_has_points(zone, self.monitor, self.grid_spacing)
        cx1, cy1 = self._abs_to_canvas(zone.x1, zone.y1)
        cx2, cy2 = self._abs_to_canvas(zone.x2, zone.y2)
        if i.get("warn"): self.canvas.delete(i["warn"]); i["warn"] = None
        if not has_pts and zone.zone_type == "include":
            i["warn"] = self.canvas.create_text((cx1 + cx2) // 2, cy2 + 14, text=f"⚠ Too small — no points", fill=WARN_COLOR, font=("Consolas", 9, "bold"))

    def _clear_all(self):
        for z in list(self.zones): self._erase_zone_items(z)
        self.zones.clear()
        self._undo_stack.clear()
        self._sync_undo_btn()

    def _finish(self):
        # Flush any zones still on canvas (shown/not yet saved) into zones_per_page
        for z in self.zones:
            if z.zone_type in ("alarm_panel", "equipment_page", "alarm_list", "event_list"):
                self.zones_per_page.setdefault(self._current_page, {})[z.zone_type] = z
        self.on_done(self.zones_per_page)
        self.destroy()

class HelpInspectorPanel(tk.Toplevel):
    def __init__(self, master, tool_title: str):
        super().__init__(master)
        self.tool_title = tool_title
        self.title(f"Info — {tool_title}")
        self.configure(bg="#0f0f0f")
        self.resizable(True, True)
        self.attributes("-topmost", True)
        w, h = 600, 400
        x = master.winfo_x() + (master.winfo_width() // 2) - (w // 2)
        y = master.winfo_y() + (master.winfo_height() // 2) - (h // 2)
        self.geometry(f"{w}x{h}+{x}+{y}")

        self.main = tk.Frame(self, bg="#0f0f0f", padx=10, pady=10)
        self.main.pack(fill="both", expand=True)

        self.lbl_section = tk.Label(self.main, text="(Select a section)", bg="#0f0f0f", fg="#fff", font=("Consolas", 13, "bold"))
        self.lbl_section.pack(anchor="w")

        self.txt_description = tk.Text(self.main, height=10, bg="#0a0a0a", fg="#ddd", font=("Consolas", 10), wrap="word")
        self.txt_description.pack(fill="both", expand=True, pady=(6, 10))

        bot = tk.Frame(self.main, bg="#0f0f0f")
        bot.pack(fill="x", pady=(0, 6))

        self.btn_ack = tk.Button(bot, text="Acknowledge", bg="#2979FF", fg="#fff", relief="flat", padx=14, pady=8, cursor="hand2", command=self._on_ack)
        self.btn_ack.pack(side="right")
        self.btn_next = tk.Button(bot, text="Next →", bg="#222", fg="#ccc", relief="flat", padx=14, pady=8, cursor="hand2", command=self._on_next)
        self.btn_next.pack(side="right", padx=(0, 8))

        self._acknowledged = False
        self._sections_order = []
        self._section_map = {}
        self._current_idx = 0

        self._build_section_buttons()
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self.withdraw()

    def _build_section_buttons(self):
        top = tk.Frame(self.main, bg="#0f0f0f")
        top.pack(fill="x", pady=(0, 8))
        self.btn_sections = {}

        def add_btn(name):
            btn = tk.Button(top, text=name, bg="#1a1a1a", fg="#ccc", relief="flat", padx=10, pady=6, cursor="hand2", command=lambda n=name: self.show_section(name=n))
            btn.pack(side="left", padx=4)
            self.btn_sections[name] = btn

        for _name in ["Whole App", "Modes", "Zone Editor Overlay", "Preview (Minimap)", "Overlay Controls", "Run & Logging"]:
            add_btn(_name)

    def register_tool_content(self, content: dict, order: list[str]):
        self._section_map = content
        self._sections_order = order[:]
        self._current_idx = 0
        if self._sections_order:
            self.show_section(self._sections_order[0], reset_ack=True)

    def show_section(self, name: str, reset_ack: bool = True):
        if reset_ack: self._acknowledged = False
        desc = self._section_map.get(name, "(No content for this section)")
        self.lbl_section.config(text=name)
        self.txt_description.delete("1.0", "end")
        self.txt_description.insert("1.0", desc)
        self._current_idx = self._sections_order.index(name) if name in self._sections_order else 0
        self.btn_ack.config(bg="#2979FF", fg="#fff", text="Acknowledge")
        
        master = self.master
        w, h = 600, 400
        x = master.winfo_x() + (master.winfo_width() // 2) - (w // 2)
        y = master.winfo_y() + (master.winfo_height() // 2) - (h // 2)
        self.geometry(f"{w}x{h}+{x}+{y}")
        self.deiconify()
        self.lift()

    def _on_ack(self):
        self._acknowledged = True
        self.btn_ack.config(text="Acknowledged ✓", bg="#111", fg="#8bc34a")

    def _on_next(self):
        if not self._sections_order: return
        nxt = min(self._current_idx + 1, len(self._sections_order) - 1)
        self._current_idx = nxt
        self.show_section(self._sections_order[nxt], reset_ack=True)

    def _on_close(self): self.withdraw()

# M5: ScreenSelectorPanel relocated to adapters/driving/ui_tkinter/components/; shim.
from adapters.driving.ui_tkinter.components.screen_selector import ScreenSelectorPanel

class IdentifyOverlay(tk.Toplevel):
    def __init__(self, master, monitor):
        super().__init__(master)
        self.geometry(f"{monitor.width}x{monitor.height}+{monitor.x}+{monitor.y}")
        self.overrideredirect(True); self.attributes("-topmost", True); self.attributes("-alpha", 0.8)
        self.configure(bg="#000000")
        tk.Label(self, text=str(monitor.display_num), font=("Consolas", 350, "bold"), bg="#000000", fg="#FFFFFF").pack(expand=True)
        self.after(500, self.destroy)

class Tooltip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tip = None
        widget.bind("<Enter>", self._show)
        widget.bind("<Leave>", self._hide)

    def _show(self, _=None):
        try:
            x = self.widget.winfo_rootx() + self.widget.winfo_width() // 2
            y = self.widget.winfo_rooty() + self.widget.winfo_height() + 4
            self.tip = tk.Toplevel(self.widget)
            self.tip.overrideredirect(True); self.tip.attributes("-topmost", True)
            tk.Label(self.tip, text=self.text, bg="#1a1a1a", fg="#ccc", font=("Consolas", 9), padx=8, pady=4).pack()
            self.tip.geometry(f"+{x}+{y}")
        except Exception: pass

    def _hide(self, _=None):
        if self.tip:
            try: self.tip.destroy()
            except Exception: pass
            self.tip = None


# ── Metadata Browser Dialog ───────────────────────────────────────────────────
class MetadataBrowserDialog(tk.Toplevel):
    """
    Shows all registered IO list profiles in the metadata DB.
    User can: load a profile into the current session, delete a profile,
    or see its point count / last imported date.
    """
    def __init__(self, master, on_load_profile):
        super().__init__(master)
        self.on_load_profile = on_load_profile
        self.result_profile_id = None
        self.title("📦 IO List Metadata Store")
        self.configure(bg="#0f0f0f")
        self.resizable(True, True)
        w, h = 780, 460
        x = master.winfo_x() + (master.winfo_width() - w) // 2
        y = master.winfo_y() + (master.winfo_height() - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")
        self.attributes("-topmost", True)
        self.grab_set()
        self._build()
        self._refresh_list()

    def _build(self):
        hdr = tk.Frame(self, bg="#0f0f0f")
        hdr.pack(fill="x", padx=16, pady=(14, 4))
        tk.Label(hdr, text="📦  REGISTERED IO LIST PROFILES", bg="#0f0f0f",
                 fg="#fff", font=("Consolas", 12, "bold")).pack(side="left")
        tk.Label(hdr, text="Import Excel → auto-saved here for reuse",
                 bg="#0f0f0f", fg="#555", font=("Consolas", 9)).pack(side="left", padx=12)

        cols_frame = tk.Frame(self, bg="#161616")
        cols_frame.pack(fill="x", padx=16)
        for txt, w in [("Profile Name", 34), ("Sheet", 20), ("Points", 6), ("Last Imported", 18)]:
            tk.Label(cols_frame, text=txt, bg="#161616", fg="#888",
                     font=("Consolas", 9, "bold"), width=w, anchor="w").pack(side="left", padx=4, pady=4)

        list_frame = tk.Frame(self, bg="#0a0a0a")
        list_frame.pack(fill="both", expand=True, padx=16, pady=4)
        sb = tk.Scrollbar(list_frame, orient="vertical")
        self.listbox = tk.Listbox(list_frame, bg="#0a0a0a", fg="#ccc",
                                  font=("Consolas", 10), relief="flat",
                                  selectbackground="#2979FF", selectforeground="#fff",
                                  activestyle="none", yscrollcommand=sb.set)
        sb.config(command=self.listbox.yview)
        sb.pack(side="right", fill="y")
        self.listbox.pack(fill="both", expand=True)
        self.listbox.bind("<<ListboxSelect>>", self._on_select)
        self.listbox.bind("<Double-1>", lambda e: self._do_load())

        self._profiles = []

        bot = tk.Frame(self, bg="#0f0f0f")
        bot.pack(fill="x", padx=16, pady=10)

        s = dict(font=("Consolas", 10, "bold"), relief="flat", padx=14, pady=6, cursor="hand2")
        self.btn_load = tk.Button(bot, text="✓ Load into Session", bg="#2979FF", fg="#fff",
                                  command=self._do_load, state="disabled", **s)
        self.btn_load.pack(side="left", padx=(0, 6))
        self.btn_del = tk.Button(bot, text="🗑 Delete Profile", bg="#8B0000", fg="#fff",
                                 command=self._do_delete, state="disabled", **s)
        self.btn_del.pack(side="left", padx=6)
        tk.Button(bot, text="✕ Close", bg="#222", fg="#aaa",
                  command=self._close, **s).pack(side="right")

        self.lbl_info = tk.Label(bot, text="", bg="#0f0f0f", fg="#aaa",
                                 font=("Consolas", 9))
        self.lbl_info.pack(side="left", padx=12)

    def _refresh_list(self):
        try:
            if not self.winfo_exists():
                return
        except Exception:
            return
        self.listbox.delete(0, "end")
        self._profiles = _metadata_list_profiles()
        if not self._profiles:
            self.listbox.insert("end", "  No profiles registered yet. Import an IO list first.")
            self.listbox.config(fg="#444")
            return
        self.listbox.config(fg="#ccc")
        for p in self._profiles:
            src_file = pathlib.Path(p["source_file"]).name
            line = f"  {p['name']:<42}  {p['point_count']:>4} pts   {p['imported_at'][:16]}"
            self.listbox.insert("end", line)

    def _on_select(self, _=None):
        sel = self.listbox.curselection()
        has = bool(sel) and bool(self._profiles)
        state = "normal" if has else "disabled"
        self.btn_load.config(state=state)
        self.btn_del.config(state=state)
        if has:
            p = self._profiles[sel[0]]
            src = pathlib.Path(p["source_file"]).name
            self.lbl_info.config(text=f"Source: {src}")

    def _do_load(self):
        sel = self.listbox.curselection()
        if not sel or not self._profiles: return
        p = self._profiles[sel[0]]
        self.on_load_profile(p["id"])
        self.destroy()

    def _do_delete(self):
        sel = self.listbox.curselection()
        if not sel or not self._profiles: return
        p = self._profiles[sel[0]]
        if messagebox.askyesno("Delete Profile",
                               f"Delete profile:\n{p['name']}\n\nThis removes all stored points. Cannot be undone.",
                               parent=self):
            _metadata_delete_profile(p["id"])
            self._refresh_list()
            self.btn_load.config(state="disabled")
            self.btn_del.config(state="disabled")
            self.lbl_info.config(text="Profile deleted.")

    def _close(self):
        """Close the dialog — routes through WM_DELETE_WINDOW so app can unregister listeners."""
        self.protocol("WM_DELETE_WINDOW", self.destroy)  # reset to safe default
        self.destroy()


# ── Coordinate Pick Overlay ───────────────────────────────────────────────────
class CoordinatePickOverlay(tk.Toplevel):
    """
    A full-screen transparent overlay that captures exactly ONE left-click,
    returns the coordinates, and automatically dismisses itself.
    """
    def __init__(self, master, monitor, on_picked, on_cancel=None):
        super().__init__(master)
        self.monitor = monitor
        self.on_picked = on_picked
        self.on_cancel = on_cancel   # called on Esc so callers can restore their window

        # Cover the entire selected monitor
        self.geometry(f"{monitor.width}x{monitor.height}+{monitor.x}+{monitor.y}")
        self.overrideredirect(True)
        self.attributes("-topmost", True)
        self.attributes("-alpha", 0.15)  # Subtle visible tint to show pick-mode is active
        self.configure(bg="#AA00FF")

        self.canvas = tk.Canvas(self, bg="#AA00FF", highlightthickness=0, cursor="target")
        self.canvas.pack(fill="both", expand=True)

        # Draw picking instructions centered on screen
        self.canvas.create_text(
            monitor.width // 2, 50,
            text="🎯 CLICK ANYWHERE TO CAPTURE COORDINATE  │  Esc to cancel",
            fill="#ffffff", font=("Consolas", 14, "bold")
        )

        self.canvas.bind("<Button-1>", self._on_click)
        self.bind("<Escape>", lambda e: self._on_cancel())
        self.focus_force()

    def _on_click(self, event):
        # Calculate absolute screen coordinates based on display offsets
        abs_x = event.x_root
        abs_y = event.y_root
        self.on_picked(abs_x, abs_y)
        self.destroy()

    def _on_cancel(self):
        cb = self.on_cancel
        self.destroy()
        if callable(cb):
            try:
                cb()
            except Exception:
                pass


# ── Suite Card Config Dialog ───────────────────────────────────────────────────
class SuiteCardConfigDialog(tk.Toplevel):
    """
    Per-card configuration for ISCS suite cards.
    Lets user set: profile (IO list), protocol, navigation coords, page list.
    Navigation coords are captured by clicking a 'Pick' button then clicking on screen.
    """
    def __init__(self, master, card_cfg=None):
        super().__init__(master)
        self.result = None
        self._pick_mode = None
        self._pick_win  = None

        existing = card_cfg or {}
        # Auto-load template silently for new cards (no existing config)
        if not card_cfg:
            tmpl = _load_template()
            nav_tmpl = tmpl.get("navigation", {})
            if nav_tmpl:
                existing.setdefault("navigation", {})
                for k, v in nav_tmpl.items():
                    existing["navigation"].setdefault(k, v)

        self.title("⚙ Suite Card Configuration")
        self.configure(bg="#0f0f0f")
        self.resizable(True, True)
        w, h = 680, 680
        x = master.winfo_x() + (master.winfo_width() - w) // 2
        y = master.winfo_y() + (master.winfo_height() - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")
        self.attributes("-topmost", True)
        self.grab_set()
        self._build(existing)

    def _build(self, cfg):
        # ── Header
        tk.Label(self, text="⚙  SUITE CARD CONFIGURATION", bg="#0f0f0f",
                 fg="#fff", font=("Consolas", 12, "bold")).pack(anchor="w", padx=16, pady=(14, 4))

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=12, pady=6)

        # ── Tab 1: IO List / Profile
        tab_io = tk.Frame(nb, bg="#0f0f0f")
        nb.add(tab_io, text=" IO List ")
        self._build_io_tab(tab_io, cfg)

        # ── Tab 2: Protocol
        tab_proto = tk.Frame(nb, bg="#0f0f0f")
        nb.add(tab_proto, text=" Protocol ")
        self._build_proto_tab(tab_proto, cfg)

        # ── Tab 3: Navigation
        tab_nav = tk.Frame(nb, bg="#0f0f0f")
        nb.add(tab_nav, text=" Navigation ")
        self._build_nav_tab(tab_nav, cfg)

        # ── Tab 4: Zones
        tab_zones = tk.Frame(nb, bg="#0f0f0f")
        nb.add(tab_zones, text=" Zones ")
        self._build_zones_tab(tab_zones)

        # ── Bottom buttons
        bot = tk.Frame(self, bg="#0f0f0f")
        bot.pack(fill="x", padx=16, pady=10)
        s = dict(font=("Consolas", 10, "bold"), relief="flat", padx=14, pady=6, cursor="hand2")
        tk.Button(bot, text="✓ Save Card", bg="#2979FF", fg="#fff",
                  command=self._save, **s).pack(side="left", padx=(0, 6))
        tk.Button(bot, text="Cancel", bg="#222", fg="#aaa",
                  command=self.destroy, **s).pack(side="left")

    # ── IO List tab
    def _build_io_tab(self, parent, cfg):
        ls = dict(bg="#0f0f0f", fg="#aaa", font=("Consolas", 10))
        es = dict(bg="#1a1a1a", fg="#fff", insertbackground="#fff",
                  font=("Consolas", 10), relief="flat", bd=6)

        tk.Label(parent, text="Card Name:", **ls).grid(row=0, column=0, sticky="w", padx=14, pady=(14, 4))
        self.v_card_name = tk.StringVar(value=cfg.get("card_name", "New Card"))
        tk.Entry(parent, textvariable=self.v_card_name, width=36, **es).grid(row=0, column=1, columnspan=2, sticky="ew", padx=8, pady=(14, 4))

        tk.Label(parent, text="Loaded Profile:", **ls).grid(row=1, column=0, sticky="w", padx=14, pady=4)
        self.lbl_profile = tk.Label(parent, text=cfg.get("profile_name", "None"), bg="#161616",
                                    fg="#2979FF", font=("Consolas", 10), padx=8, pady=4)
        self.lbl_profile.grid(row=1, column=1, sticky="ew", padx=8, pady=4)
        self._profile_id     = cfg.get("profile_id", None)
        self._profile_points = cfg.get("profile_points", [])

        # ── Auto-fill from app's currently loaded IO list if this is a new card ──
        if self._profile_id is None:
            app_points = list(getattr(self.master, "iscs_excel_points", []))
            if app_points:
                self._profile_points = app_points
                try:
                    conn = _metadata_get_db()
                    cur  = conn.execute("SELECT id, name FROM profiles ORDER BY imported_at DESC LIMIT 1")
                    row  = cur.fetchone()
                    conn.close()
                    if row:
                        self._profile_id = row["id"]
                        self.lbl_profile.config(text=row["name"], fg="#00C853")
                    else:
                        self.lbl_profile.config(text="Current Session", fg="#FFD600")
                except Exception:
                    self.lbl_profile.config(text="Current Session", fg="#FFD600")

        btn_s = dict(font=("Consolas", 9, "bold"), relief="flat", padx=10, pady=4, cursor="hand2")
        tk.Button(parent, text="📦 Browse Profiles", bg="#333", fg="#ccc",
                  command=self._browse_profiles, **btn_s).grid(row=1, column=2, padx=4, pady=4)

        tk.Label(parent, text="Point Count:", **ls).grid(row=2, column=0, sticky="w", padx=14, pady=4)
        self.lbl_point_count = tk.Label(parent,
            text=str(len(self._profile_points)) if self._profile_points else "—",
            bg="#0f0f0f", fg="#00C853", font=("Consolas", 11, "bold"))
        self.lbl_point_count.grid(row=2, column=1, sticky="w", padx=8)

        sep = tk.Frame(parent, bg="#222", height=1)
        sep.grid(row=3, column=0, columnspan=3, sticky="ew", padx=14, pady=10)

        tk.Label(parent, text="Run Filter", bg="#0f0f0f", fg="#888",
                 font=("Consolas", 9, "bold")).grid(row=4, column=0, sticky="w", padx=14)

        tk.Label(parent, text="Severity filter (comma-sep, blank=all):", **ls).grid(
            row=5, column=0, sticky="w", padx=14, pady=4)
        self.v_sev_filter = tk.StringVar(value=cfg.get("severity_filter", ""))
        tk.Entry(parent, textvariable=self.v_sev_filter, width=20, **es).grid(
            row=5, column=1, sticky="w", padx=8, pady=4)

        tk.Label(parent, text="Point range (start–end, blank=all):", **ls).grid(
            row=6, column=0, sticky="w", padx=14, pady=4)
        range_f = tk.Frame(parent, bg="#0f0f0f")
        range_f.grid(row=6, column=1, sticky="w", padx=8)
        rng = cfg.get("point_range", [0, 9999])
        self.v_range_start = tk.StringVar(value=str(rng[0]))
        self.v_range_end   = tk.StringVar(value=str(rng[1]))
        tk.Entry(range_f, textvariable=self.v_range_start, width=8, **es).pack(side="left")
        tk.Label(range_f, text=" – ", bg="#0f0f0f", fg="#888", font=("Consolas", 10)).pack(side="left")
        tk.Entry(range_f, textvariable=self.v_range_end,   width=8, **es).pack(side="left")

        parent.columnconfigure(1, weight=1)

    def _browse_profiles(self):
        def on_load(profile_id):
            prof, points = _metadata_load_profile(profile_id)
            if prof:
                self._profile_id     = profile_id
                self._profile_points = points
                self.lbl_profile.config(text=prof["name"])
                self.lbl_point_count.config(text=str(len(points)))
        MetadataBrowserDialog(self, on_load_profile=on_load)

    # ── Protocol tab
    def _build_proto_tab(self, parent, cfg):
        proto_cfg = cfg.get("protocol", {})
        ls = dict(bg="#0f0f0f", fg="#aaa", font=("Consolas", 10))
        es = dict(bg="#1a1a1a", fg="#fff", insertbackground="#fff",
                  font=("Consolas", 10), relief="flat", bd=6)

        tk.Label(parent, text="Protocol type:", **ls).grid(row=0, column=0, sticky="w", padx=14, pady=(14, 4))
        self.v_proto_type = tk.StringVar(value=proto_cfg.get("type", "MODBUS"))
        for i, p in enumerate(["MODBUS", "SNMP"]):
            tk.Radiobutton(parent, text=p, variable=self.v_proto_type, value=p,
                           bg="#0f0f0f", fg="#ccc", selectcolor="#333",
                           activebackground="#0f0f0f",
                           font=("Consolas", 10)).grid(row=0, column=i+1, padx=6, pady=(14, 4), sticky="w")

        rows = [
            ("Host / IP:", "v_host",    proto_cfg.get("host", "127.0.0.1"),    20),
            ("Port:",      "v_port",    str(proto_cfg.get("port", 502)),        8),
            ("Unit ID:",   "v_unit_id", str(proto_cfg.get("unit_id", 1)),       8),
        ]
        for i, (lbl, attr, default, width) in enumerate(rows, start=1):
            tk.Label(parent, text=lbl, **ls).grid(row=i, column=0, sticky="w", padx=14, pady=6)
            v = tk.StringVar(value=default)
            setattr(self, attr, v)
            tk.Entry(parent, textvariable=v, width=width, **es).grid(
                row=i, column=1, columnspan=2, sticky="w", padx=8, pady=6)

        tk.Label(parent, text="Trigger value (alarm):", **ls).grid(
            row=4, column=0, sticky="w", padx=14, pady=6)
        self.v_trigger_val = tk.StringVar(value=str(proto_cfg.get("trigger_value", 1)))
        tk.Entry(parent, textvariable=self.v_trigger_val, width=8, **es).grid(
            row=4, column=1, sticky="w", padx=8, pady=6)
        tk.Label(parent, text="Reset value (normal):", **ls).grid(
            row=5, column=0, sticky="w", padx=14, pady=6)
        self.v_reset_val = tk.StringVar(value=str(proto_cfg.get("reset_value", 0)))
        tk.Entry(parent, textvariable=self.v_reset_val, width=8, **es).grid(
            row=5, column=1, sticky="w", padx=8, pady=6)

        tk.Label(parent,
                 text="Trigger / reset values override IO list states when set.",
                 bg="#0f0f0f", fg="#555", font=("Consolas", 8)).grid(
            row=6, column=0, columnspan=3, sticky="w", padx=14, pady=2)

        parent.columnconfigure(1, weight=1)

    # ── Navigation tab
    def _build_nav_tab(self, parent, cfg):
        nav = cfg.get("navigation", {})
        ls  = dict(bg="#0f0f0f", fg="#aaa", font=("Consolas", 10))

        tk.Label(parent,
                 text="Click 'Pick' then click the target on screen. Coords captured automatically.",
                 bg="#0f0f0f", fg="#555", font=("Consolas", 9)).pack(anchor="w", padx=14, pady=(10, 6))

        self._nav_vars = {}
        nav_items = [
            ("subsystem_tab",       "Subsystem Tab button",           "Click after selecting AMS/FAS/etc in SysView"),
            ("alarm_list_btn",      "Alarm List page button",         "Global top-bar icon or nav button for alarm list"),
            ("event_list_btn",      "Event List page button",         "Global top-bar icon or nav button for event list"),
            ("home_btn",            "Home / reset button",            "Global home icon"),
            ("rightclick_row1",     "Right-click coord (Row 1)",      "Fixed coord to right-click on first alarm row in alarm panel"),
            ("rightclick_page_btn", "Context menu 'Page' button",     "Fixed coord of 'Page' option in right-click context menu"),
        ]

        coord_frame = tk.Frame(parent, bg="#0f0f0f")
        coord_frame.pack(fill="both", expand=True, padx=14)

        for row_i, (key, label, hint) in enumerate(nav_items):
            existing_coord = nav.get(key, {})
            x_val = existing_coord.get("x", "")
            y_val = existing_coord.get("y", "")

            tk.Label(coord_frame, text=label, **ls).grid(
                row=row_i*2, column=0, sticky="w", pady=(10, 0))
            tk.Label(coord_frame, text=hint, bg="#0f0f0f", fg="#555",
                     font=("Consolas", 8)).grid(
                row=row_i*2+1, column=0, sticky="w", pady=(0, 4))

            vx = tk.StringVar(value=str(x_val))
            vy = tk.StringVar(value=str(y_val))
            self._nav_vars[key] = (vx, vy)

            coord_inner = tk.Frame(coord_frame, bg="#0f0f0f")
            coord_inner.grid(row=row_i*2, column=1, padx=8, pady=(10, 0), sticky="w")
            tk.Label(coord_inner, text="X:", bg="#0f0f0f", fg="#888", font=("Consolas", 9)).pack(side="left")
            tk.Entry(coord_inner, textvariable=vx, width=6, bg="#1a1a1a", fg="#fff",
                     insertbackground="#fff", font=("Consolas", 10), relief="flat", bd=4).pack(side="left", padx=(2, 8))
            tk.Label(coord_inner, text="Y:", bg="#0f0f0f", fg="#888", font=("Consolas", 9)).pack(side="left")
            tk.Entry(coord_inner, textvariable=vy, width=6, bg="#1a1a1a", fg="#fff",
                     insertbackground="#fff", font=("Consolas", 10), relief="flat", bd=4).pack(side="left", padx=(2, 8))

            btn = tk.Button(coord_inner, text="🎯 Pick",
                            bg="#333", fg="#ccc", font=("Consolas", 9, "bold"),
                            relief="flat", padx=8, pady=2, cursor="hand2",
                            command=lambda k=key: self._pick_coord(k))
            btn.pack(side="left")

        # Pages list (left nav buttons for this subsystem)
        sep = tk.Frame(parent, bg="#222", height=1)
        sep.pack(fill="x", padx=14, pady=8)

        tk.Label(parent, text="Subsystem Page List  (left sidebar nav buttons):",
                 bg="#0f0f0f", fg="#aaa", font=("Consolas", 10)).pack(anchor="w", padx=14)
        tk.Label(parent, text="Add each left-nav page button (name + coords). Engine navigates to correct page per point.",
                 bg="#0f0f0f", fg="#555", font=("Consolas", 8)).pack(anchor="w", padx=14, pady=(0, 4))

        pages_outer = tk.Frame(parent, bg="#0f0f0f")
        pages_outer.pack(fill="both", expand=True, padx=14, pady=4)

        self._pages_frame = tk.Frame(pages_outer, bg="#0f0f0f")
        self._pages_frame.pack(fill="both", expand=True)
        self._page_rows = []

        for pg in nav.get("pages", []):
            self._add_page_row(pg.get("name", ""), pg.get("x", ""), pg.get("y", ""))

        tk.Button(pages_outer, text="+ Add Page", bg="#222", fg="#ccc",
                  font=("Consolas", 9, "bold"), relief="flat", padx=10, pady=4,
                  cursor="hand2", command=lambda: self._add_page_row()).pack(anchor="w", pady=4)

    def _build_zones_tab(self, parent):
        """Zones tab — shows template zone status, Load/Load All buttons."""
        ls  = dict(bg="#0f0f0f", fg="#aaa", font=("Consolas", 10))
        dim = dict(bg="#0f0f0f", fg="#555", font=("Consolas", 8))
        bs  = dict(font=("Consolas", 9, "bold"), relief="flat", padx=10, pady=4, cursor="hand2")

        tk.Label(parent, text="Zone Template", bg="#0f0f0f",
                 fg="#fff", font=("Consolas", 11, "bold")).pack(anchor="w", padx=14, pady=(14, 2))
        tk.Label(parent,
                 text="Zones are shared across all cards. When you draw and save a zone\n"
                      "in the overlay, it's stored in iscs_template.json automatically.\n"
                      "Use Load to pull saved zones into this card.",
                 **dim).pack(anchor="w", padx=14, pady=(0, 10))

        zone_types = [
            ("alarm_panel",    "🚨 Alarm Panel",    ALARM_PANEL_COLOR, "#fff"),
            ("equipment_page", "🖥  Equipment Page", EQUIP_ZONE_COLOR,  "#000"),
            ("alarm_list",     "📋 Alarm List",     ALARM_LIST_COLOR,  "#000"),
            ("event_list",     "📅 Event List",     EVENT_LIST_COLOR,  "#000"),
        ]

        self._zone_status_labels = {}
        tmpl = _load_template()

        for zt, label, color, fg in zone_types:
            row = tk.Frame(parent, bg="#161616", pady=6)
            row.pack(fill="x", padx=14, pady=3)

            # Color indicator
            tk.Label(row, text="  ", bg=color, width=2).pack(side="left", padx=(8, 6))
            tk.Label(row, text=label, bg="#161616", fg="#ccc",
                     font=("Consolas", 10, "bold"), width=20, anchor="w").pack(side="left")

            # Status label
            has_zone = zt in tmpl.get("zones", {})
            status_text = "✓ in template" if has_zone else "— not saved yet"
            status_color = "#00C853" if has_zone else "#555"
            lbl = tk.Label(row, text=status_text, bg="#161616",
                           fg=status_color, font=("Consolas", 9))
            lbl.pack(side="left", padx=8)
            self._zone_status_labels[zt] = lbl

            # Load button
            tk.Button(row, text="⬇ Load", bg="#1a2e1a", fg="#00C853",
                      command=lambda t=zt: self._load_zone_from_template(t),
                      state="normal" if has_zone else "disabled",
                      **bs).pack(side="right", padx=8)

        # Load All button
        sep = tk.Frame(parent, bg="#222", height=1)
        sep.pack(fill="x", padx=14, pady=10)

        btn_row = tk.Frame(parent, bg="#0f0f0f")
        btn_row.pack(fill="x", padx=14)

        tk.Button(btn_row, text="⬇ Load All Zones from Template",
                  bg="#1a2e1a", fg="#00C853",
                  font=("Consolas", 10, "bold"), relief="flat",
                  padx=14, pady=6, cursor="hand2",
                  command=self._load_all_zones_from_template).pack(side="left")

        tk.Button(btn_row, text="🔄 Refresh", bg="#222", fg="#888",
                  font=("Consolas", 9), relief="flat",
                  padx=10, pady=6, cursor="hand2",
                  command=lambda: self._refresh_zone_status()).pack(side="left", padx=8)

        tk.Label(parent,
                 text="Note: Zones are applied when the suite card runs.\n"
                      "To draw/update zones, use the Draw Zones button on the main screen.",
                 **dim).pack(anchor="w", padx=14, pady=(10, 0))

    def _load_zone_from_template(self, zone_type: str):
        """Load a single zone from template into this card's zones_per_page."""
        tmpl = _load_template()
        zd = tmpl.get("zones", {}).get(zone_type)
        if not zd:
            return
        # Store in result zones_per_page under "Global"
        if not hasattr(self, '_template_zones'):
            self._template_zones = {}
        self._template_zones[zone_type] = Zone.from_dict(zd)
        # Update status label
        lbl = self._zone_status_labels.get(zone_type)
        if lbl:
            lbl.config(text="✓ loaded", fg="#5599ff")

    def _load_all_zones_from_template(self):
        """Load all zones from template."""
        tmpl = _load_template()
        zones = tmpl.get("zones", {})
        if not zones:
            return
        if not hasattr(self, '_template_zones'):
            self._template_zones = {}
        for zt, zd in zones.items():
            self._template_zones[zt] = Zone.from_dict(zd)
            lbl = self._zone_status_labels.get(zt)
            if lbl:
                lbl.config(text="✓ loaded", fg="#5599ff")

    def _refresh_zone_status(self):
        """Refresh the zone status labels from template."""
        tmpl = _load_template()
        for zt, lbl in self._zone_status_labels.items():
            has_zone = zt in tmpl.get("zones", {})
            lbl.config(text="✓ in template" if has_zone else "— not saved yet",
                       fg="#00C853" if has_zone else "#555")

    def _add_page_row(self, name="", x="", y=""):
        row_frame = tk.Frame(self._pages_frame, bg="#161616", pady=3)
        row_frame.pack(fill="x", pady=2)
        vname = tk.StringVar(value=name)
        vx    = tk.StringVar(value=str(x))
        vy    = tk.StringVar(value=str(y))
        tk.Entry(row_frame, textvariable=vname, width=22, bg="#1a1a1a", fg="#fff",
                 insertbackground="#fff", font=("Consolas", 9), relief="flat", bd=4).pack(side="left", padx=4)
        tk.Label(row_frame, text="X:", bg="#161616", fg="#888", font=("Consolas", 9)).pack(side="left")
        tk.Entry(row_frame, textvariable=vx, width=5, bg="#1a1a1a", fg="#fff",
                 insertbackground="#fff", font=("Consolas", 9), relief="flat", bd=4).pack(side="left", padx=2)
        tk.Label(row_frame, text="Y:", bg="#161616", fg="#888", font=("Consolas", 9)).pack(side="left")
        tk.Entry(row_frame, textvariable=vy, width=5, bg="#1a1a1a", fg="#fff",
                 insertbackground="#fff", font=("Consolas", 9), relief="flat", bd=4).pack(side="left", padx=2)
        tk.Button(row_frame, text="🎯", bg="#222", fg="#ccc", relief="flat", padx=6, pady=2,
                  cursor="hand2",
                  command=lambda idx=len(self._page_rows): self._pick_page_coord(idx)).pack(side="left", padx=4)
        tk.Button(row_frame, text="✕", bg="#8B0000", fg="#fff", relief="flat", padx=6, pady=2,
                  cursor="hand2",
                  command=lambda f=row_frame, r=(vname, vx, vy, row_frame): self._del_page_row(r)).pack(side="left", padx=2)
        self._page_rows.append((vname, vx, vy, row_frame))

    def _del_page_row(self, row_tuple):
        vname, vx, vy, frame = row_tuple
        if row_tuple in self._page_rows:
            self._page_rows.remove(row_tuple)
        frame.destroy()

    def _pick_coord(self, key):
        self.withdraw()
        self._pick_mode = ("nav", key)
        self._show_pick_prompt()

    def _pick_page_coord(self, idx):
        self.withdraw()
        self._pick_mode = ("page", idx)
        self._show_pick_prompt()

    def _show_pick_prompt(self):
        try:
            import __main__
            mon = getattr(__main__.app, 'active_mon', self.master.active_mon)
        except Exception:
            mon = self.master.active_mon

        def on_picked(x, y):
            self.after(0, lambda: self._apply_picked_coord((x, y)))

        def on_cancel():
            try:
                self.deiconify(); self.lift()
            except Exception:
                pass

        CoordinatePickOverlay(self, mon, on_picked, on_cancel)

    def _apply_picked_coord(self, xy):
        x, y = int(xy[0]), int(xy[1])
        mode = self._pick_mode
        if mode:
            kind, key_or_idx = mode
            if kind == "nav" and key_or_idx in self._nav_vars:
                self._nav_vars[key_or_idx][0].set(str(x))
                self._nav_vars[key_or_idx][1].set(str(y))
                # Auto-save to template
                _save_template({"navigation": {key_or_idx: {"x": x, "y": y}}})
            elif kind == "page" and isinstance(key_or_idx, int) and key_or_idx < len(self._page_rows):
                self._page_rows[key_or_idx][1].set(str(x))
                self._page_rows[key_or_idx][2].set(str(y))
                # Auto-save pages to template
                self._save_pages_to_template()
        self._pick_mode = None
        if self._pick_win:
            try: self._pick_win.destroy()
            except Exception: pass
            self._pick_win = None
        self.deiconify()

    def _save_pages_to_template(self):
        """Save current page rows to template."""
        pages = []
        for vname, vx, vy, _ in self._page_rows:
            n = vname.get().strip()
            if not n: continue
            try: pages.append({"name": n, "x": int(vx.get()), "y": int(vy.get())})
            except ValueError: pages.append({"name": n, "x": 0, "y": 0})
        _save_template({"navigation": {"pages": pages}})

    def _cancel_pick(self):
        self._pick_mode = None
        if self._pick_win:
            try: self._pick_win.destroy()
            except Exception: pass
        self.deiconify()

    def _save(self):
        nav = {}
        for key, (vx, vy) in self._nav_vars.items():
            try: nav[key] = {"x": int(vx.get()), "y": int(vy.get())}
            except ValueError: nav[key] = {}
        pages = []
        for vname, vx, vy, _ in self._page_rows:
            n = vname.get().strip()
            if not n: continue
            try: pages.append({"name": n, "x": int(vx.get()), "y": int(vy.get())})
            except ValueError: pages.append({"name": n, "x": 0, "y": 0})
        nav["pages"] = pages

        try: trigger_val = int(self.v_trigger_val.get())
        except: trigger_val = 1
        try: reset_val   = int(self.v_reset_val.get())
        except: reset_val = 0
        try: port = int(self.v_port.get())
        except: port = 502
        try: unit_id = int(self.v_unit_id.get())
        except: unit_id = 1

        sev_raw = [s.strip() for s in self.v_sev_filter.get().split(",") if s.strip()]
        sev_filter = []
        for s in sev_raw:
            try: sev_filter.append(int(s))
            except: pass

        try: rng_start = int(self.v_range_start.get())
        except: rng_start = 0
        try: rng_end = int(self.v_range_end.get())
        except: rng_end = 9999

        self.result = {
            "card_name":       self.v_card_name.get().strip() or "Unnamed Card",
            "profile_id":      self._profile_id,
            "profile_name":    self.lbl_profile.cget("text"),
            "profile_points":  self._profile_points,
            "protocol": {
                "type":          self.v_proto_type.get(),
                "host":          self.v_host.get().strip(),
                "port":          port,
                "unit_id":       unit_id,
                "trigger_value": trigger_val,
                "reset_value":   reset_val,
            },
            "navigation":      nav,
            "severity_filter": sev_filter,
            "point_range":     [rng_start, rng_end],
            "template_zones":  getattr(self, '_template_zones', {}),
        }
        self.destroy()

    @classmethod
    def ask(cls, master, card_cfg=None):
        dlg = cls(master, card_cfg)
        master.wait_window(dlg)
        return dlg.result


# ── Toast Notification ────────────────────────────────────────────────────────
class Toast(tk.Toplevel):
    """
    Small non-blocking notification that slides in at bottom-right of parent
    and auto-dismisses after `duration_ms` milliseconds.
    kind: 'success' | 'error' | 'info'
    """
    _COLORS = {
        "success": {"bg": "#1a3a1a", "border": "#00C853", "icon": "✓", "fg": "#00C853"},
        "error":   {"bg": "#3a1a1a", "border": "#FF1744", "icon": "✕", "fg": "#FF1744"},
        "info":    {"bg": "#1a1a2a", "border": "#2979FF", "icon": "ℹ", "fg": "#2979FF"},
    }

    def __init__(self, master, message, kind="success", duration_ms=3500):
        super().__init__(master)
        self.overrideredirect(True)
        self.attributes("-topmost", True)
        self.attributes("-alpha", 0.0)

        c = self._COLORS.get(kind, self._COLORS["info"])
        self.configure(bg=c["border"])  # 1px border via bg

        inner = tk.Frame(self, bg=c["bg"], padx=14, pady=10)
        inner.pack(padx=1, pady=1)

        tk.Label(inner, text=c["icon"], bg=c["bg"], fg=c["border"],
                 font=("Consolas", 14, "bold")).pack(side="left", padx=(0, 10))
        tk.Label(inner, text=message, bg=c["bg"], fg="#ffffff",
                 font=("Consolas", 10), wraplength=320, justify="left").pack(side="left")

        self.update_idletasks()
        pw = master.winfo_rootx() + master.winfo_width()
        ph = master.winfo_rooty() + master.winfo_height()
        tw = self.winfo_reqwidth()
        th = self.winfo_reqheight()
        x = pw - tw - 18
        y = ph - th - 18
        self.geometry(f"+{x}+{y}")

        self._fade_in(duration_ms)

    def _fade_in(self, duration_ms, step=0):
        alpha = step / 10
        try:
            self.attributes("-alpha", alpha)
        except Exception:
            return
        if step < 10:
            self.after(20, self._fade_in, duration_ms, step + 1)
        else:
            self.after(duration_ms, self._fade_out)

    def _fade_out(self, step=10):
        alpha = step / 10
        try:
            self.attributes("-alpha", alpha)
        except Exception:
            return
        if step > 0:
            self.after(30, self._fade_out, step - 1)
        else:
            try:
                self.destroy()
            except Exception:
                pass

    @classmethod
    def show(cls, master, message, kind="success", duration_ms=3500):
        cls(master, message, kind=kind, duration_ms=duration_ms)


# ── Main Application ──────────────────────────────────────────────────────────
# M5: Tk driving-adapter dispatcher + views, extracted from this file one at a time.
from adapters.driving.ui_tkinter.composition import build_tk_core_api
from adapters.driving.ui_tkinter.dispatcher import TkEventDispatcher
from adapters.driving.ui_tkinter.views.log_sink import LogSink
from adapters.driving.ui_tkinter.views.run_progress_view import RunProgressView
from adapters.driving.ui_tkinter.views.stats_view import StatsView
from adapters.driving.ui_tkinter.views.settings_view import SettingsView
from adapters.driving.ui_tkinter.views.run_controls import RunControls
from core.services.workspace import WorkspaceSession


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        # M5: the Tk app's R-HEX-2 dispatcher — views marshal onto the loop via this.
        self._dispatcher = TkEventDispatcher(self)
        # M3.5: the live zone working set lives in a core WorkspaceSession.
        self._workspace = WorkspaceSession()

        try:
            myappid = 'willowglen.WilloWisp.v1'  # Arbitrary unique string
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
        except Exception:
            pass
        self.title("WilloWisp  — Unified Test Automation & System Validation Suite")
        self.geometry("1100x850")
        self.minsize(860, 600)
        self.configure(bg="#0f0f0f")

        # ── Set Native Window Icon (Sharp, Square Center Crop) ────────────────
        logo_path = BASE_DIR / "wispTest.png"
        self.logo_img = None
        if PIL_AVAILABLE and logo_path.exists():
            try:
                img = Image.open(logo_path)
                w, h = img.size
                
                if w > h:
                    # Extract the center square (the "W" wrench icon) to avoid horizontal squishing
                    left = (w - h) // 2
                    right = left + h
                    icon_img = img.crop((left, 0, right, h))
                else:
                    icon_img = img
                
                # Resize cleanly to a standard sharp icon size
                icon_img = icon_img.resize((32, 32), Image.LANCZOS)
                self.logo_img = ImageTk.PhotoImage(icon_img)
                self.iconphoto(True, self.logo_img)
            except Exception as e:
                logger.warning(f"Failed to load window icon: {e}")

        self.monitors = detect_monitors()
        self.active_mon = self.monitors[0]
        self.valid_points, self.all_points = [], []
        # zones / zones_per_page now live in self._workspace (M3.5), via properties.
        self.run_mode = tk.StringVar(value="sequence")  
        self.grid_spacing = GRID_SPACING
        self.suite_panel = None
        self._suite_pane_visible = False
        
        self.click_engine, self.hud, self.crosshair_overlay = None, None, None
        self._last_run_hk = 0

        self.protocols = ProtocolManager(APP_CONFIG)
        self.iscs_excel_points = []
        self._profile_update_listeners = []   # callbacks notified when a new IO list is registered

        # M5: the Tk app runs on the Core API facade, wired to its OWN registry +
        # live config provider (no duplicate). Views forward intents through this.
        self.core_api = build_tk_core_api(
            self, registry=CORE_REGISTRY, config_provider=_config_provider,
            protocols=self.protocols, monitors=self.monitors,
        )

        def _tk_error_handler(exc, val, tb):
            logger.error("Tkinter Callback Exception!")
            logger.error("".join(traceback.format_exception(exc, val, tb)))
        self.report_callback_exception = _tk_error_handler

        self._build_ui()
        self.help_panel = self._init_help_panel()
        self._register_hotkeys()
        self.after(0, lambda: self.screen_selector._select(0))
        self._log("WilloWisp v1 ready. Modes: Sequence | Grid | Suite Runner.")

    def _set_taskbar_icon(self):
        """Forces the custom window to register with the Windows taskbar."""
        try:
            hwnd = ctypes.windll.user32.GetParent(self.winfo_id()) or self.winfo_id()
            style = ctypes.windll.user32.GetWindowLongW(hwnd, -20)
            style = style & ~0x00000080  # Turn off WS_EX_TOOLWINDOW
            style = style | 0x00040000   # Turn on WS_EX_APPWINDOW
            ctypes.windll.user32.SetWindowLongW(hwnd, -20, style)
            self.withdraw()
            self.after(10, self.deiconify)
        except Exception:
            pass    
    def _build_help_content(self):
        content = {
        "Whole App": (
            "WilloWisp - UI Testing & ISCS Automation Framework\n\n"
            "What it does:\n"
            "- Lets you pick a mode (Fuzzer, RPA, Suite Runner)\n"
            "- Select a monitor\n"
            "- Draw zones on the overlay\n"
            "- Preview click points\n"
            "- Run with pause/stop + screenshot/JSON logging\n\n"
            "Outputs go into test_logs/."
        ),
        "Modes": (
            "Fuzzer / Grid Scan:\n"
            "- Draw INCLUDE zones (and optional EXCLUDE zones)\n"
            "- The tool clicks a grid lattice inside INCLUDE.\n\n"
            "RPA / Targeted Sequence:\n"
            "- Draw TARGET boxes in order\n"
            "- The tool clicks the center of each TARGET sequentially.\n\n"
            "Suite Runner:\n"
            "- Draw an ALARM PANEL zone over the SCADA alarm banner.\n"
            "- Load an IO List (Excel).\n"
            "- Triggers Modbus alarms, waits, and verifies via OCR/Color matching."
        ),
        "Zone Editor Overlay": (
            "This overlay is positioned on the selected monitor.\n\n"
            "Draw / edit zones:\n"
            "- Drag to create a rectangle\n"
            "- Drag edges/corners to resize\n"
            "- Drag inside the rectangle to move\n\n"
            "Type depends on your selected mode."
        ),
        "Preview (Minimap)": (
            "The preview canvas shows a visual confirmation of your zones/click points.\n\n"
            "Use it to sanity-check before running."
        ),
        "Overlay Controls": (
            "Overlay buttons:\n"
            "- ↩ Undo: revert the last edit step\n"
            "- 🗑 Delete: delete last touched zone\n"
            "- ✕ Clear: remove all zones\n"
            "- ✓ Done: close overlay and apply zones\n"
        ),
        "Run & Logging": (
            "Run controls:\n"
            "- Run (Ctrl+5)\n"
            "- Pause (Space)\n"
            "- Stop (Esc)\n\n"
            "Logging:\n"
            "- screenshots per click / per ISCS test\n"
            "- periodic heartbeat screenshots\n"
            "- results.json or Test_Execution_Summary.csv"
            ),
        }
        return content, list(content.keys())
            
    def _init_help_panel(self):
        panel = HelpInspectorPanel(self, tool_title="WilloWisp + ISCS Framework")
        content, order = self._build_help_content()
        panel.register_tool_content(content, order=order)
        panel.withdraw()
        return panel

    def _register_hotkeys(self):
        if not KEYBOARD_AVAILABLE: return
        try:
            keyboard.add_hotkey("ctrl+5", self._hk_run, suppress=False) 
            keyboard.add_hotkey("ctrl+f12", self._hk_stop, suppress=True)
            keyboard.add_hotkey("escape", self._hk_stop, suppress=False)
            keyboard.add_hotkey("space", self._hk_space, suppress=False)
        except Exception: pass

    def _unregister_hotkeys(self):
        if not KEYBOARD_AVAILABLE: return
        try:
            for hk in ("ctrl+5", "ctrl+f12", "escape", "space"): keyboard.remove_hotkey(hk)
        except Exception: pass

    def _hk_run(self):
        now = time.time()
        if now - self._last_run_hk < 0.5: return
        self._last_run_hk = now
        self.after(0, self._run_test)
    def _hk_stop(self): self.after(0, self._stop_test)
    def _hk_space(self): self.after(0, self._toggle_pause)

    # M3.5: zones / zones_per_page delegate to the core WorkspaceSession (state owner).
    @property
    def zones(self):
        return self._workspace.zones

    @zones.setter
    def zones(self, value):
        self._workspace.zones = value

    @property
    def zones_per_page(self):
        return self._workspace.zones_per_page

    @zones_per_page.setter
    def zones_per_page(self, value):
        self._workspace.zones_per_page = value

    def _build_ui(self):
        mode_f = tk.Frame(self, bg="#1a1a1a", pady=4, padx=4)
        mode_f.pack(fill="x", padx=20, pady=(10, 0))
        tk.Label(mode_f, text="OPERATING MODE:", bg="#1a1a1a", fg="#aaa", font=("Consolas", 9, "bold")).pack(side="left", padx=10)

        btn_style = dict(font=("Consolas", 10, "bold"), relief="flat", padx=16, pady=6, cursor="hand2", bd=0)

        self.btn_mode_seq = tk.Button(mode_f, text="🎯 Targeted Sequence (RPA)", command=lambda: self._set_mode("sequence"), **btn_style)
        self.btn_mode_seq.pack(side="left", padx=4)

        self.btn_mode_grid = tk.Button(mode_f, text="▦ Grid Scan (Fuzzer)", command=lambda: self._set_mode("grid"), **btn_style)
        self.btn_mode_grid.pack(side="left", padx=4)

        self.btn_mode_iscs = tk.Button(mode_f, text="🚨 Suite Runner", command=lambda: self._set_mode("iscs"), **btn_style)
        self.btn_mode_iscs.pack(side="left", padx=4)

        # Info Button cleanly integrated on the far right of the Operating Mode bar
        self.btn_info = tk.Button(mode_f, text="Info", bg="#222", fg="#ccc", relief="flat", padx=10, pady=6, cursor="hand2", command=lambda: self.help_panel.show_section(name="Whole App"))
        self.btn_info.pack(side="right", padx=(0, 8))

        self.sf = tk.Frame(self, bg="#0f0f0f")
        self.sf.pack(fill="x", padx=20, pady=(10, 0))
        self.screen_selector = ScreenSelectorPanel(self.sf, self.monitors, self._on_screen_selected)
        self.screen_selector.pack(side="left")
        self.btn_refresh = tk.Button(self.sf, text="↻ Refresh", bg="#222", fg="#ccc", font=("Consolas", 9, "bold"), cursor="hand2", relief="flat", padx=10, command=self._refresh_monitors)
        self.btn_refresh.pack(side="left", padx=15, pady=8)
        self.btn_ocr_monitor = tk.Button(self.sf, text="🔍 OCR", bg="#222", fg="#00BCD4",
                                         font=("Consolas", 9, "bold"), cursor="hand2",
                                         relief="flat", padx=10,
                                         command=self._open_ocr_monitor)
        self.btn_ocr_monitor.pack(side="left", padx=(0, 8), pady=8)
        Tooltip(self.btn_ocr_monitor, "Live OCR Monitor — draw a zone and see what Tesseract reads")

        tb = tk.Frame(self, bg="#161616", pady=4)
        tb.pack(fill="x", padx=20, pady=(4, 0))
        sf = dict(font=("Consolas", 10, "bold"), relief="flat", padx=11, pady=7, cursor="hand2")

        row1 = tk.Frame(tb, bg="#161616")
        row1.pack(fill="x")
        self.btn_overlay = tk.Button(row1, text="⊞ Draw Zones", bg="#222", fg="#ccc", command=self._open_overlay, **sf)
        self.btn_overlay.pack(side="left", padx=3)
        self.btn_clear_ws = tk.Button(row1, text="✕ Clear", bg="#222", fg="#555", command=self._clear_workspace, **sf, state="disabled")
        self.btn_clear_ws.pack(side="left", padx=3)
        self.btn_preview = tk.Button(row1, text="👁 Preview on Screen", bg="#222", fg="#aaa", command=self._toggle_preview, **sf, state="disabled")
        self.btn_preview.pack(side="left", padx=3)
        
        self.btn_load_excel = tk.Button(row1, text="📊 Load IO List", bg="#222", fg="#ccc", command=self._load_excel, **sf)
        self.btn_load_excel.pack(side="left", padx=3)
        self.btn_metadata = tk.Button(row1, text="📦 Profiles", bg="#222", fg="#ccc",
                                      command=self._open_metadata_browser, **sf)
        self.btn_metadata.pack(side="left", padx=3)
        Tooltip(self.btn_metadata, "Browse & reuse previously imported IO List profiles")

        self.btn_suite = tk.Button(row1, text="📋 Suite", bg="#222", fg="#ccc", command=self._toggle_suite, **sf)
        self.btn_suite.pack(side="left", padx=3)
        tk.Button(row1, text="⚙ Settings", bg="#222", fg="#ccc", command=self._settings_dialog, **sf).pack(side="left", padx=3)
        tk.Button(row1, text="💾 Save", bg="#222", fg="#ccc", command=self._save_zones, **sf).pack(side="left", padx=3)
        tk.Button(row1, text="📂 Load", bg="#222", fg="#ccc", command=self._load_zones, **sf).pack(side="left", padx=3)
        
        # M5: Run/Pause/Stop button cluster extracted to the RunControls view.
        self._tb_row2_full = tk.Frame(tb, bg="#161616")
        self._tb_row2_full.pack(fill="x", pady=(2, 2))
        self._tb_row2_compact = tk.Frame(tb, bg="#161616")
        self.run_controls = RunControls(
            self._tb_row2_full, self._tb_row2_compact,
            on_run=self._run_test, on_pause=self._toggle_pause, on_stop=self._stop_test,
            tooltip=Tooltip)

        self._tb_compact = False
        self.after(0, lambda: self.bind("<Configure>", self._on_resize))

        # M5: the stats strip is the StatsView (owns all stat labels).
        self.stats = StatsView(
            self, self._dispatcher,
            monitor_text=f"Display {self.active_mon.display_num}",
            spacing_text=f"{GRID_SPACING}px",
        ).pack(fill="x", padx=20)

        self._paned = tk.PanedWindow(self, orient="horizontal", bg="#0f0f0f", sashwidth=5, sashrelief="flat", sashpad=2, handlesize=0)
        self._paned.pack(fill="both", expand=True, padx=0, pady=0)

        left_pane = tk.Frame(self._paned, bg="#0f0f0f")
        self._paned.add(left_pane, stretch="always", minsize=400)

        pf = tk.Frame(left_pane, bg="#161616")
        pf.pack(fill="both", expand=True, padx=20, pady=8)
        self.preview_canvas = tk.Canvas(pf, bg="#111", highlightthickness=1, highlightbackground="#222")
        self.preview_canvas.pack(fill="both", expand=True, padx=8, pady=6)
        self.preview_canvas.bind("<Configure>", lambda e: self.after(50, self._draw_minimap))

        pf2 = tk.Frame(left_pane, bg="#0f0f0f")
        pf2.pack(fill="x", padx=20, pady=(0, 2))
        # M5: progress bar + status label extracted to the RunProgressView.
        self.run_progress = RunProgressView(pf2, self._dispatcher).pack(fill="x")

        lf = tk.Frame(left_pane, bg="#0f0f0f")
        lf.pack(fill="x", padx=20, pady=(0, 14))
        self._suite_pane = tk.Frame(self._paned, bg="#0f0f0f")   # unrelated; kept in App

        # M5: log rendering extracted to the LogSink driving-adapter view.
        self.log_sink = LogSink(lf, self._dispatcher).pack(fill="x")

        self._update_mode_buttons()

    def _notify_profile_listeners(self):
        """Fire all registered refresh callbacks (e.g. open MetadataBrowserDialogs)."""
        dead = []
        for cb in self._profile_update_listeners:
            try:
                cb()
            except Exception:
                dead.append(cb)
        for cb in dead:
            self._profile_update_listeners.remove(cb)

    def _open_metadata_browser(self):
        def on_load(profile_id):
            self._load_profile_from_metadata(profile_id)
        dlg = MetadataBrowserDialog(self, on_load_profile=on_load)
        # Register the dialog's refresh so it auto-updates when a new IO list is uploaded
        self._profile_update_listeners.append(dlg._refresh_list)
        # Override _close so the listener is unregistered when the dialog closes (any path)
        _listeners = self._profile_update_listeners
        _refresh_fn = dlg._refresh_list
        def _on_close():
            if _refresh_fn in _listeners:
                _listeners.remove(_refresh_fn)
            dlg.destroy()
        dlg._close = _on_close
        dlg.protocol("WM_DELETE_WINDOW", _on_close)

    def _load_profile_from_metadata(self, profile_id):
        prof, points = _metadata_load_profile(profile_id)
        if not prof or not points:
            messagebox.showerror("Metadata Error", "Could not load profile from metadata store.", parent=self)
            return
        self.iscs_excel_points = points
        self._log(f"ISCS: Loaded profile '{prof['name']}' — {len(points)} points (from metadata store).")
        Toast.show(self,
                   f"Profile loaded\n{prof['name']}\n{len(points)} points",
                   kind="info")
        # If a suite card config dialog is open and has no profile yet, sync it
        self._sync_open_card_config(prof, points)
        self._refresh()

    def _sync_open_card_config(self, prof, points):
        """If SuiteCardConfigDialog is open and no profile is selected, pre-fill it."""
        try:
            for widget in self.winfo_children():
                if isinstance(widget, SuiteCardConfigDialog):
                    if widget._profile_id is None:
                        widget._profile_id = prof["id"]
                        widget._profile_points = points
                        widget.lbl_profile.config(text=prof["name"])
                        widget.lbl_point_count.config(text=str(len(points)))
        except Exception:
            pass

    def _load_excel(self):
        if not PANDAS_AVAILABLE:
            messagebox.showerror("Error", "Please run: pip install pandas openpyxl", parent=self)
            return

        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xlsm *.xls"), ("All files", "*.*")], parent=self)
        if not path: return

        load_win = tk.Toplevel(self)
        load_win.title("Loading...")
        w, h = 300, 100
        x = self.winfo_x() + (self.winfo_width() // 2) - (w // 2)
        y = self.winfo_y() + (self.winfo_height() // 2) - (h // 2)
        load_win.geometry(f"{w}x{h}+{x}+{y}")
        load_win.configure(bg="#0f0f0f")
        load_win.overrideredirect(True)
        load_win.attributes("-topmost", True)
        tk.Label(load_win, text="Loading Excel file...\nPlease wait.", font=("Consolas", 11), bg="#0f0f0f", fg="#fff").pack(expand=True, fill="both")
        load_win.update()

        def parse_thread():
            try:
                wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
                sheets = wb.sheetnames
                self.after(0, lambda: self._excel_file_loaded(wb, sheets, load_win, path))
            except Exception as e:
                self.after(0, lambda: self._excel_load_failed(e, load_win))

        threading.Thread(target=parse_thread, daemon=True).start()

    def _excel_load_failed(self, err, load_win):
        load_win.destroy()
        messagebox.showerror("Excel Error", f"Failed to load file.\n{err}", parent=self)

    def _excel_file_loaded(self, wb, sheets, load_win, path):
        load_win.destroy()
        sheet_dlg = SheetSelectorDialog(self, sheets)
        self.wait_window(sheet_dlg)
        selected_sheet = sheet_dlg.result
        if not selected_sheet: return

        ws = wb[selected_sheet]
        row_idx, headers = detect_header_row(ws)
        if not headers:
            messagebox.showerror("Error", "Could not find a valid header row.", parent=self)
            return

        auto_map = auto_map_columns(headers)

        map_dlg = ColumnMapperDialog(self, headers, auto_map)
        self.wait_window(map_dlg)
        final_map = map_dlg.result
        if not final_map: return

        points = []

        all_headers_norm = [_normalize(h) for h in headers]
        state_cols = _find_state_table_cols(all_headers_norm)
        print("DEBUG norm_headers:", all_headers_norm)  # ADD THIS
        print("DEBUG state_cols:", state_cols)           # ADD THIS
        all_headers_norm = [_normalize(h) for h in headers]
        

        for row in ws.iter_rows(min_row=row_idx+1, values_only=True):
            if not any(row): continue

            def get_val(key, default=None, _row=row):
                col_idx = final_map.get(key)
                if col_idx is None or col_idx >= len(_row) or _row[col_idx] is None:
                    return default
                return _row[col_idx]

            pid = str(get_val("point_id", "")).strip()
            if not pid: continue

            states = _extract_states(row, state_cols)
            if not states:
                print(f"DEBUG states EMPTY for {pid} — check IO list columns")

            severity = get_val("severity", 0)
            try: severity = int(severity)
            except: severity = 0

            payload = {}
            for col_key in final_map.keys():
                val = get_val(col_key)
                if col_key == "fc":
                    if val is None: val = 3
                    s = str(val).strip().lower()
                    m = re.search(r'(\d+)', s)
                    val = int(m.group(1)) if m else 3
                elif col_key in ["reg", "bit", "device_address", "addr_size"]:
                    try: val = int(val)
                    except: val = 0
                if col_key not in ["point_id", "severity"]:   # protocol now included in payload
                    payload[col_key] = val

            raw_proto = str(get_val("protocol", "MODBUS")).strip().upper()

            point = {
                "point_id":       pid,
                "equipment_desc": str(get_val("equipment_desc", "")).strip(),
                "location":       str(get_val("location", "")).strip(),
                "attribute_desc": str(get_val("attribute_desc", "")).strip(),
                "station_code":   str(get_val("station_code", "")).strip(),
                "data_type":      str(get_val("data_type", "DI")).strip().upper(),
                "severity":       severity,
                "protocol":       raw_proto if raw_proto in ["MODBUS", "SNMP"] else "MODBUS",
                "states":         states,
                "payload":        payload,
                "alarm_list_desc": "",
            }
            eq = point["equipment_desc"]
            at = point["attribute_desc"]
            point["alarm_list_desc"] = f"{eq} : {at}" if eq and at else (eq or at)
            points.append(point)

        try:
            wb.close()
        except Exception:
            pass

        self.iscs_excel_points = points
        self._log(f"ISCS: Loaded {len(points)} points from '{selected_sheet}'.")

        # Update all suite cards with the new points so suite runner uses fresh IO list
        if self.suite_panel:
            for sc in self.suite_panel.scenarios:
                sc.iscs_points = points
            self.suite_panel._rebuild_cards()

        saved_ok, save_result = _metadata_save_profile(path, selected_sheet, final_map, points)

        if saved_ok:
            self._log(f"ISCS: Profile registered — {len(points)} points saved to DB.")
            Toast.show(self,
                       f"IO List loaded & registered\n"
                       f"{len(points)} points  ·  sheet: {selected_sheet}",
                       kind="success")
        else:
            self._log(f"ISCS: WARNING — Profile DB save failed: {save_result}")
            Toast.show(self,
                       f"IO List loaded ({len(points)} pts) but\n"
                       f"DB registration failed:\n{save_result}",
                       kind="error", duration_ms=6000)

        # Notify any open Profiles browser / MetadataBrowserDialog to refresh their list
        self._notify_profile_listeners()

        self._refresh()

    def _on_resize(self, event):
        if event.widget is not self: return
        if event.width < 860 and not self._tb_compact:
            self._tb_row2_full.pack_forget(); self._tb_row2_compact.pack(fill="x", pady=(2, 2)); self._tb_compact = True
        elif event.width >= 860 and self._tb_compact:
            self._tb_row2_compact.pack_forget(); self._tb_row2_full.pack(fill="x", pady=(2, 2)); self._tb_compact = False

    def _log(self, msg):
        # M5 shim → LogSink. DELETION TICKET: once log lines flow as a LogMessage
        # event the LogSink subscribes to, callers stop calling this and it's removed.
        self.log_sink.write(msg)

    def _refresh_monitors(self):
        self.monitors = detect_monitors()
        self.active_mon = next((m for m in self.monitors if m.display_num == self.active_mon.display_num), self.monitors[0])
        self.screen_selector.destroy()
        self.screen_selector = ScreenSelectorPanel(self.sf, self.monitors, self._on_screen_selected)
        self.screen_selector.pack(side="left", before=self.btn_refresh)
        self.screen_selector._highlight(self.monitors.index(self.active_mon))
        self._refresh()

    def _toggle_suite(self):
        if self._suite_pane_visible:
            self._paned.remove(self._suite_pane)
            self._suite_pane_visible = False
            self.btn_suite.config(bg="#222", fg="#ccc")
        else:
            if self.suite_panel is None:
                self.suite_panel = SuitePanel(self._suite_pane, self)
                
            self._paned.add(self._suite_pane, stretch="never", minsize=320, width=360)
            self._suite_pane_visible = True
            self.btn_suite.config(bg=POINT_COLOR, fg="#000")

    def _on_screen_selected(self, monitor: Monitor):
        IdentifyOverlay(self, monitor)
        prev = self.active_mon
        self.active_mon = monitor
        mon_color = SCREEN_COLORS[self.monitors.index(monitor) % len(SCREEN_COLORS)]
        self.stats.set("monitor", f"Display {monitor.display_num}", mon_color)
        self._close_preview()
        if self.zones and prev is not monitor:
            # Clear canvas zones only — zones_per_page (saved zones) preserved
            self.zones.clear()
        self._refresh()
        self._update_overlay_btn()

    def _find_monitor_by_info(self, monitor_info):
        return min(self.monitors, key=lambda m: abs(m.x - monitor_info["x"]) + abs(m.y - monitor_info["y"]))

    def _set_mode(self, mode: str):
        if self.run_mode.get() == mode: return
        self.run_mode.set(mode)
        self._update_mode_buttons()
        self._on_mode_change()
        self._update_overlay_btn()
        self._refresh()

    def _update_mode_buttons(self):
        if not hasattr(self, "btn_mode_seq"): return
        mode = self.run_mode.get()
        
        has_excel_btn = hasattr(self, "btn_load_excel")

        if mode == "sequence":
            self.btn_mode_seq.config(bg=TARGET_COLOR,  fg="#000")
            self.btn_mode_grid.config(bg="#2a2a2a",     fg="#666")
            self.btn_mode_iscs.config(bg="#2a2a2a",     fg="#666")
            if has_excel_btn: self.btn_load_excel.config(state="disabled")
            
        elif mode == "grid":
            self.btn_mode_seq.config(bg="#2a2a2a",      fg="#666")
            self.btn_mode_grid.config(bg=INCLUDE_COLOR, fg="#000")
            self.btn_mode_iscs.config(bg="#2a2a2a",     fg="#666")
            if has_excel_btn: self.btn_load_excel.config(state="disabled")
            
        elif mode == "iscs":
            self.btn_mode_seq.config(bg="#2a2a2a",      fg="#666")
            self.btn_mode_grid.config(bg="#2a2a2a",     fg="#666")
            self.btn_mode_iscs.config(bg=ALARM_PANEL_COLOR, fg="#fff")
            if has_excel_btn: self.btn_load_excel.config(state="normal")

    def _on_mode_change(self):
        if self.zones:
            if messagebox.askyesno("Change Mode", "Changing modes will clear current zones. Continue?"):
                self.zones.clear()
            else:
                self.run_mode.set("grid" if self.run_mode.get() == "sequence" else "sequence")
                self._update_mode_buttons()
                self._update_overlay_btn()
                return
        
        mode = self.run_mode.get()
        if mode == "grid":
            self.stats.set("include_title", "INCLUDE")
            self.stats.set("exclude_title", "EXCLUDE")
        elif mode == "sequence":
            self.stats.set("include_title", "TARGETS")
            self.stats.set("exclude_title", "-")
        elif mode == "iscs":
            self.stats.set("include_title", "PANELS")
            self.stats.set("exclude_title", "POINTS")
            
        self._refresh()

    def _open_overlay(self):
        self._close_preview()
        self.withdraw()
        time.sleep(0.15)
        # Get pages from the selected suite card config if available
        pages = []
        zpp = {}
        if self.suite_panel and getattr(self.suite_panel, '_selected_idx', None) is not None:
            idx = self.suite_panel._selected_idx
            sc = self.suite_panel.scenarios[idx]
            pages = sc.card_cfg.get("navigation", {}).get("pages", [])
            zpp = sc.zones_per_page
        OverlayWindow(self, self.run_mode.get(), self.zones, self.active_mon,
                      self.grid_spacing, self._overlay_done, pages=pages, zones_per_page=zpp)

    def _overlay_done(self, zones_per_page=None):
        self.deiconify()
        # Store zones_per_page on app so direct run (_run_test) can use it
        if zones_per_page is not None:
            self.zones_per_page = zones_per_page
        if self.suite_panel and getattr(self.suite_panel, '_selected_idx', None) is not None:
            idx = self.suite_panel._selected_idx
            selected_sc = self.suite_panel.scenarios[idx]
            selected_sc.zones = [Zone.from_dict(z.to_dict()) for z in self.zones]
            if zones_per_page is not None:
                selected_sc.zones_per_page = zones_per_page
            self.suite_panel._rebuild_cards()
        self._refresh()
        self._update_overlay_btn()

    def _clear_workspace(self):
        # M3.5: reset the live working set via the core WorkspaceSession.
        self._workspace.clear()

        # Deselect the card — card's own zones are NOT touched
        if self.suite_panel:
            self.suite_panel._selected_idx = None
            self.suite_panel._rebuild_cards()

        self._refresh()
        self._update_overlay_btn()

    def _update_overlay_btn(self):
        has_zones = bool(self.zones) or bool(self.zones_per_page)
        if has_zones:
            self.btn_overlay.config(text="✏ Edit Zones")
            self.btn_clear_ws.config(state="normal", fg="#FFE6E6")
        else:
            self.btn_overlay.config(text="⊞ Draw Zones")
            self.btn_clear_ws.config(state="disabled", fg="#555")  

    def _refresh_stats_only(self):
        self.valid_points, self.all_points = generate_points(self.run_mode.get(), self.active_mon, self.grid_spacing, self.zones)
        self._update_stats()
        has_pts = len(self.zones) > 0 if self.run_mode.get() == "iscs" else len(self.valid_points) > 0
        self.run_controls.set_run_enabled(has_pts)
        self.btn_preview.config(state="normal" if len(self.valid_points)>0 else "disabled")

    def _refresh(self):
        self.valid_points, self.all_points = generate_points(self.run_mode.get(), self.active_mon, self.grid_spacing, self.zones)
        self._update_stats()
        self._draw_minimap()
        
        if self.run_mode.get() == "iscs":
            has_zones = bool(self.zones) or bool(self.zones_per_page)
            has_pts = has_zones and (len(getattr(self, 'iscs_excel_points', [])) > 0)
        else:
            has_pts = len(self.valid_points) > 0

        self.run_controls.set_run_enabled(has_pts)
        self.btn_preview.config(state="normal" if len(self.valid_points)>0 else "disabled")

        if self.crosshair_overlay:
            self._close_preview()
            self._open_preview()

    def _update_stats(self):
        self.stats.set("zones", str(len(self.zones)))
        self.stats.set("points", str(len(self.valid_points)))
        self.stats.set("spacing", f"{self.grid_spacing}px")
        
        if self.run_mode.get() == "grid":
            inc = sum(1 for z in self.zones if z.zone_type == "include")
            exc = sum(1 for z in self.zones if z.zone_type == "exclude")
            self.stats.set("include", str(inc), INCLUDE_COLOR)
            self.stats.set("exclude", str(exc), EXCLUDE_COLOR)
        elif self.run_mode.get() == "sequence":
            tgt = sum(1 for z in self.zones if z.zone_type == "target")
            self.stats.set("include", str(tgt), TARGET_COLOR)
            self.stats.set("exclude", "-", "#444")
        elif self.run_mode.get() == "iscs":
            panels = sum(1 for z in self.zones if z.zone_type == "alarm_panel")
            pts = len(self.iscs_excel_points) if hasattr(self, 'iscs_excel_points') else 0
            self.stats.set("include", str(panels), ALARM_PANEL_COLOR)
            self.stats.set("exclude", str(pts), "#00C853")

    def _capture_monitor_thumbnail(self, monitor) -> "ImageTk.PhotoImage | None":
        if not PIL_AVAILABLE: return None
        try:
            bbox = (monitor.x, monitor.y, monitor.x + monitor.width, monitor.y + monitor.height)
            img = ImageGrab.grab(bbox=bbox, all_screens=True)
            self.preview_canvas.update_idletasks()
            cw = self.preview_canvas.winfo_width()
            ch = self.preview_canvas.winfo_height()
            if cw < 10 or ch < 10: return None
            ratio = min(cw / monitor.width, ch / monitor.height)
            new_w = int(monitor.width  * ratio)
            new_h = int(monitor.height * ratio)
            from PIL import Image as _Img
            img = img.resize((new_w, new_h), _Img.LANCZOS)
            return ImageTk.PhotoImage(img), new_w, new_h
        except Exception:
            return None

    def _draw_minimap(self, zones=None, monitor=None):
        c   = self.preview_canvas
        c.delete("all")
        c.update_idletasks()
        cw, ch = c.winfo_width(), c.winfo_height()
        if cw < 10 or ch < 10: return

        mon   = monitor or self.active_mon
        zones = zones   if zones is not None else self.zones

        result = self._capture_monitor_thumbnail(mon)
        offset_x = offset_y = 0
        if result:
            thumb, tw, th = result
            offset_x = (cw - tw) // 2
            offset_y = (ch - th) // 2
            c.create_image(offset_x, offset_y, anchor="nw", image=thumb)
            c._thumb_ref = thumb
            sx = tw / mon.width
            sy = th / mon.height
        else:
            c.create_rectangle(1, 1, cw-1, ch-1, outline="#1a1a1a", fill="#111")
            sx = cw / mon.width
            sy = ch / mon.height

        for zone in zones:
            zx1 = offset_x + (zone.x1 - mon.x) * sx
            zy1 = offset_y + (zone.y1 - mon.y) * sy
            zx2 = offset_x + (zone.x2 - mon.x) * sx
            zy2 = offset_y + (zone.y2 - mon.y) * sy
            
            # Map exact zone colors
            color = (TARGET_COLOR  if zone.zone_type == "target"  else
                     ALARM_PANEL_COLOR if zone.zone_type == "alarm_panel" else
                     EQUIP_ZONE_COLOR if zone.zone_type == "equipment_page" else
                     ALARM_LIST_COLOR if zone.zone_type == "alarm_list" else
                     EVENT_LIST_COLOR if zone.zone_type == "event_list" else
                     INCLUDE_COLOR if zone.zone_type == "include" else
                     EXCLUDE_COLOR)

            c.create_rectangle(zx1, zy1, zx2, zy2, fill=color, outline=color, stipple="gray25")
            c.create_rectangle(zx1, zy1, zx2, zy2, fill="", outline=color, width=1)

    def _toggle_preview(self):
        if self.crosshair_overlay: self._close_preview()
        else: self._open_preview()

    def _open_preview(self):
        if not self.valid_points: return
        self.crosshair_overlay = CrosshairOverlay(self, self.valid_points, self.all_points, self.active_mon, self.run_mode.get())
        self.btn_preview.config(text="✕ Close Preview", bg="#333", fg=POINT_COLOR)

    def _close_preview(self):
        if self.crosshair_overlay:
            self.crosshair_overlay.destroy(); self.crosshair_overlay = None
        self.btn_preview.config(text="👁 Preview on Screen", bg="#222", fg="#aaa")

    def _shake_window(self, win, count=6, dx=8):
        orig_x = win.winfo_x()
        orig_y = win.winfo_y()
        def _step(n, direction):
            if n <= 0:
                win.geometry(f"+{orig_x}+{orig_y}")
                return
            win.geometry(f"+{orig_x + direction * dx}+{orig_y}")
            win.after(30, _step, n - 1, -direction)
        _step(count, 1)
        
    def _settings_dialog(self):
        # M5: the Settings dialog is a SettingsView that reads/writes config via the
        # Core API. The App only supplies the side effects the view must not own.
        if getattr(self, "_settings_view", None) is None:
            self._settings_view = SettingsView(
                self, self.core_api, on_applied=self._on_settings_applied,
                shake=self._shake_window)
        self._settings_view.show()

    def _on_settings_applied(self, cfg):
        # Config is already written to APP_CONFIG + saved by the facade. Apply the
        # non-config side effects: legacy globals (ISCS_Engine still reads these until
        # M6.1), the grid-spacing attr, OCR re-init, and a canvas refresh.
        global CLICK_DELAY, MOUSE_DRIFT_PX
        self.grid_spacing = int(cfg.get("grid_spacing", self.grid_spacing))
        CLICK_DELAY = float(cfg.get("click_delay", CLICK_DELAY))
        MOUSE_DRIFT_PX = int(cfg.get("mouse_drift_px", MOUSE_DRIFT_PX))
        initialize_tesseract()
        if self.zones:
            self._refresh()

    def _save_zones(self):
        path = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("Zone config", "*.json")])
        if not path: return
        zpp_serial = {}
        for page, zt_dict in self.zones_per_page.items():
            zpp_serial[page] = {zt: z.to_dict() for zt, z in zt_dict.items() if z is not None}
        data = {
            "mode": self.run_mode.get(),
            "monitor": {"name": self.active_mon.name, "x": self.active_mon.x, "y": self.active_mon.y, "width": self.active_mon.width, "height": self.active_mon.height},
            "zones": [z.to_dict() for z in self.zones],
            "zones_per_page": zpp_serial,
        }
        with open(path, "w") as f: json.dump(data, f, indent=2)
        self._log(f"Saved → {path}")

    def _load_zones(self):
        path = filedialog.askopenfilename(filetypes=[("Zone config", "*.json")])
        if not path: return
        with open(path) as f: data = json.load(f)
        self.zones.clear()
        self.run_mode.set(data.get("mode", "grid"))
        self._update_mode_buttons()
        self.zones = [Zone.from_dict(d) for d in data.get("zones", [])]
        zpp = {}
        for page, zt_dict in data.get("zones_per_page", {}).items():
            zpp[page] = {zt: Zone.from_dict(zd) for zt, zd in zt_dict.items()}
        self.zones_per_page = zpp
        self._refresh()
        self._update_overlay_btn()
        self._log(f"Loaded ← {path}")
        
    def set_execution_state(self, state: str):
        if state == "idle":
            has_pts = False
            if self.run_mode.get() == "iscs":
                has_pts = (len(self.zones) > 0) and (len(getattr(self, 'iscs_excel_points', [])) > 0)
            else:
                has_pts = len(self.valid_points) > 0

            self.run_controls.set_run_state("idle", can_run=has_pts)
            self.btn_overlay.config(state="normal")
            self.btn_clear_ws.config(state="normal" if self.zones else "disabled")
            self.btn_preview.config(state="normal" if len(self.valid_points)>0 else "disabled")
            self.btn_refresh.config(state="normal")
            self.screen_selector.unlock()

        elif state == "running":
            self.run_controls.set_run_state("running")
            self.btn_overlay.config(state="disabled")
            self.btn_clear_ws.config(state="disabled")
            self.btn_preview.config(state="disabled")
            self.btn_refresh.config(state="disabled")
            self.screen_selector.lock()

        elif state == "paused":
            self.run_controls.set_run_state("paused")
            self.stats.set("state", "PAUSED", PAUSE_COLOR)

        elif state == "stopping":
            self.run_controls.set_run_state("stopping")
            self.stats.set("state", "STOPPING…", EXCLUDE_COLOR)

    # ── Test Execution ────────────────────────────────────────────────────────
    def _run_test(self):
        if self.click_engine and self.click_engine.is_alive(): return

        if self.run_mode.get() == "iscs":
            if not getattr(self, 'iscs_excel_points', []):
                messagebox.showerror("Error", "Load IO List first!", parent=self)
                return
            # Check alarm_panel zone exists either in flat zones or in zones_per_page
            zpp = getattr(self, 'zones_per_page', {})
            has_alarm_panel = (
                any(z.zone_type == "alarm_panel" for z in self.zones) or
                any("alarm_panel" in page_zones for page_zones in zpp.values())
            )
            if not has_alarm_panel:
                messagebox.showerror("Error", "Draw an Alarm Panel zone first!", parent=self)
                return
            if not TESSERACT_AVAILABLE:
                messagebox.showwarning("Missing Tesseract OCR", "Tesseract OCR was not found. Verification will fail. Check Settings.", parent=self)
                
            self._close_preview()
            self.set_execution_state("running")
            self.run_progress.set_fraction(0)
            self.stats.set("state", "RUNNING", ALARM_PANEL_COLOR)

            self.hud = HudOverlay(self, len(self.iscs_excel_points), self.active_mon)
            self._log(f"ISCS Test Started: {len(self.iscs_excel_points)} targets.")

            self.click_engine = ISCS_Engine(
                self.iscs_excel_points, self.zones, self.protocols, APP_CONFIG, LOG_DIR,
                self._cb_progress, self._cb_paused, self._cb_done, self._log,
                zones_per_page=zpp
            )
            self.click_engine.start()
            return

        if not self.valid_points: return
        if not ConfirmDialog.ask(self, self.run_mode.get(), len(self.valid_points), self.active_mon.label): return

        self._close_preview()
        self.set_execution_state("running") 
        self.run_progress.set_fraction(0)
        self.stats.set("state", "RUNNING", INCLUDE_COLOR)

        self.hud = HudOverlay(self, len(self.valid_points), self.active_mon)
        self._log(f"Test Started: {self.run_mode.get().upper()} MODE — {len(self.valid_points)} targets.")

        self.click_engine = ClickEngine(
            self.run_mode.get(), self.valid_points, LOG_DIR, CLICK_DELAY, self.active_mon,
            self._cb_progress, self._cb_paused, self._cb_done)
        self.click_engine.start()

    def _toggle_pause(self):
        if not self.click_engine or not self.click_engine.is_alive(): return
        
        if self.click_engine.is_paused:
            logger.info("User requested RESUME.")
            self.click_engine.resume()
            self.set_execution_state("running") 
            self.stats.set("state", "RUNNING", INCLUDE_COLOR)
            self._log("Resumed.")
        else:
            logger.info("User requested PAUSE.")
            self.click_engine.pause("manual")
            self.set_execution_state("paused") 
            self._log("Paused.")

    def _stop_test(self):
        if not self.click_engine or not self.click_engine.is_alive(): return
        
        logger.info("User requested STOP.")
        self.click_engine.stop()
        self.set_execution_state("stopping") 
        if self.hud and self.hud.winfo_exists():
            self.hud.update_stopped()
        self._log("Stop requested.")

    def _cb_progress(self, done, total, val1, val2):
        pct = done / total * 100
        self.run_progress.set_fraction(pct)
        
        if self.run_mode.get() == "iscs":
            msg = f"Testing {done}/{total}  →  {val1} ({val2})"
        else:
            msg = f"Clicking {done}/{total}  →  ({val1}, {val2})"
            
        self.run_progress.set_text(msg)
        
        if self.hud and self.hud.winfo_exists():
            if self.run_mode.get() == "iscs":
                tgt_text = f"Point: {val1}  [{val2[:4]}]"
            else:
                tgt_text = ""
                
            self.after(0, lambda: self.hud.update_running(done, total, tgt_text))

    def _cb_paused(self, idx, total, reason):
        self.after(0, lambda: self._on_auto_paused(idx, total, reason))

    def _cb_done(self, session_dir, log_path, was_stopped):
        self.after(0, lambda: self._test_finished(session_dir, log_path, was_stopped))

    def _on_auto_paused(self, idx, total, reason):
        self.set_execution_state("paused") 
        self._log(f"Auto-paused ({reason}). Space to resume.")
        if self.hud and self.hud.winfo_exists():
            self.hud.update_paused(idx, total, reason)

    def _test_finished(self, session_dir, log_path, was_stopped):
        if self.hud and self.hud.winfo_exists():
            self.hud.blink_and_destroy()
        
        self.set_execution_state("idle") 

        if not was_stopped: self.run_progress.set_fraction(100)
        self.stats.set("state", "STOPPED" if was_stopped else "DONE", EXCLUDE_COLOR if was_stopped else INCLUDE_COLOR)
        
        results_len = len(self.click_engine.results) if hasattr(self.click_engine, 'results') else 0
        end_str = f"{'Stopped' if was_stopped else 'Complete'} — {results_len} items."
        self.run_progress.set_text(end_str)
        self._log(f"{'Stopped' if was_stopped else 'Done'} → {session_dir}")

    def _open_ocr_monitor(self):
        # If already open, bring to front
        if hasattr(self, '_ocr_monitor_win') and self._ocr_monitor_win and \
                self._ocr_monitor_win.winfo_exists():
            self._ocr_monitor_win.lift()
            self._ocr_monitor_win.focus_force()
            return
        self._ocr_monitor_win = OcrMonitorPanel(self, self.active_mon)

    def destroy(self):
        self._unregister_hotkeys()
        self._close_preview()
        try:
            self.protocols.stop_all()
        except Exception:
            pass
        super().destroy()

if __name__ == "__main__":
    _load_plugins()        # discover ported capabilities (override legacy adapters by key)
    _wire_subscribers()    # event-driven report generation (P2.3)
    app = App()
    app.mainloop()
